from flask import Flask, render_template, request, redirect, url_for, session, jsonify,flash
from flask_sqlalchemy import SQLAlchemy
from flask_cors import CORS
import google.generativeai as genai
from dotenv import load_dotenv
from datetime import datetime, date, time
import qrcode
from io import BytesIO
import base64
from sqlalchemy import desc
from sqlalchemy import Date
import random
from datetime import datetime
from openpyxl import load_workbook
from werkzeug.utils import secure_filename 
import os
import re
import json
from datetime import datetime, timedelta


app = Flask(__name__)
CORS(app)
UPLOAD_FOLDER = 'static/uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = os.path.join(app.root_path, 'static', 'uploads')
app.config['SECRET_KEY'] = 'your_very_secret_key'
db = SQLAlchemy(app)
from datetime import time

class LessonStudent(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    lesson_id = db.Column(db.Integer, nullable=False)
    student_id = db.Column(db.Integer, nullable=False)

class LessonEmployee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    lesson_id = db.Column(db.Integer, nullable=False)
    student_id = db.Column(db.Integer, nullable=False)


class Workshop_Register(db.Model):
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    workshop_id = db.Column(db.Integer, nullable=False)
    user_id = db.Column(db.Integer, nullable=False)
    account_type = db.Column(db.String(50), nullable=False)
    name = db.Column(db.String(255), nullable=False)
    email = db.Column(db.String(255), nullable=False)
    country = db.Column(db.String(100), nullable=True)
    city = db.Column(db.String(100), nullable=True)
    accepted = db.Column(db.Boolean, nullable=False, default=False)

class Reward2(db.Model):
    id = db.Column(db.Integer, primary_key=True,  default=lambda: random.randint(100000, 9999999))
    name = db.Column(db.String(100), nullable=False)
    name1 = db.Column(db.String(100), nullable=True)
    name2 = db.Column(db.String(100), nullable=True)
    name3 = db.Column(db.String(100), nullable=True)
    name4 = db.Column(db.String(100), nullable=True)
    image = db.Column(db.String(200))
    start_date = db.Column(Date, nullable=False)
    end_date = db.Column(Date, nullable=False)
    start_time = db.Column(db.Time, nullable=False)  
    end_time = db.Column(db.Time, nullable=False)  
    total_gifts = db.Column(db.Integer, nullable=False)
    total_gifts1 = db.Column(db.Integer, nullable=True)
    total_gifts2 = db.Column(db.Integer, nullable=True)
    total_gifts3 = db.Column(db.Integer, nullable=True)
    total_gifts4 = db.Column(db.Integer, nullable=True)
    is_internal = db.Column(db.Boolean, nullable=False)
    exam_count = db.Column(db.Integer, nullable=False)
    percentage = db.Column(db.Float, nullable=False)
    account_type = db.Column(db.String(50), nullable=False)
    practice_count = db.Column(db.Integer, nullable=False, default=4)
    description = db.Column(db.String(255), nullable=True) 
class Work_shop4(db.Model):
    id = db.Column(db.Integer, primary_key=True,  default=lambda: random.randint(100000, 9999999))
    name = db.Column(db.String(100), nullable=False)
    image = db.Column(db.String(200))
    start_date = db.Column(Date, nullable=False)
    end_date = db.Column(Date, nullable=False)
    start_time = db.Column(db.Time, nullable=False)  
    end_time = db.Column(db.Time, nullable=False)  
    type = db.Column(db.String(255), nullable=False)
    days_count = db.Column(db.Integer, nullable=False)
    person_count = db.Column(db.Integer, nullable=False)
    location = db.Column(db.String(255), nullable=False)
    country = db.Column(db.String(255), nullable=True)
    city = db.Column(db.String(255), nullable=True)
    trainer = db.Column(db.String(255), nullable=True)
    description = db.Column(db.String(255), nullable=True)
    keywords = db.Column(db.Text, nullable=True) # الكلمات المفتاحية

class Notification_owner(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    subject = db.Column(db.String(100), nullable=False)
    message = db.Column(db.String(500), nullable=False)
    viewed = db.Column(db.Boolean, default=False)  
    company_code = db.Column(db.String(20), nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.now)

class Notification_organizer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    subject = db.Column(db.String(100), nullable=False)
    message = db.Column(db.String(500), nullable=False)
    viewed = db.Column(db.Boolean, default=False)  
    company_code = db.Column(db.String(20), nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.now)
    
class Notification_admin(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_user= db.Column(db.Integer,nullable=False)
    subject = db.Column(db.String(100), nullable=False)
    message = db.Column(db.String(500), nullable=False)
    viewed = db.Column(db.Boolean, default=False)  
    company_code = db.Column(db.String(20), nullable=False)
    timestamp = db.Column(db.DateTime, default=datetime.now)


class QuestionMultipleMultiple_choice(db.Model):
   
    question_id = db.Column(db.Integer, primary_key=True, default=lambda: random.randint(100000, 9999999))
    lesson_id = db.Column(db.Integer, nullable=False)
    question_name = db.Column(db.Text, nullable=True)
    audio_path = db.Column(db.String(255), nullable=True)
    photo_path = db.Column(db.String(255), nullable=True)
    difficulty = db.Column(db.String(50), default="Easy")
    importance = db.Column(db.String(50), default="Necessary")
    bloom_taxonomy = db.Column(db.String(50), default="Analyze")
    grade = db.Column(db.Integer, default=1)
    times = db.Column(db.Integer, default=7)
    type = db.Column(db.String(50), default="multiple_multiple_choice")

class Question_Multipleone_choice(db.Model):
    question_id = db.Column(db.Integer, primary_key=True,  default=lambda: random.randint(100000, 9999999))  
    lesson_id = db.Column(db.Integer, nullable=False)  
    question_name = db.Column(db.Text, nullable=True) 
    audio_path = db.Column(db.String(255), nullable=True) 
    photo_path = db.Column(db.String(255), nullable=True)  
    difficulty = db.Column(db.String(50), default="Easy")
    importance = db.Column(db.String(50), default="Necessary")  
    bloom_taxonomy = db.Column(db.String(50), default="Analyze") 
    grade = db.Column(db.Integer, default=1)  
    times = db.Column(db.Integer, default=7) 
    type = db.Column(db.String(50), default="multiple_choice")  
    
class Questionsequence1(db.Model):
    question_id = db.Column(db.Integer, primary_key=True,  default=lambda: random.randint(100000, 9999999))  
    lesson_id = db.Column(db.Integer, nullable=False)  
    question_name  = db.Column(db.Text, nullable=True) 
    audio_path = db.Column(db.String(255), nullable=True) 
    photo_path = db.Column(db.String(255), nullable=True)  
    difficulty = db.Column(db.String(50), default="Easy")
    importance = db.Column(db.String(50), default="Necessary")  
    bloom_taxonomy = db.Column(db.String(50), default="Analyze") 
    grade = db.Column(db.Integer, default=1)  
    times = db.Column(db.Integer, default=7)
    type = db.Column(db.String(50), default="sequence")
class Question_Fillin_blank4(db.Model):
    question_id = db.Column(db.Integer, primary_key=True,  default=lambda: random.randint(100000, 9999999))
    lesson_id = db.Column(db.Integer, nullable=False)
    question_text = db.Column(db.Text, nullable=True)
    question_text1 = db.Column(db.Text, nullable=True)
    question_text2 = db.Column(db.Text, nullable=True)
    question_text3 = db.Column(db.Text, nullable=True)
    question_text4 = db.Column(db.Text, nullable=True)
    question_text5 = db.Column(db.Text, nullable=True)
    question_text6 = db.Column(db.Text, nullable=True)
    difficulty = db.Column(db.String(50), default="Easy")
    importance = db.Column(db.String(50), default="Necessary")
    bloom_taxonomy = db.Column(db.String(50), default="Analyze")
    grade = db.Column(db.Integer, default=1)
    times = db.Column(db.Integer, default=7)
    answer_text = db.Column(db.Text, nullable=True)
    answer_text1 = db.Column(db.Text, nullable=True)
    answer_text2 = db.Column(db.Text, nullable=True)
    answer_text3 = db.Column(db.Text, nullable=True)
    answer_text4 = db.Column(db.Text, nullable=True)
    answer_text5 = db.Column(db.Text, nullable=True)
    answer_text6 = db.Column(db.Text, nullable=True)
    type = db.Column(db.String(20), default="fill_in_blank")

class Question_Fillin_blank_choice2(db.Model):
    question_id = db.Column(db.Integer, primary_key=True,  default=lambda: random.randint(100000, 9999999))  
    lesson_id = db.Column(db.Integer, nullable=False)  
    question_text = db.Column(db.Text, nullable=True)   
    question_text1 = db.Column(db.Text, nullable=True) 
    question_text2 = db.Column(db.Text, nullable=True) 
    question_text3 = db.Column(db.Text, nullable=True) 
    question_text4 = db.Column(db.Text, nullable=True) 
    question_text5 = db.Column(db.Text, nullable=True) 
    question_text6 = db.Column(db.Text, nullable=True) 
    difficulty = db.Column(db.String(50), default="Easy")
    importance = db.Column(db.String(50), default="Necessary")  
    bloom_taxonomy = db.Column(db.String(50), default="Analyze") 
    grade = db.Column(db.Integer, default=1)  
    times = db.Column(db.Integer, default=7) 
    answer_text = db.Column(db.Text, nullable=True) 
    answer_text1 = db.Column(db.Text, nullable=True) 
    answer_text2 = db.Column(db.Text, nullable=True) 
    answer_text3 = db.Column(db.Text, nullable=True) 
    answer_text4 = db.Column(db.Text, nullable=True) 
    answer_text5 = db.Column(db.Text, nullable=True) 
    answer_text6 = db.Column(db.Text, nullable=True)
    type = db.Column(db.String(20), default="fill_in_blank_choice")

class Article_question(db.Model):
    question_id = db.Column(db.Integer, primary_key=True,  default=lambda: random.randint(100000, 9999999))  
    lesson_id = db.Column(db.Integer, nullable=False)  
    question_text = db.Column(db.Text, nullable=True)   
    audio_path = db.Column(db.String(255), nullable=True) 
    photo_path = db.Column(db.String(255), nullable=True) 
    difficulty = db.Column(db.String(50), default="Easy")
    importance = db.Column(db.String(50), default="Necessary")  
    bloom_taxonomy = db.Column(db.String(50), default="Analyze") 
    grade = db.Column(db.Integer, default=1)  
    times = db.Column(db.Integer, default=7) 


class Answer_Multiple_Multiple_choice(db.Model):
    answer_id = db.Column(db.Integer, primary_key=True,  default=lambda: random.randint(100000, 9999999))
    question_id = db.Column(db.Integer, nullable=False)  
    answer_text = db.Column(db.Text, nullable=True) 
    audio_path = db.Column(db.String(255), nullable=True)  
    photo_path = db.Column(db.String(255), nullable=True)  
    is_correct = db.Column(db.Boolean, default=False)  

class Answer_sequence(db.Model):
    answer_id = db.Column(db.Integer, primary_key=True,  default=lambda: random.randint(100000, 9999999))  
    question_id = db.Column(db.Integer, nullable=False)  
    answer_text = db.Column(db.Text, nullable=True)  
    index = db.Column(db.Integer, default=False) 

class Answer_matching(db.Model):
    answer_id = db.Column(db.Integer, primary_key=True,  default=lambda: random.randint(100000, 9999999))  
    question_id = db.Column(db.Integer, nullable=False)  
    answer_text = db.Column(db.Text, nullable=True)  
    index = db.Column(db.Integer, default=False) 

class questionmatching1(db.Model):
    question_id = db.Column(db.Integer, primary_key=True,  default=lambda: random.randint(100000, 9999999))  
    lesson_id = db.Column(db.Integer, nullable=False)  
    question_text = db.Column(db.Text, nullable=True)   
    audio_path = db.Column(db.String(255), nullable=True) 
    photo_path = db.Column(db.String(255), nullable=True) 
    question_text1 = db.Column(db.Text, nullable=True)   
    audio_path1 = db.Column(db.String(255), nullable=True) 
    photo_path1 = db.Column(db.String(255), nullable=True) 
    question_text2 = db.Column(db.Text, nullable=True)   
    audio_path2 = db.Column(db.String(255), nullable=True) 
    photo_path2 = db.Column(db.String(255), nullable=True) 
    question_text3 = db.Column(db.Text, nullable=True)   
    audio_path3 = db.Column(db.String(255), nullable=True) 
    photo_path3 = db.Column(db.String(255), nullable=True)
    answer_text = db.Column(db.Text, nullable=True)
    answer_audio_path = db.Column(db.String(255), nullable=True)  
    answer_photo_path = db.Column(db.String(255), nullable=True) 
    answer_text1 = db.Column(db.Text, nullable=True)
    answer_audio_path1 = db.Column(db.String(255), nullable=True) 
    answer_photo_path1 = db.Column(db.String(255), nullable=True) 
    answer_text2 = db.Column(db.Text, nullable=True)  
    answer_audio_path2 = db.Column(db.String(255), nullable=True) 
    answer_photo_path2 = db.Column(db.String(255), nullable=True) 
    answer_text3 = db.Column(db.Text, nullable=True)  
    answer_audio_path3 = db.Column(db.String(255), nullable=True)
    answer_photo_path3 = db.Column(db.String(255), nullable=True)
    difficulty = db.Column(db.String(50), default="Easy")
    importance = db.Column(db.String(50), default="Necessary")  
    bloom_taxonomy = db.Column(db.String(50), default="Analyze") 
    grade = db.Column(db.Integer, default=1)  
    times = db.Column(db.Integer, default=7)

class questionclassiffication(db.Model):
    question_id = db.Column(db.Integer, primary_key=True,  default=lambda: random.randint(100000, 9999999))  
    lesson_id = db.Column(db.Integer, nullable=False)  
    label_text = db.Column(db.Text, nullable=True)   
    label_text1 = db.Column(db.Text, nullable=True)   
    label_text2 = db.Column(db.Text, nullable=True)   
    label_text3 = db.Column(db.Text, nullable=True) 
    difficulty = db.Column(db.String(50), default="Easy")
    importance = db.Column(db.String(50), default="Necessary")  
    bloom_taxonomy = db.Column(db.String(50), default="Analyze") 
    grade = db.Column(db.Integer, default=1)  
    times = db.Column(db.Integer, default=7) 
    type=db.Column(db.String(50), default="classification")

class Answer_classiffication1(db.Model):
    answer_id = db.Column(db.Integer, primary_key=True,  default=lambda: random.randint(100000, 9999999))  
    question_id = db.Column(db.Integer, nullable=False)  
    answer_text = db.Column(db.Text, nullable=True)  
    label = db.Column(db.Integer, nullable=True) 

class WordPuzzleGame2(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    lesson_id = db.Column(db.Integer, nullable=False)
    user_id = db.Column(db.Integer, nullable=False)
    company_code = db.Column(db.String(50), nullable=False)
    difficulty = db.Column(db.String(50), default="Easy")
    importance = db.Column(db.String(50), default="Necessary")
    bloom_taxonomy = db.Column(db.String(50), default="Analyze")
    grade = db.Column(db.Integer, default=1)
    times = db.Column(db.Integer, default=7)
    words = db.Column(db.Text, nullable=False)
    image_path = db.Column(db.String(255), nullable=True)  # لتخزين مسار الصورة
    audio_path = db.Column(db.String(255), nullable=True)  # لتخزين مسار الصوت

class MatchGame(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    lesson_id = db.Column(db.Integer, nullable=False)
    user_id = db.Column(db.Integer, nullable=False)
    company_code = db.Column(db.String(50), nullable=False)
    difficulty = db.Column(db.String(50), default="Easy")
    importance = db.Column(db.String(50), default="Necessary")
    bloom_taxonomy = db.Column(db.String(50), default="Analyze")
    grade = db.Column(db.Integer, default=1)
    times = db.Column(db.Integer, default=7)
    words = db.Column(db.Text, nullable=False)

class ReverseGame1(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    lesson_id = db.Column(db.Integer, nullable=False)
    user_id = db.Column(db.Integer, nullable=False)
    company_code = db.Column(db.String(50), nullable=False)
    difficulty = db.Column(db.String(50), default="Easy")
    importance = db.Column(db.String(50), default="Necessary")
    bloom_taxonomy = db.Column(db.String(50), default="Analyze")
    grade = db.Column(db.Integer, default=1)
    times = db.Column(db.Integer, default=7)
    words = db.Column(db.Text, nullable=False)
    image_path = db.Column(db.String(255), nullable=True)  # لتخزين مسار الصورة
    audio_path = db.Column(db.String(255), nullable=True)  # لتخزين مسار الصوت

class Quessgame(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    lesson_id = db.Column(db.Integer, nullable=False)
    user_id = db.Column(db.Integer, nullable=False)
    company_code = db.Column(db.String(50), nullable=False)
    difficulty = db.Column(db.String(50), default="Easy")
    importance = db.Column(db.String(50), default="Necessary")
    bloom_taxonomy = db.Column(db.String(50), default="Analyze")
    grade = db.Column(db.Integer, default=1)
    times = db.Column(db.Integer, default=7)
    words = db.Column(db.String(255), nullable=False)
    translate = db.Column(db.String(255), nullable=False)


class TraslateArabic(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    lesson_id = db.Column(db.Integer, nullable=False)
    user_id = db.Column(db.Integer, nullable=False)
    company_code = db.Column(db.String(50), nullable=False)
    difficulty = db.Column(db.String(50), default="Easy")
    importance = db.Column(db.String(50), default="Necessary")
    bloom_taxonomy = db.Column(db.String(50), default="Analyze")
    grade = db.Column(db.Integer, default=1)
    times = db.Column(db.Integer, default=7)
    words = db.Column(db.String(255), nullable=False)
    translate = db.Column(db.String(255), nullable=False)

class Traslateicon(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    lesson_id = db.Column(db.Integer, nullable=False)
    user_id = db.Column(db.Integer, nullable=False)
    company_code = db.Column(db.String(50), nullable=False)
    difficulty = db.Column(db.String(50), default="Easy")
    importance = db.Column(db.String(50), default="Necessary")
    bloom_taxonomy = db.Column(db.String(50), default="Analyze")
    grade = db.Column(db.Integer, default=1)
    times = db.Column(db.Integer, default=7)
    words = db.Column(db.String(255), nullable=False)
    translate = db.Column(db.String(255), nullable=False)

class Chapter(db.Model):
    id = db.Column(db.Integer, primary_key=True, default=lambda: random.randint(100000, 9999999))
    id_chapter=db.Column(db.Integer, nullable=False)
    title = db.Column(db.String(120), nullable=False)  
    image_path = db.Column(db.String(200), nullable=True) 
    created_at = db.Column(db.DateTime, default=datetime.now)  
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    employee_id = db.Column(db.Integer, nullable=False)  

class Lesson8(db.Model):
    id = db.Column(db.Integer, primary_key=True, default=lambda: random.randint(100000, 9999999))
    id_chapter = db.Column(db.Integer, nullable=False)
    title = db.Column(db.String(120), nullable=False)  
    image_path = db.Column(db.String(200), nullable=True)  
    file_path = db.Column(db.String(200), nullable=True)  
    file_path1 = db.Column(db.String(200), nullable=True)
    file_path2 = db.Column(db.String(200), nullable=True)
    file_path3 = db.Column(db.String(200), nullable=True)
    file_path4 = db.Column(db.String(200), nullable=True)
    audio_path = db.Column(db.String(200), nullable=True)  # مسار الصوت
    video_path = db.Column(db.String(200), nullable=True)  # مسار الفيديو
    description = db.Column(db.Text, nullable=True)  # وصف الدرس
    link = db.Column(db.String(300), nullable=True)  # رابط الدرس
    created_at = db.Column(db.DateTime, default=datetime.now)  
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    employee_id = db.Column(db.Integer, nullable=False)

class Lesson_custom1(db.Model):
    id = db.Column(db.Integer, primary_key=True, default=lambda: random.randint(100000, 9999999))
    title = db.Column(db.String(120), nullable=False)  
    image_path = db.Column(db.String(200), nullable=True)  
    file_path = db.Column(db.String(200), nullable=True)  
    file_path1 = db.Column(db.String(200), nullable=True)
    file_path2 = db.Column(db.String(200), nullable=True)
    file_path3 = db.Column(db.String(200), nullable=True)
    file_path4 = db.Column(db.String(200), nullable=True)
    audio_path = db.Column(db.String(200), nullable=True)  # مسار الصوت
    video_path = db.Column(db.String(200), nullable=True)  # مسار الفيديو
    description = db.Column(db.Text, nullable=True)  # وصف الدرس
    link = db.Column(db.String(300), nullable=True)  # رابط الدرس
    created_at = db.Column(db.DateTime, default=datetime.now)  
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    employee_id = db.Column(db.Integer, nullable=False)

class lessons_workshop(db.Model):
    id = db.Column(db.Integer, primary_key=True, default=lambda: random.randint(100000, 9999999))
    workshop_id = db.Column(db.Integer, nullable=False)
    title = db.Column(db.String(120), nullable=False)  
    image_path = db.Column(db.String(200), nullable=True)  
    file_path = db.Column(db.String(200), nullable=True)  # مسار الملف
    audio_path = db.Column(db.String(200), nullable=True)  # مسار الصوت
    video_path = db.Column(db.String(200), nullable=True)  # مسار الفيديو
    description = db.Column(db.Text, nullable=True)  # وصف الدرس
    link = db.Column(db.String(300), nullable=True)  # رابط الدرس
    created_at = db.Column(db.DateTime, default=datetime.now)  
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    employee_id = db.Column(db.Integer, nullable=False)

class Exam(db.Model):
    exam_id = db.Column(db.Integer, primary_key=True)
    lesson_id = db.Column(db.Integer, nullable=False)
    exam_name = db.Column(db.String(100), nullable=False)
    description = db.Column(db.String(255), nullable=False)
    exam_type = db.Column(db.String(50), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow) 

class category(db.Model):
    id = db.Column(db.Integer, primary_key=True, default=lambda: random.randint(100000, 9999999))
    title = db.Column(db.String(120), nullable=False)  
    image_path = db.Column(db.String(200), nullable=True)  
    created_at = db.Column(db.DateTime, default=datetime.now)  
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now) 
    employee_id = db.Column(db.Integer, nullable=False)  

class UserAnswer4(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, nullable=False)  # ID المستخدم
    question_id = db.Column(db.Integer, nullable=False)  # ID السؤال
    lesson_id = db.Column(db.Integer, nullable=False)  # ID الدرس
    is_correct = db.Column(db.Boolean, default=False)  # هل الإجابة صحيحة؟
    time_taken = db.Column(db.Float, nullable=False)  # الوقت المستغرق لحل السؤال
    grade = db.Column(db.Integer, nullable=False)  
    grade_real=db.Column(db.Integer, nullable=False) 
    time_allowed = db.Column(db.Float, nullable=False) 
    type = db.Column(db.String(200), nullable=True) 
    created_at = db.Column(db.DateTime, default=datetime.now)
    exams_taken = db.Column(db.Integer, default=0)  # إضافة هذا العمود

class UserExam1(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    lesson_id = db.Column(db.Integer, nullable=False)
    exams_taken = db.Column(db.Integer, default=0)
    last_exam_date = db.Column(db.DateTime, default=datetime.now)

class ManagerEmployee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin=db.Column(db.Integer,nullable=False)
    code = db.Column(db.String(10), nullable=False)
    username = db.Column(db.String(50), nullable=False)
    email = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    birthdate = db.Column(db.Date, nullable=True)
    stage = db.Column(db.String(50), nullable=True)
    university = db.Column(db.String(100), nullable=True)
    major = db.Column(db.String(100), nullable=True)
    phone = db.Column(db.String(15), nullable=True)
    country = db.Column(db.String(100), nullable=True)
    province = db.Column(db.String(100), nullable=True)
    col1 = db.Column(db.String(100), nullable=True)
    col2 = db.Column(db.String(100), nullable=True)
    col3 = db.Column(db.String(100), nullable=True)
    col4 = db.Column(db.String(100), nullable=True)
    col5 = db.Column(db.String(100), nullable=True)
    branch = db.Column(db.String(100), nullable=True) 
    manager_id = db.Column(db.Integer, nullable=True)
    manager_name = db.Column(db.String(50), nullable=True)

class prvatebank(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    lesson_id = db.Column(db.Integer, nullable=False)
    student_id = db.Column(db.Integer, nullable=False)

class head(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin=db.Column(db.Integer,nullable=False)
    code = db.Column(db.String(10), nullable=False)
    username = db.Column(db.String(50), nullable=False)
    email = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    birthdate = db.Column(db.Date, nullable=True)
    stage = db.Column(db.String(50), nullable=True)
    university = db.Column(db.String(100), nullable=True)
    major = db.Column(db.String(100), nullable=True)
    phone = db.Column(db.String(15), nullable=True)
    country = db.Column(db.String(100), nullable=True)
    province = db.Column(db.String(100), nullable=True)
    col1 = db.Column(db.String(100), nullable=True)
    col2 = db.Column(db.String(100), nullable=True)
    col3 = db.Column(db.String(100), nullable=True)
    col4 = db.Column(db.String(100), nullable=True)
    col5 = db.Column(db.String(100), nullable=True)
    branch = db.Column(db.String(100), nullable=True) 
class SubmittionLesson(db.Model):
    id = db.Column(db.Integer, primary_key=True, default=lambda: random.randint(100000, 9999999))
    lesson_id = db.Column(db.Integer, nullable=False)
    employee_id = db.Column(db.Integer, nullable=False)
    role = db.Column(db.String(50), nullable=True)
class Employee6(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin=db.Column(db.Integer,nullable=False)
    code = db.Column(db.String(10), nullable=False)
    username = db.Column(db.String(50), nullable=False)
    email = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    birthdate = db.Column(db.Date, nullable=True)
    stage = db.Column(db.String(50), nullable=True)
    university = db.Column(db.String(100), nullable=True)
    major = db.Column(db.String(100), nullable=True)
    phone = db.Column(db.String(15), nullable=True)
    country = db.Column(db.String(100), nullable=True)
    province = db.Column(db.String(100), nullable=True)
    col1 = db.Column(db.String(100), nullable=True)
    col2 = db.Column(db.String(100), nullable=True)
    col3 = db.Column(db.String(100), nullable=True)
    col4 = db.Column(db.String(100), nullable=True)
    col5 = db.Column(db.String(100), nullable=True)
    branch = db.Column(db.String(100), nullable=True) 
    manager_id = db.Column(db.Integer, db.ForeignKey('employee.id'), nullable=True)
    manager_name = db.Column(db.String(50), nullable=True)
    is_manager = db.Column(db.Boolean, default=False)
    is_head = db.Column(db.String(100), default="fff")
    permission_add_exam = db.Column(db.Boolean, default=False) 
    permission_add_permission = db.Column(db.Boolean, default=False) 

class organizer2(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), nullable=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(50), unique=True, nullable=False)  
    col1 = db.Column(db.String(50), nullable=True)
    col2 = db.Column(db.String(50), nullable=True)
    col3 = db.Column(db.String(50), nullable=True)
    col4 = db.Column(db.String(50), nullable=True)
    col5 = db.Column(db.String(50), nullable=True)
    password = db.Column(db.String(100), nullable=False)

class  administrator2(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(50), unique=True, nullable=False)  
    col1 = db.Column(db.String(50), nullable=True)
    col2 = db.Column(db.String(50), nullable=True)
    col3 = db.Column(db.String(50), nullable=True)
    col4 = db.Column(db.String(50), nullable=True)
    col5 = db.Column(db.String(50), nullable=True)
    code = db.Column(db.String(20), nullable=True)
    password = db.Column(db.String(100), nullable=False)
    
class Employee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    job_number = db.Column(db.String(20), unique=True, nullable=False)
    job_title = db.Column(db.String(50), nullable=False)
    specialization = db.Column(db.String(50), nullable=False)
    company_name = db.Column(db.String(50), nullable=False)
    department = db.Column(db.String(50), nullable=False)
    admin_title = db.Column(db.String(50), nullable=False)
    manager_id = db.Column(db.Integer, db.ForeignKey('employee.id'), nullable=True)

class ColAdministrator(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), unique=True, nullable=False)
    c1 = db.Column(db.String(50), nullable=False)
    c2 = db.Column(db.String(50), nullable=False)
    c3 = db.Column(db.String(50), nullable=False)
    c4 = db.Column(db.String(50), nullable=False)
    c5 = db.Column(db.String(50), nullable=False)

class ColOrganizer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), unique=True, nullable=False)
    c1 = db.Column(db.String(50), nullable=False)
    c2 = db.Column(db.String(50), nullable=False)
    c3 = db.Column(db.String(50), nullable=False)
    c4 = db.Column(db.String(50), nullable=False)
    c5 = db.Column(db.String(50), nullable=False)

class ColStudent(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), nullable=False)
    id_admin = db.Column(db.Integer, nullable=False)
    c1 = db.Column(db.String(50), nullable=True)
    c2 = db.Column(db.String(50), nullable=True)
    c3 = db.Column(db.String(50), nullable=True)
    c4 = db.Column(db.String(50), nullable=True)
    c5 = db.Column(db.String(50), nullable=True)

class ColEmployee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin = db.Column(db.Integer, nullable=False)
    code = db.Column(db.String(20), nullable=False)
    c1 = db.Column(db.String(50), nullable=True)
    c2 = db.Column(db.String(50), nullable=True)
    c3 = db.Column(db.String(50), nullable=True)
    c4 = db.Column(db.String(50), nullable=True)
    c5 = db.Column(db.String(50), nullable=True)

class Manager(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    job_number = db.Column(db.String(20), unique=True, nullable=False)
    job_title = db.Column(db.String(50), nullable=False)
    specialization = db.Column(db.String(50), nullable=False)
    company_name = db.Column(db.String(50), nullable=False)
    department = db.Column(db.String(50), nullable=False)
    admin_title = db.Column(db.String(50), nullable=False)
    manager_id = db.Column(db.Integer, db.ForeignKey('manager.id'), nullable=True)

class CompanyName(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    code=db.Column(db.String(8), nullable=False)

class specializationEmployee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin=db.Column(db.Integer,nullable=False)
    name = db.Column(db.String(120), nullable=False)
    code=db.Column(db.String(8), nullable=False)    

class Administrative_Title(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin=db.Column(db.Integer,nullable=False)
    name = db.Column(db.String(120), nullable=False)
    code=db.Column(db.String(8), nullable=False) 

class JobTitle(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin=db.Column(db.Integer,nullable=False)
    name = db.Column(db.String(120), nullable=False)
    code=db.Column(db.String(8), nullable=False)

class GradeStudent(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin = db.Column(db.Integer, nullable=False)
    name = db.Column(db.String(120), nullable=False)
    code=db.Column(db.String(8), nullable=False)

class DepartmentEmployee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin=db.Column(db.Integer,nullable=False)
    name = db.Column(db.String(120), nullable=False)
    code=db.Column(db.String(8), nullable=False)

class DepartmentEmployee1(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin=db.Column(db.Integer,nullable=False)
    name = db.Column(db.String(120), nullable=False)
    code=db.Column(db.String(8), nullable=False)

class DepartmentStudent(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin = db.Column(db.Integer, nullable=False)
    name = db.Column(db.String(120), nullable=False)
    code=db.Column(db.String(8), nullable=False)

class ClassStudent(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin = db.Column(db.Integer, nullable=False)
    name = db.Column(db.String(120), nullable=False)
    code=db.Column(db.String(8), nullable=False)

class yearStudent(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin = db.Column(db.Integer, nullable=False)
    name = db.Column(db.String(120), nullable=False)
    code=db.Column(db.String(8), nullable=False)

class specializationStudent(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin = db.Column(db.Integer, nullable=False)
    name = db.Column(db.String(120), nullable=False)
    code=db.Column(db.String(8), nullable=False)

class ColumnPreference(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), nullable=False)
    column_name = db.Column(db.String(100), nullable=False)
    code=db.Column(db.String(8), nullable=False)
    visible = db.Column(db.Boolean, default=True)

class Personal(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password = db.Column(db.String(100), nullable=False)
    account_type = db.Column(db.String(50), nullable=False)

class Sector(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    foundation_name = db.Column(db.String(100), nullable=False)
    username = db.Column(db.String(50), nullable=False, unique=False)
    email = db.Column(db.String(100), nullable=False, unique=True)
    code = db.Column(db.String(8), nullable=False, unique=True)
    password = db.Column(db.String(200), nullable=False)
    address = db.Column(db.String(200), nullable=True)
    phone_number = db.Column(db.String(20), nullable=True)
    type = db.Column(db.String(50), nullable=True)
    state = db.Column(db.String(100), nullable=True)
    country = db.Column(db.String(100), nullable=True)
    description = db.Column(db.Text, nullable=True)
    image_path = db.Column(db.String(200), nullable=True)

class Sectorimage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(8), nullable=False, unique=True)
    image_path = db.Column(db.String(200), nullable=True)


class SectorStudent8(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_admin = db.Column(db.Integer, nullable=False)
    username = db.Column(db.String(150), nullable=False)
    email = db.Column(db.String(150), nullable=False)
    code = db.Column(db.String(8), nullable=False)
    password = db.Column(db.String(200), nullable=False)
    accepted = db.Column(db.Boolean, nullable=False, default=False)
    birthdate = db.Column(db.Date, nullable=True)
    stage = db.Column(db.String(100), nullable=True)
    university = db.Column(db.String(150), nullable=True)
    major = db.Column(db.String(150), nullable=True)
    phone = db.Column(db.String(20), nullable=True)
    country = db.Column(db.String(100), nullable=True)
    province = db.Column(db.String(100), nullable=True)
    col1 = db.Column(db.String(100), nullable=True)
    col2 = db.Column(db.String(100), nullable=True)
    col3 = db.Column(db.String(100), nullable=True)
    col4 = db.Column(db.String(100), nullable=True)
    col5 = db.Column(db.String(100), nullable=True)
    
class mangers(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    foundation_name = db.Column(db.String(100), nullable=False)
    username = db.Column(db.String(50), nullable=False, unique=False)
    email = db.Column(db.String(100), nullable=False, unique=True)
    code = db.Column(db.String(8), nullable=False, unique=True)
    password = db.Column(db.String(200), nullable=False)
    address = db.Column(db.String(200), nullable=True)
    phone_number = db.Column(db.String(20), nullable=True)
    type = db.Column(db.String(50), nullable=True)
    state = db.Column(db.String(100), nullable=True)
    country = db.Column(db.String(100), nullable=True)
    description = db.Column(db.Text, nullable=True)

class Education(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    foundation_name = db.Column(db.String(100), nullable=False)
    username = db.Column(db.String(50), nullable=False, unique=False)
    email = db.Column(db.String(100), nullable=False, unique=True)
    password = db.Column(db.String(200), nullable=False)

class Questionfort_or_f(db.Model):
    question_id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    lesson_id = db.Column(db.Integer,  nullable=False)
    audio_path = db.Column(db.String(255), default="")
    photo_path = db.Column(db.String(255), default="")
    question_name = db.Column(db.String(255), default="السؤال الأول")
    difficulty = db.Column(db.String(50), default="Easy")
    importance = db.Column(db.String(50), default="Necessary")
    bloom_taxonomy = db.Column(db.String(50), default="Analyze")
    grade = db.Column(db.Integer, default=1)
    times = db.Column(db.Integer, default=7)
    correct_answer = db.Column(db.Boolean, default=True)
    type = db.Column(db.String(50), default="true_false")

with app.app_context():
    #db.drop_all()
    db.create_all()


@app.route('/parent_grades')
def parent_grades():
    user_id = request.args.get('user_id')  # احصل على user_id من الـ query string
    if not user_id:
        return "User ID is required in the query string."

    # استعلام لاسترجاع جميع إجابات المستخدم، مرتبة حسب lesson_id (الامتحان)
    user_answers = UserAnswer4.query.filter_by(user_id=user_id).order_by(UserAnswer4.lesson_id).all()

    # تجميع الدرجات لكل امتحان
    exam_grades = {}
    for answer in user_answers:
        if answer.lesson_id not in exam_grades:
            user_exam = UserExam1.query.filter_by(lesson_id=answer.lesson_id).first()
            exams_taken = user_exam.exams_taken if user_exam else 0 
            exam_grades[answer.lesson_id] = {'total_grade': 0,'exams_taken': exams_taken ,'max_total_grade': 0, 'exam_name': f'Exam {answer.lesson_id}'}  # يمكنك استبدال exam_name باسم الامتحان الفعلي إذا كان لديك
        exam_grades[answer.lesson_id]['total_grade'] += answer.grade
        
        exam_grades[answer.lesson_id]['max_total_grade'] += answer.grade_real

    return render_template('parent_grades.html', exam_grades=exam_grades, user_id=user_id,exams_taken=exams_taken)

company_nour = """
"أنت روبوت دردشة تم تدريبه خصيصًا على فهم نظام عمل موقع شركة نور التعليمية في عمان. مهمتك هي الإجابة على أسئلة المستخدمين حول كيفية عمل الموقع، مع التركيز على عملية التسجيل، والأدوار المختلفة للمستخدمين، وكيفية تفاعلهم مع بعضهم البعض.

إليك ملخص لكيفية عمل الموقع:

التسجيل:

أول من يسجل هو المدير.

يقوم المدير بإنشاء أكواد خاصة.

يعطي المدير هذه الأكواد إلى المنظمين.

يستخدم المنظمون الأكواد للتسجيل في الموقع باستخدام البريد الإلكتروني وكلمة المرور.

بعد التسجيل، يقوم المنظمون بإنشاء فروع الشركة.

يعطي المنظمون أكوادًا أخرى لرؤساء الأقسام.

يستخدم رؤساء الأقسام الأكواد للتسجيل في الموقع باستخدام البريد الإلكتروني وكلمة المرور.

بعد التسجيل، يقوم رؤساء الأقسام بإضافة بيانات الموظفين والطلاب.

يعطي رؤساء الأقسام أكوادًا أخرى للموظفين والطلاب.

يستخدم الموظفون والطلاب الأكواد للتسجيل في الموقع باستخدام البريد الإلكتروني وكلمة المرور بالإضافة إلى بيانات أخرى.

بعد تسجيل الطالب، يجب أن يتم قبول تسجيله من قبل رئيس القسم قبل أن يتمكن من الوصول إلى المواد التعليمية.

الأدوار والصلاحيات:

المدير: لديه صلاحيات كاملة على الموقع.

المنظمون: مسؤولون عن إنشاء فروع الشركة.

رؤساء الأقسام: مسؤولون عن إضافة بيانات الموظفين والطلاب، وقبول تسجيل الطلاب.

الموظفون: يقومون بإنشاء امتحانات، وورش عمل، ودروس مخصصة. يمكنهم الاستعانة بالذكاء الاصطناعي لتحديد الوقت المناسب للامتحانات. تتضمن أنواع أسئلة الامتحانات: أكمل، صح وخطأ، اختيار من متعدد، وتوصيل/إدراج.

الطلاب: يتلقون الامتحانات والدروس، ويتم تقييمهم من قبل الموظفين.

التفاعل:

يرسل الموظفون الامتحانات والدروس إلى الطلاب.

يتابع الموظفون درجات الطلاب وتقدمهم.

يمكن للموظفين منح جوائز للطلاب المتميزين.

عندما يسألك المستخدم سؤالاً، قم بما يلي:

فهم السؤال: تأكد من أنك فهمت السؤال جيدًا قبل الإجابة.

الاستناد إلى الملخص: استخدم الملخص أعلاه كأساس لإجاباتك.

التفصيل: قدم إجابات مفصلة وواضحة.

الأمثلة: استخدم أمثلة لتوضيح الأفكار إذا لزم الأمر.

اللهجة: حافظ على لهجة ودية ومهنية.

مثال:

المستخدم: "كيف يمكنني التسجيل كطالب؟"

أنت: "للتسجيل كطالب، تحتاج إلى الحصول على كود من رئيس القسم المسؤول عنك. بعد ذلك، يمكنك الذهاب إلى صفحة التسجيل وإدخال الكود مع بريدك الإلكتروني وكلمة المرور وبعض البيانات الأخرى المطلوبة. بعد التسجيل، سيحتاج رئيس القسم إلى قبول تسجيلك قبل أن تتمكن من الوصول إلى المواد التعليمية."

تذكر أن هدفك هو مساعدة المستخدمين على فهم كيفية عمل موقع شركة نور التعليمية."
"""


def generate_chatbot_nour(prompt):
    augmented_prompt = f"{company_nour}\n\n{prompt}"
    try:
        response = model.generate_content(augmented_prompt)
        return response.text
    except Exception as e:
        print(f"Error generating response: {e}")
        return "حدث خطأ أثناء معالجة طلبك. يرجى المحاولة مرة أخرى."

@app.route('/chatbot_nour', methods=['POST'])
def chatbot_nour():
    message = request.form['message']
    response = generate_chatbot_nour(message)
    return jsonify({'response': response})

@app.route('/chat_nour', methods=['GET', 'POST'])
def chat_nour():
    return render_template('chat_nour.html')
@app.route('/student_grades')
def student_grades():
    user_id = session.get('student_id')  # احصل على user_id من الـ query string

    if not user_id:
        return "User ID is required in the query string."

    # استعلام لاسترجاع جميع إجابات المستخدم، مرتبة حسب lesson_id (الامتحان)
    user_answers = UserAnswer4.query.filter_by(user_id=user_id).order_by(UserAnswer4.lesson_id).all()

    # تجميع الدرجات لكل امتحان
    exam_grades = {}
    for answer in user_answers:
        if answer.lesson_id not in exam_grades:
            user_exam = UserExam1.query.filter_by(lesson_id=answer.lesson_id).first()
            exams_taken = user_exam.exams_taken if user_exam else 0 
            exam_grades[answer.lesson_id] = {'total_grade': 0,'exams_taken': exams_taken ,'max_total_grade': 0, 'exam_name': f'Exam {answer.lesson_id}'}  # يمكنك استبدال exam_name باسم الامتحان الفعلي إذا كان لديك
        exam_grades[answer.lesson_id]['total_grade'] += answer.grade
        
        exam_grades[answer.lesson_id]['max_total_grade'] += answer.grade_real

    return render_template('student_grades.html', exam_grades=exam_grades, user_id=user_id,exams_taken=exams_taken)
@app.route('/student_lessons')
def student_lessons():
    student_id = session.get('student_id')
    if not student_id:
        return redirect(url_for('access'))  # Redirect if student_id is missing

    # Fetch all the lessons IDs for the provided student_id
    lesson_employees = SubmittionLesson.query.filter_by(employee_id=student_id).all()
    
    lessons_data = []
    for lesson_employee in lesson_employees:
         lesson = Lesson8.query.filter_by(id=lesson_employee.lesson_id).first()
         if lesson:
            lessons_data.append(lesson)

    return render_template('student_lessons.html', lessons=lessons_data)
@app.route('/send_lesson',  methods=['GET', 'POST'])
def send_lesson():
    lesson_id = request.args.get('id') 
    if not lesson_id:
        return jsonify({"message": "Lesson ID not provided"}), 400
    try:
        lesson = Lesson8.query.filter_by(id=lesson_id).first()
        if not lesson:
           return jsonify({"message": "Lesson not found"}), 404

        creator_employee_id = lesson.employee_id
        creator_employee = Employee6.query.filter_by(id=creator_employee_id).first()
        if not creator_employee:
            return jsonify({"message": "Creator employee not found"}), 404

        admin_id = creator_employee.id_admin
        print(admin_id)
        employees = Employee6.query.filter_by(id_admin=admin_id).all()
        students = SectorStudent8.query.filter_by(id_admin=admin_id).all()
        added_count = 0
        for employee in employees:
    
            existing_record = SubmittionLesson.query.filter_by(lesson_id=lesson_id, employee_id=employee.id).first()
            if not existing_record:
                lesson_record = SubmittionLesson(lesson_id=lesson_id, employee_id=employee.id, role="employee")
                db.session.add(lesson_record)
                print(f"Adding lesson_record - lesson_id: {lesson_record.lesson_id}, employee_id: {lesson_record.employee_id}, role: {lesson_record.role}")
                added_count+=1
            else:
                 print(f"Record already exist for lesson: {lesson_id}, and employee id: {employee.id}, skipping")

        db.session.commit()

        for student in students:
    
            existing_record = SubmittionLesson.query.filter_by(lesson_id=lesson_id, employee_id=student.id).first()
            if not existing_record:
                lesson_record = SubmittionLesson(lesson_id=lesson_id, employee_id=student.id, role="student")
                db.session.add(lesson_record)
                print(f"Adding lesson_record - lesson_id: {lesson_record.lesson_id}, employee_id: {lesson_record.employee_id}, role: {lesson_record.role}")
                added_count+=1
            else:
                 print(f"Record already exist for lesson: {lesson_id}, and employee id: {student.id}, skipping")

        db.session.commit()
        
        if added_count>0:
            return redirect(url_for('lesson_details', lesson_id=lesson_id))
        else:
            return redirect(url_for('lesson_details', lesson_id=lesson_id))
       
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': str(e)}), 500
@app.route('/update_preferences', methods=['GET', 'POST'])
def update_preferences():
    if request.method == 'POST':
        selected_columns = request.form.getlist('columns')
        email = "user@example.com"

        ColumnPreference.query.filter_by(email=email).delete()

        for column in ['id', 'username', 'email', 'password', 'account_type']:
            preference = ColumnPreference(
                email=email,
                column_name=column,
                visible=(column in selected_columns)
            )
            db.session.add(preference)
        db.session.commit()

        return redirect(url_for('success', username="user"))

    return render_template('update_preferences.html')
@app.route('/exam', methods=['GET', 'POST'])
def exam():
    lesson_id = request.args.get('lesson_id')  # استخراج قيمة lesson_id
    data_type = request.args.get('type')      # استخراج قيمة type

    if lesson_id:
        print(f"Lesson ID: {lesson_id}")
    else:
        print("Lesson ID not provided.")
    if data_type=='s':
        id=session.get('student_id')
    else:
        id=session.get('employee_id')
    if not lesson_id:
        return "Lesson ID is required", 400

    mcq_questions = Question_Multipleone_choice.query.filter_by(lesson_id=lesson_id).all()
    tf_questions = Questionfort_or_f.query.filter_by(lesson_id=lesson_id).all()
    mmcq_questions = QuestionMultipleMultiple_choice.query.filter_by(lesson_id=lesson_id).all()
    sequence_questions = Questionsequence1.query.filter_by(lesson_id=lesson_id).all()
    fillin_blank_questions = Question_Fillin_blank4.query.filter_by(lesson_id=lesson_id).all()
    fillin_blank_choice_questions = Question_Fillin_blank_choice2.query.filter_by(lesson_id=lesson_id).all()
    classification_questions = questionclassiffication.query.filter_by(lesson_id=lesson_id).all()

    all_questions = mcq_questions + tf_questions + mmcq_questions + sequence_questions + fillin_blank_questions + fillin_blank_choice_questions + classification_questions

    if 'question_order' not in session or session.get('current_lesson_id') != lesson_id:
        session['question_order'] = [q.question_id for q in all_questions]
        session['current_lesson_id'] = lesson_id
        random.shuffle(session['question_order'])
        session['current_question_index'] = 0

    current_question_index = session.get('current_question_index', 0)

    if current_question_index >= len(session['question_order']):
        session.pop('question_order', None)
        session.pop('current_question_index', None)
        session.pop('current_lesson_id', None)
         # Update UserExam1 table
        user_exam = UserExam1.query.filter_by(lesson_id=lesson_id).first()
       
        if user_exam:
            user_exam.exams_taken += 1
            user_exam.last_exam_date = datetime.now()  # Update last exam date
        else:
            user_exam = UserExam1( lesson_id=lesson_id, exams_taken=1)
            db.session.add(user_exam)

        db.session.commit()
        return redirect(url_for('index2', lesson_id=lesson_id))

    # get the question using the question id from the question order
    current_question_id = session['question_order'][current_question_index]
    current_question = None
    for question in all_questions:
        if question.question_id == current_question_id:
            current_question = question
            break

    if not current_question:
        session.pop('question_order', None)
        session.pop('current_question_index', None)
        session.pop('current_lesson_id', None)
        return "Error no question found", 404

    # Handle POST request
    if request.method == 'POST':
        start_time = request.form.get('start_time')
        end_time = datetime.now()
        time_taken = (end_time - datetime.fromisoformat(start_time)).total_seconds()

        answer = request.form.getlist('answer')
        is_correct = False
        question_grade = 0
        answer_ids = request.form.getlist('answer_id')

        if isinstance(current_question, Question_Multipleone_choice):
            correct_answer = Answer_Multiple_Multiple_choice.query.filter_by(
                question_id=current_question.question_id, is_correct=True
            ).first()
            if correct_answer and correct_answer.answer_text == answer[0] if isinstance(answer, list) and answer else None:
                is_correct = True
                question_grade = current_question.grade

        elif isinstance(current_question, Questionfort_or_f):
            if str(current_question.correct_answer).lower() == answer[0].lower() if isinstance(answer, list) and answer else None:
                is_correct = True
                question_grade = current_question.grade

        elif isinstance(current_question, QuestionMultipleMultiple_choice):
            correct_answers = Answer_Multiple_Multiple_choice.query.filter_by(
                question_id=current_question.question_id, is_correct=True
            ).all()
            correct_answers_list = {ans.answer_text for ans in correct_answers}
            user_answers_set = set(answer)

            if user_answers_set == correct_answers_list:
                is_correct = True
                question_grade = current_question.grade

        elif isinstance(current_question, Questionsequence1):
            correct_answers = Answer_sequence.query.filter_by(question_id=current_question.question_id).all()
            correct_answers_dict = {str(ans.answer_id): str(ans.index) for ans in correct_answers}
            user_answers = request.form.getlist('answer')
            user_answer_ids = request.form.getlist('answer_id')

            is_correct = True
            if len(user_answers) != len(correct_answers_dict) or len(user_answer_ids) != len(correct_answers_dict):
                is_correct = False
            else:
                for answer_id, user_index in zip(user_answer_ids, user_answers):
                    if answer_id not in correct_answers_dict or correct_answers_dict[answer_id] != user_index:
                        is_correct = False
                        break

            if is_correct:
                question_grade = current_question.grade
        elif isinstance(current_question, Question_Fillin_blank4):

            # استخراج جميع الاجابات الممكنه والاسئله
            question_texts = [
                current_question.question_text,
                current_question.question_text1,
                current_question.question_text2,
                current_question.question_text3,
                current_question.question_text4,
                current_question.question_text5,
                current_question.question_text6
            ]

            answer_texts = [
                current_question.answer_text,
                current_question.answer_text1,
                current_question.answer_text2,
                current_question.answer_text3,
                current_question.answer_text4,
                current_question.answer_text5,
                current_question.answer_text6
            ]

            # تنظيف القوائم من القيم الفارغة (None)
            question_texts = [text for text in question_texts if text]
            answer_texts = [text for text in answer_texts if text]

            # استخراج الاجابات من الريكوست
            user_answers_str = request.form.get('answer')

            # تحويل الاجابات الي قاموس
            user_answers = json.loads(user_answers_str) if user_answers_str else []
            # التحقق من تطابق الاجابات
            is_correct = True
            if len(question_texts) != len(user_answers):
                is_correct = False
            else:
                for i, user_answer in enumerate(user_answers):
                    if i < len(answer_texts) and answer_texts[i] != user_answer:
                        is_correct = False
                        break

            if is_correct:
                question_grade = current_question.grade
        elif isinstance(current_question, Question_Fillin_blank_choice2):
            # استخراج جميع الاجابات الممكنه والاسئله
            question_texts = [
                current_question.question_text,
                current_question.question_text1,
                current_question.question_text2,
                current_question.question_text3,
                current_question.question_text4,
                current_question.question_text5,
                current_question.question_text6
            ]

            answer_texts = [
                current_question.answer_text,
                current_question.answer_text1,
                current_question.answer_text2,
                current_question.answer_text3,
                current_question.answer_text4,
                current_question.answer_text5,
                current_question.answer_text6
            ]

            # تنظيف القوائم من القيم الفارغة (None)
            question_texts = [text for text in question_texts if text]
            answer_texts = [text for text in answer_texts if text]

            # استخراج الاجابات من الريكوست
            user_answers = request.form.getlist('answer')

            # التحقق من تطابق الاجابات
            is_correct = True
            if len(question_texts) != len(user_answers):
                is_correct = False
            else:
                for i, question_text in enumerate(question_texts):
                    if i < len(answer_texts) and answer_texts[i] != user_answers[i]:
                        is_correct = False
                        break

            if is_correct:
                question_grade = current_question.grade
        elif isinstance(current_question, questionclassiffication):
            labels = [
                current_question.label_text,
                current_question.label_text1,
                current_question.label_text2,
                current_question.label_text3
            ]
            labels = [label for label in labels if label]
            user_answers = request.form.getlist('answer')
            answer_ids = request.form.getlist('answer_id')
            correct_answers = Answer_classiffication1.query.filter(
                Answer_classiffication1.question_id == current_question.question_id).all()
            correct_answers_dict = {str(ans.answer_id): ans.label for ans in correct_answers}
          
            is_correct = True
            if len(user_answers) != len(correct_answers_dict) or len(user_answers) != len(answer_ids):
                is_correct = False
            else:
                 for answer_id, label  in zip(answer_ids, user_answers) :
                        if answer_id not in correct_answers_dict or str(correct_answers_dict[answer_id]) != label:
                            is_correct = False
                            break
            if is_correct:
                question_grade = current_question.grade

        new_answer = UserAnswer4(
            user_id=id,
            question_id=current_question.question_id,
            time_allowed=current_question.times,
            lesson_id=lesson_id,
            is_correct=is_correct,
            time_taken=time_taken,
            grade=question_grade,
            grade_real=current_question.grade, 
            type=id
        )
        db.session.add(new_answer)
        db.session.commit()

        session['current_question_index'] += 1
        return redirect(url_for('exam', lesson_id=lesson_id))

    question_data = {
        'type': current_question.type,
        'difficulty': current_question.difficulty,
        'importance': current_question.importance,
        'bloom_taxonomy': current_question.bloom_taxonomy,
        'grade': current_question.grade,
        'times': current_question.times,
        'question_text': current_question.question_name if isinstance(current_question,
                                                                     Question_Multipleone_choice) or isinstance(
            current_question, QuestionMultipleMultiple_choice) or isinstance(current_question,
                                                                            Questionsequence1) else (
            current_question.question_text if isinstance(current_question, Question_Fillin_blank4) or isinstance(
                current_question, Question_Fillin_blank_choice2) else (current_question.question_name if hasattr(current_question,'question_name') else current_question.label_text if isinstance(current_question,questionclassiffication) else '')),
    }
    if not (isinstance(current_question, Question_Fillin_blank4) or isinstance(current_question,
                                                                            Question_Fillin_blank_choice2) or isinstance(current_question,questionclassiffication)):
        question_data['audio_path'] = current_question.audio_path
        question_data['photo_path'] = current_question.photo_path
    if isinstance(current_question, Question_Multipleone_choice):
        answers = Answer_Multiple_Multiple_choice.query.filter_by(question_id=current_question.question_id).all()
        question_data['answers'] = [{'answer_text': answer.answer_text} for answer in answers]
    elif isinstance(current_question, QuestionMultipleMultiple_choice):
        answers = Answer_Multiple_Multiple_choice.query.filter_by(question_id=current_question.question_id).all()
        question_data['answers'] = [{'answer_text': answer.answer_text} for answer in answers]
    elif isinstance(current_question, Questionsequence1):
        answers = Answer_sequence.query.filter_by(question_id=current_question.question_id).all()
        question_data['answers'] = [
            {'answer_text': answer.answer_text, 'index': answer.index, 'answer_id': answer.answer_id} for answer in
            answers]
    elif isinstance(current_question, Question_Fillin_blank4):
        question_data['question_texts'] = [
            current_question.question_text,
            current_question.question_text1,
            current_question.question_text2,
            current_question.question_text3,
            current_question.question_text4,
            current_question.question_text5,
            current_question.question_text6
        ]
        question_data['answers'] = [
            current_question.answer_text,
            current_question.answer_text1,
            current_question.answer_text2,
            current_question.answer_text3,
            current_question.answer_text4,
            current_question.answer_text5,
            current_question.answer_text6
        ]
        question_data['answers'] = [answer for answer in question_data['answers'] if answer]
        question_data['question_texts'] = [question for question in question_data['question_texts'] if question]
    elif isinstance(current_question, Question_Fillin_blank_choice2):
        question_data['question_texts'] = [
            current_question.question_text,
            current_question.question_text1,
            current_question.question_text2,
            current_question.question_text3,
            current_question.question_text4,
            current_question.question_text5,
            current_question.question_text6
        ]
        question_data['answers'] = [
            current_question.answer_text,
            current_question.answer_text1,
            current_question.answer_text2,
            current_question.answer_text3,
            current_question.answer_text4,
            current_question.answer_text5,
            current_question.answer_text6
        ]
        question_data['answers'] = [answer for answer in question_data['answers'] if answer]
        question_data['question_texts'] = [question for question in question_data['question_texts'] if question]
    elif isinstance(current_question, questionclassiffication):
        labels = [
            current_question.label_text,
            current_question.label_text1,
            current_question.label_text2,
            current_question.label_text3
        ]
        question_data['labels'] = [label for label in labels if label]
        answers = Answer_classiffication1.query.filter(
            Answer_classiffication1.question_id == current_question.question_id).all()
        question_data['answers'] = [{'answer_text': answer.answer_text, 'answer_id': answer.answer_id} for answer in
                                   answers]

    return render_template(
        'exam.html',
        question=question_data,
        question_index=current_question_index,
        total_questions=len(all_questions),
        start_time=datetime.now().isoformat()
    )
@app.route('/exam_summary')
def exam_summary():
    lesson_id = request.args.get('lesson_id', type=int)
    user_id = session.get('user_id')

    if not lesson_id or not user_id:
      return "Lesson ID and User ID is required", 400

    user_answers = UserAnswer4.query.filter_by(user_id=user_id, lesson_id=lesson_id).all()

    total_grade = sum([answer.grade for answer in user_answers if answer.is_correct])
    total_questions = len(user_answers)
    correct_answers_count = sum([1 for answer in user_answers if answer.is_correct])
    incorrect_answers_count = total_questions - correct_answers_count

    correct_in_time_count = sum([1 for answer in user_answers if answer.is_correct and answer.time_taken <= answer.time_allowed])

    return render_template(
      'exam_summary.html', 
      total_grade=total_grade,
      total_questions=total_questions,
      correct_answers_count=correct_answers_count,
      incorrect_answers_count=incorrect_answers_count,
      answers=user_answers,  
      correct_in_time_count = correct_in_time_count

    )
@app.route('/index2')
def index2():
    lesson_id = request.args.get('lesson_id', type=int)
    user_id = session.get('user_id')

    if not lesson_id or not user_id:
      return "Lesson ID and User ID is required", 400

    user_answers = UserAnswer4.query.filter_by(user_id=user_id, lesson_id=lesson_id).all()

    total_grade = sum([answer.grade for answer in user_answers if answer.is_correct])
    total_questions = len(user_answers)
    correct_answers_count = sum([1 for answer in user_answers if answer.is_correct])
    incorrect_answers_count = total_questions - correct_answers_count
    correct_in_time_count = sum([1 for answer in user_answers if answer.is_correct and answer.time_taken <= answer.time_allowed])
    late=total_questions - correct_in_time_count
    
    percentage_correct_In_Time= (correct_in_time_count / total_questions) * 100
    
    # حساب النسبة المئوية
    if total_questions > 0:
        percentage_correct = (correct_answers_count / total_questions) * 100
    else:
        percentage_correct = 0  # أو None، أو قيمة افتراضية أخرى

    # Get exam data from UserExam1 table
    user_exam = UserExam1.query.filter_by(lesson_id=lesson_id).first()
    print(f"UserExam1 object: {late}")
    print(f"UserExam1 object: {user_exam}")  # Print the UserExam1 object
    if user_exam:
        num_practices = user_exam.exams_taken
        last_practice_date = user_exam.last_exam_date
    else:
        num_practices = 0
        last_practice_date = None

    print(f"num_practices: {num_practices}, type: {type(num_practices)}")  # Print num_practices and its type
    print(f"last_practice_date: {last_practice_date}, type: {type(last_practice_date)}")  # Print last_practice_date and its type

    # Calculate time since last practice
    if last_practice_date:
        time_since_last_practice = datetime.now() - last_practice_date
        if time_since_last_practice < timedelta(hours=24):
            last_practice_text = f"{int(time_since_last_practice.total_seconds() // 3600)} hours ago"
        elif time_since_last_practice < timedelta(days=30):
             last_practice_text = f"{int(time_since_last_practice.days)} days ago"
        else:
            last_practice_text = last_practice_date.strftime("%Y-%m-%d")
    else:
        last_practice_text = "No practices yet"

    print(f"last_practice_text: {last_practice_text}, type: {type(last_practice_text)}")  # Print last_practice_text and its type

    return render_template(
      'index2.html', 
      total_grade=total_grade,
      total_questions=total_questions,
      correct_answers_count=correct_answers_count,
      incorrect_answers_count=incorrect_answers_count,
      answers=user_answers,  
      percentage_correct_In_Time=percentage_correct_In_Time,
      correct_in_time_count = correct_in_time_count,
      percentage_correct=percentage_correct,  
      num_practices=num_practices,  # عدد مرات الممارسة
      last_practice_text=last_practice_text,   # تاريخ آخر ممارسة
      late=late,
      
      lesson_id=lesson_id
    )
@app.route('/signup_education', methods=['GET', 'POST'])
def signup_education():
    if request.method == 'POST' and request.form.get('foundation_name')!=None:
        foundation_name = request.form.get('foundation_name')
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        existing_Education = Education.query.filter_by(email=email).first()
        if existing_Education:
            return render_template('signup_education.html', error="Email already exists. Please choose a different one.")

        if password != confirm_password:
            return render_template('signup_education.html', error="Passwords do not match.")

        new_Education = Education(
            foundation_name=foundation_name,
            username=username,
            email=email,
            password=password 
        )
        db.session.add(new_Education)
        db.session.commit()

        return redirect(url_for('success', username=username))

    return render_template('signup_education.html')   

@app.route('/access', methods=['GET', 'POST'])
def access():
    return render_template('access.html')  

@app.route('/accept', methods=['GET', 'POST'])
def accept():
    return render_template('accept.html')  

@app.route('/custom_lesson_details', methods=['GET', 'POST'])
def custom_lesson_details():
    lesson_id = request.args.get('lesson_id', type=int)
    lesson = Lesson_custom1.query.filter_by(id=lesson_id).first()

    if not lesson:
        return "Lesson not found", 404

    if request.method == 'POST':
        # تعديل البيانات
        title = request.form.get('title', lesson.title)
        link = request.form.get('link', lesson.link)
        description = request.form.get('description', lesson.description)
        
        lesson.title = title
        lesson.link = link
        lesson.description = description

        # معالجة الملفات المرفوعة
        files = request.files.getlist('file')
        image = request.files.get('image')
        video = request.files.get('video')
        audio = request.files.get('audio')

        if files:
            for file in files:
                if not lesson.file_path:
                    lesson.file_path = save_file(file, 'file', lesson_id, lesson.file_path)
                elif not lesson.file_path1:
                    lesson.file_path1 = save_file(file, 'file1', lesson_id, lesson.file_path1)
                elif not lesson.file_path2:
                    lesson.file_path2 = save_file(file, 'file2', lesson_id, lesson.file_path2)
                elif not lesson.file_path3:
                    lesson.file_path3 = save_file(file, 'file3', lesson_id, lesson.file_path3)
                elif not lesson.file_path4:
                    lesson.file_path4 = save_file(file, 'file4', lesson_id, lesson.file_path4)

        if image:
            lesson.image_path = save_file(image, 'image', lesson_id, lesson.image_path)
        if video:
            lesson.video_path = save_file(video, 'video', lesson_id, lesson.video_path)
        if audio:
            lesson.audio_path = save_file(audio, 'audio', lesson_id, lesson.audio_path)

        db.session.commit()
        return redirect(url_for('custom_lesson_details', lesson_id=lesson_id))

    return render_template('custom_lesson_details.html', lesson=lesson)

@app.route('/Custom_Lesson', methods=['GET', 'POST'])
def Custom_Lesson():
    lessons = Lesson_custom1.query.all()

    lessons_with_details = []
    for lesson in lessons:
        student_count = LessonStudent.query.filter_by(lesson_id=lesson.id).count()
        
        employee_count = LessonEmployee.query.filter_by(lesson_id=lesson.id).count()
        
        lessons_with_details.append({
            'lesson': lesson,
            'student_count': student_count,
            'employee_count': employee_count,
        })

    return render_template('Custom_Lesson.html', lessons_with_details=lessons_with_details)


@app.route('/add_CustumLesson', methods=['GET', 'POST'])
def add_CustumLesson():
    if request.method == 'POST':

        title = request.form.get('title')
        image = request.files.get('image')
        employee_id = session.get('employee_id')

        if not title or not employee_id:
            return "Title and Employee ID are required", 400

        new_lesson = Lesson_custom1(
            title=title,
            employee_id=employee_id,
            image_path=None  
        )
        db.session.add(new_lesson)
        db.session.commit()

        if image:
            image_path = f"static/uploads/{new_lesson.id}_L_{image.filename}"
            image.save(image_path)

            # تحديث مسار الصورة
            new_lesson.image_path = image_path
            db.session.commit()

        return redirect(url_for('add_student', lesson_id=new_lesson.id))

    return render_template('add_CustumLesson.html')
    
@app.route('/add_students_to_lesson', methods=['POST'])
def add_students_to_lesson():
    lesson_id = request.args.get('lesson_id', type=int)
    if not lesson_id:
        return "Lesson ID is required", 400

    # الحصول على قائمة IDs الطلاب المحددين
    student_ids = request.form.getlist('student_ids')

    if not student_ids:
        return "No students selected", 400

    all_entries = LessonStudent.query.all()
    print("Current LessonStudent Table Data:")
    for entry in all_entries:
        print(f"Lesson ID: {entry.lesson_id}, Student ID: {entry.student_id}")

    for student_id in student_ids:
        exists = LessonStudent.query.filter_by(lesson_id=lesson_id, student_id=student_id).first()
        if not exists:
            new_entry = LessonStudent(
                lesson_id=lesson_id,
                student_id=student_id
            )
            db.session.add(new_entry)

    db.session.commit()
    return redirect(url_for('add_Emolpoyee', lesson_id=lesson_id))

@app.route('/add_emloyees_to_lesson', methods=['POST'])
def add_emloyees_to_lesson():
    lesson_id = request.args.get('lesson_id', type=int)
    if not lesson_id:
        return "Lesson ID is required", 400

    # الحصول على قائمة IDs الطلاب المحددين
    student_ids = request.form.getlist('student_ids')

    if not student_ids:
        return "No students selected", 400

    all_entries = LessonEmployee.query.all()
    print("Current LessonEmployee Table Data:")
    for entry in all_entries:
        print(f"Lesson ID: {entry.lesson_id}, Student ID: {entry.student_id}")

    for student_id in student_ids:
        exists = LessonEmployee.query.filter_by(lesson_id=lesson_id, student_id=student_id).first()
        if not exists:
            new_entry = LessonEmployee(
                lesson_id=lesson_id,
                student_id=student_id
            )
            db.session.add(new_entry)

    db.session.commit()
    return redirect(url_for('Custom_Lesson', lesson_id=lesson_id))

@app.route('/add_student', methods=['GET', 'POST'])
def add_student():
    employee_id = session.get('employee_id')
    sector_code = session.get('sector_code')
    print(employee_id)
    # جلب بيانات الموظف باستخدام ID
    employee = Employee6.query.filter_by(id=employee_id).first()
    if not employee:
        return "Employee not found", 404

    # استخراج id_admin الخاص بالموظف
    admin_id = employee.id_admin

    # تصفية الموظفين الذين يتبعون نفس id_admin
    employees_under_same_admin = Employee6.query.filter_by(id_admin=admin_id).all()

    students_in_sector = SectorStudent8.query.filter_by(id_admin=admin_id).all()

    return render_template(
        'add_student.html',
        employees=employees_under_same_admin,
        students=students_in_sector,
        sector_code=sector_code
    )


@app.route('/add_Emolpoyee', methods=['GET', 'POST'])
def add_Emolpoyee():
    employee_id = session.get('employee_id')
    sector_code = session.get('sector_code')
    print(employee_id)
    # جلب بيانات الموظف باستخدام ID
    employee = Employee6.query.filter_by(id=employee_id).first()
    if not employee:
        return "Employee not found", 404

    admin_id = employee.id_admin

    employees_under_same_admin = Employee6.query.filter_by(id_admin=admin_id).all()

    students_in_sector = SectorStudent8.query.filter_by(id_admin=admin_id).all()

    return render_template(
        'add_Emolpoyee.html',
        employees=employees_under_same_admin,
        students=students_in_sector,
        sector_code=sector_code
    )

@app.route('/edit_question_true_or_false', methods=['GET', 'POST'])
def edit_question_true_or_false():
    sector_code = session.get('sector_code')
    lesson_id = request.args.get('lesson_id')
    questions = Questionfort_or_f.query.filter_by(lesson_id=lesson_id).all()
    return render_template('edit_question_true_or_false.html',lesson_id=lesson_id,questions=questions,sector_code=sector_code)

@app.route('/add_question_T_or_F/<int:lesson_id>', methods=['POST'])
def add_question_T_or_F(lesson_id):
    question = Questionfort_or_f(
        lesson_id=lesson_id,
        question_name="Default question Click to edit",
        difficulty="Easy",
        importance="Necessary",
        bloom_taxonomy="Analyze",
        grade=1,
        times=7,
        correct_answer=True
    )
    db.session.add(question)
    db.session.commit()

    return redirect(url_for('edit_question_true_or_false', lesson_id=lesson_id))

@app.route('/edit_question_Multiple_Multiple', methods=['GET', 'POST'])
def edit_question_Multiple_Multiple():
    sector_code = session.get('sector_code')
    lesson_id = request.args.get('lesson_id')
    questions = QuestionMultipleMultiple_choice.query.filter_by(lesson_id=lesson_id).all()
    return render_template('edit_question_Multiple_Multiple.html',lesson_id=lesson_id,questions=questions,sector_code=sector_code) 

@app.route('/edit_fill_in_blank', methods=['GET', 'POST'])
def edit_fill_in_blank():
    sector_code = session.get('sector_code')
    lesson_id = request.args.get('lesson_id')
    questions = Question_Fillin_blank4.query.filter_by(lesson_id=lesson_id).all()
    return render_template('edit_fill_in_blank.html',lesson_id=lesson_id,questions=questions,sector_code=sector_code) 


@app.route('/add_fill_in_blank/<int:lesson_id>', methods=['POST'])
def add_fill_in_blank(lesson_id):

    question = Question_Fillin_blank4(
        lesson_id=lesson_id,
        question_text="Default question Click to edit",
        question_text1="Default question Click to edit",
        question_text2="Default question Click to edit",
        difficulty="Easy",
        importance="Necessary",
        bloom_taxonomy="Analyze",
        grade=1,
        times=7,
        answer_text="Default answer",
        answer_text1="Default answer",
        answer_text2="Default answer"
    )
    db.session.add(question)
    db.session.commit()
  
    return redirect(url_for('edit_fill_in_blank', lesson_id=lesson_id))

def calculate_expected_time_fill(question):
    expected_time = 0
    
    # حساب الوقت بناءً على طول النص في السؤال
    if question.question_text:
        expected_time += len(question.question_text) // 5  # ثانية لكل 5 حروف
    if question.question_text1:
        expected_time += len(question.question_text1) // 5  # ثانية لكل 5 حروف
    if question.question_text2:
        expected_time +=5
        expected_time += len(question.question_text2) // 5  # ثانية لكل 5 حروف
        expected_time +=5
    if question.question_text3:
        expected_time +=5
        expected_time += len(question.question_text3) // 5  # ثانية لكل 5 حروف
    if question.question_text4:
        expected_time +=5
        expected_time += len(question.question_text4) // 5  # ثانية لكل 5 حروف

    if question.question_text5:
        expected_time +=5
        expected_time += len(question.question_text5) // 5  # ثانية لكل 5 حروف
    if question.question_text6:
        expected_time +=5
        expected_time += len(question.question_text6) // 5  # ثانية لكل 5 حروف


    # إضافة 20 ثانية إذا كانت هناك صورة
    #if question.photo_path:
    #    expected_time += 20

    # إضافة وقت الصوت و 5 ثواني إذا كان هناك ملف صوتي
    #if question.audio_path:
    #     try:
    #        audio_path = os.path.join(app.config['UPLOAD_FOLDER'], question.audio_path)
    #        if audio_path.lower().endswith(('.mp3')):
    #              audio = MP3(audio_path)
    #              audio_duration = audio.info.length
    #              expected_time += audio_duration + 5
    #        elif audio_path.lower().endswith(('.wav')):
    #             audio = WAVE(audio_path)
    #             audio_duration = audio.info.length
    #             expected_time += audio_duration + 5
    #        elif audio_path.lower().endswith(('.ogg')):
    #              audio = OggVorbis(audio_path)
    #              audio_duration = audio.info.length
    #              expected_time += audio_duration + 5
    #        elif audio_path.lower().endswith(('.aac')):
    #              audio = AAC(audio_path)
    #              audio_duration = audio.info.length
    #              expected_time += audio_duration + 5
    #        else:
    #              expected_time += 5
    #     except Exception as e:
    #        print(f"Error getting audio duration: {e}")
    #        expected_time += 5
    return int(expected_time+7)

@app.route('/fill_in_blank', methods=['GET', 'POST'])
def fill_in_blank():
    sector_code = session.get('sector_code')
    question_id = request.args.get('question_id', type=int)
    question = Question_Fillin_blank4.query.get_or_404(question_id)
    
    if request.method == 'POST':
        question.question_text = request.form['question_text']
        question.question_text1 = request.form['question_text1']
        question.question_text2 = request.form['question_text2']
        question.question_text3 = request.form['question_text3']
        question.question_text4 = request.form['question_text4']
        question.question_text5 = request.form['question_text5']
        question.question_text6 = request.form['question_text6']
     
        question.difficulty = request.form['difficulty']

        question.importance = request.form['importance']
        question.bloom_taxonomy = request.form['bloom_taxonomy']
        question.grade = request.form['grade']
        question.times = request.form['times']
        question.answer_text = request.form['answer']
        question.answer_text1 = request.form['answer1']
        question.answer_text2 = request.form['answer2']
        question.answer_text3 = request.form['answer3']
        question.answer_text4 = request.form['answer4']
        question.answer_text5 = request.form['answer5']
        question.answer_text6 = request.form['answer6']
        try:
            db.session.commit()
        except Exception as e:
            print(f"Error saving data: {e}")
    expected_time = calculate_expected_time_fill(question)

    return render_template('fill_in_blank.html', question=question,sector_code=sector_code,expected_time=expected_time)

@app.route('/delete_fill/<int:question_id>', methods=['POST'])
def delete_fill(question_id):
    question = Question_Fillin_blank4.query.get_or_404(question_id)
    lesson_id = question.lesson_id 
    try:
        db.session.delete(question)
        db.session.commit()
      
    except Exception as e:
        db.session.rollback()
        

    return redirect(url_for('edit_fill_in_blank', lesson_id=lesson_id))

@app.route('/delete_fill_choice/<int:question_id>', methods=['POST'])
def delete_fill_choice(question_id):
    question = Question_Fillin_blank_choice2.query.get_or_404(question_id)
    lesson_id = question.lesson_id 
    try:
        db.session.delete(question)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
    return redirect(url_for('edit_fill_in_blank_choice', lesson_id=lesson_id))

@app.route('/edit_fill_in_blank_choice', methods=['GET', 'POST'])
def edit_fill_in_blank_choice():
    sector_code = session.get('sector_code')
    lesson_id = request.args.get('lesson_id')
    questions = Question_Fillin_blank_choice2.query.filter_by(lesson_id=lesson_id).all()
    return render_template('edit_fill_in_blank_choice.html',lesson_id=lesson_id,questions=questions,sector_code=sector_code) 


@app.route('/add_fill_in_blank_choice/<int:lesson_id>', methods=['POST'])
def add_fill_in_blank_choice(lesson_id):
    question = Question_Fillin_blank_choice2(
        lesson_id=lesson_id,
        question_text="Default question Click to edit",
        question_text1="Default question Click to edit",
        question_text2="Default question Click to edit",
        difficulty="Easy",
        importance="Necessary",
        bloom_taxonomy="Analyze",
        grade=1,
        times=7,
        answer_text="الاجابه",
        answer_text1="الاجابه",
        answer_text2="الاجابه"
    )
    db.session.add(question)
    db.session.commit()
    
    return redirect(url_for('edit_fill_in_blank_choice', lesson_id=lesson_id))


@app.route('/fill_in_blank_choice', methods=['GET', 'POST'])
def fill_in_blank_choice():
    sector_code = session.get('sector_code')
    question_id = request.args.get('question_id', type=int)
    question = Question_Fillin_blank_choice2.query.get_or_404(question_id)
    
    if request.method == 'POST':
        question.question_text = request.form['question_text']
        question.question_text1 = request.form['question_text1']
        question.question_text2 = request.form['question_text2']
        question.question_text3 = request.form['question_text3']
        question.question_text4 = request.form['question_text4']
        question.question_text5 = request.form['question_text5']
        question.question_text6 = request.form['question_text6']
     
        question.difficulty = request.form['difficulty']
        question.importance = request.form['importance']
        question.bloom_taxonomy = request.form['bloom_taxonomy']
        question.grade = request.form['grade']
        question.times = request.form['times']
        question.answer_text = request.form['answer']
        question.answer_text1 = request.form['answer1']
        question.answer_text2 = request.form['answer2']
        question.answer_text3 = request.form['answer3']
        question.answer_text4 = request.form['answer4']
        question.answer_text5 = request.form['answer5']
        question.answer_text6 = request.form['answer6']
        try:
            db.session.commit()
        except Exception as e:
            print(f"Error saving data: {e}")
    expected_time = calculate_expected_time_fill(question)
    return render_template('fill_in_blank_choice.html', question=question,sector_code=sector_code,expected_time=expected_time)

@app.route('/Article', methods=['GET', 'POST'])
def article():
    sector_code = session.get('sector_code')
    lesson_id = request.args.get('lesson_id')
    questions = Article_question.query.filter_by(lesson_id=lesson_id).all()
    return render_template('Article.html',lesson_id=lesson_id,questions=questions,sector_code=sector_code) 

@app.route('/accept_member/<int:member_id>/<int:workshop_id>', methods=['POST'])
def accept_member(member_id, workshop_id):
    member = Workshop_Register.query.get(member_id)
    if member:
        member.accepted = True
        db.session.commit()
    return redirect(url_for('workshop_member', workshop_id=workshop_id))

@app.route('/delete_member/<int:member_id>/<int:workshop_id>', methods=['POST'])
def delete_member(member_id, workshop_id):
    member = Workshop_Register.query.get(member_id)
    if member:
        db.session.delete(member)
        db.session.commit()
    return redirect(url_for('workshop_member', workshop_id=workshop_id))

@app.route('/workshop_Register', methods=['GET'])
def workshop_register():
    workshop_id = request.args.get('workshop_id')
    if not workshop_id:
        return redirect(url_for('home'))
    
    student_id = session.get('student_id')
    if student_id:

        student = SectorStudent8.query.filter_by(id=student_id).first()
        if not student:
            return redirect(url_for('home'))

        existing_registration = Workshop_Register.query.filter_by(workshop_id=workshop_id, user_id=student.id).first()
        if existing_registration:
            return render_template('success.html', workshop_id=workshop_id)

        registration = Workshop_Register(
            workshop_id=workshop_id,
            user_id=student.id,
            account_type="Student",
            name=student.username,
            email=student.email,
            country=student.country,
            city=student.province
        )
        db.session.add(registration)
        db.session.commit()
        return render_template('success.html', workshop_id=workshop_id)

    employee_id = session.get('employee_id')
    if employee_id:

        employee = Employee6.query.filter_by(id=employee_id).first()
        if not employee:
            return redirect(url_for('home'))

        existing_registration = Workshop_Register.query.filter_by(workshop_id=workshop_id, user_id=employee.id).first()
        if existing_registration:
            return render_template('success.html', workshop_id=workshop_id)

        registration = Workshop_Register(
            workshop_id=workshop_id,
            user_id=employee.id,
            account_type="Employee",
            name=employee.username,
            email=employee.email,
            country=employee.country,
            city=employee.province
        )
        db.session.add(registration)
        db.session.commit()
        return render_template('success.html', workshop_id=workshop_id)

    

    return redirect(url_for('home'))


@app.route('/workshop_member', methods=['GET'])
def workshop_member():
    workshop_id = request.args.get('workshop_id')
    if not workshop_id:
        return redirect(url_for('home')) 

    members = Workshop_Register.query.filter_by(workshop_id=workshop_id).all()

    return render_template('workshop_member.html', workshop_id=workshop_id, members=members)

def generate_qr_code(url):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(url)
    qr.make(fit=True)

    img = qr.make_image(fill='black', back_color='white')
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    qr_code_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    return qr_code_base64

@app.route('/workshop_details', methods=['GET', 'POST'])
def workshop_details():
    sector_code = session.get('sector_code')
    workshop_id = request.args.get('workshop_id')
    workshop = Work_shop4.query.get_or_404(workshop_id)
    
    workshop_url = f"http://127.0.0.1:5000/workshop_Register?workshop_id={workshop_id}"
    qr_code_base64 = generate_qr_code(workshop_url)

    if request.method == 'POST':
        workshop.name = request.form.get('name')
        workshop.type = request.form.get('type')
        workshop.days_count = int(request.form.get('days_count'))
        workshop.person_count = int(request.form.get('person_count'))
        workshop.location = request.form.get('location')
        workshop.country = request.form.get('country')
        workshop.city = request.form.get('city')
        workshop.description = request.form.get('description')
        workshop.trainer = request.form.get('trainer')
        workshop.keywords = request.form.get('keywords')  # حفظ الكلمات المفتاحية
        
        workshop.start_date = datetime.strptime(request.form.get('start_date'), '%Y-%m-%d').date()
        workshop.end_date = datetime.strptime(request.form.get('end_date'), '%Y-%m-%d').date()
        
        def parse_time(time_str):
            try:
                return datetime.strptime(time_str, '%H:%M:%S').time()
            except ValueError:
                return datetime.strptime(time_str + ":00", '%H:%M:%S').time()
        
        workshop.start_time = parse_time(request.form.get('start_time'))
        workshop.end_time = parse_time(request.form.get('end_time'))
        
        if 'image' in request.files:
            image = request.files['image']
            if image and allowed_file(image.filename):
                filename = f"{workshop.id}_{secure_filename(image.filename)}"  
                image_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                if workshop.image and os.path.exists(os.path.join(app.config['UPLOAD_FOLDER'], workshop.image)):
                    os.remove(os.path.join(app.config['UPLOAD_FOLDER'], workshop.image))  
                image.save(image_path)
                workshop.image = f"uploads/{filename}"
        
        db.session.commit()
        flash('Workshop updated successfully!', 'success')
        return redirect(url_for('workshop_details', workshop_id=workshop.id))
    
    return render_template('workshop_details.html', sector_code=sector_code, workshop=workshop, qr_code=qr_code_base64)

@app.route('/workshop_details_search/<int:workshop_id>')
def workshop_details_search(workshop_id):
    sector_code = session.get('sector_code')
    workshop = Work_shop4.query.get_or_404(workshop_id)
    workshop_url = f"http://127.0.0.1:5000/workshop_Register?workshop_id={workshop_id}"
    qr_code_base64 = generate_qr_code(workshop_url)
    
    return render_template('workshop_details_search.html', sector_code=sector_code, workshop=workshop, qr_code=qr_code_base64)

@app.route('/gift_details', methods=['GET', 'POST'])
def gift_details():
    sector_code = session.get('sector_code')
    gift_id = request.args.get('gift_id')
    gift = Reward2.query.get_or_404(gift_id)

    if request.method == 'POST':
        title = request.form['title']
        number_of_gifts = request.form['number_of_gifts']
        title1 = request.form.get('title1')  # الحقل name1
        number_of_gifts1 = request.form.get('number_of_gifts1')  # الحقل total_gifts1
        title2 = request.form.get('title2')  # الحقل name2
        number_of_gifts2 = request.form.get('number_of_gifts2')  # الحقل total_gifts2
        title3 = request.form.get('title3')  # الحقل name3
        number_of_gifts3 = request.form.get('number_of_gifts3')  # الحقل total_gifts3
        title4 = request.form.get('title4')  # الحقل name4
        number_of_gifts4 = request.form.get('number_of_gifts4')  # الحقل total_gifts4

        description = request.form['description']
        start_date = request.form['start_date']
        end_date = request.form['end_date']
        start_time = request.form['start_time']
        end_time = request.form['end_time']
        gift_type = request.form['gift_type']
        percentage = request.form['percentage']
        number_of_tests = request.form['number_of_tests']
        number_of_practices = request.form['number_of_practices']
        account_type = request.form['account_type']

        # تحويل التواريخ والأوقات
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        start_time = datetime.strptime(start_time, '%H:%M:%S').time()
        end_time = datetime.strptime(end_time, '%H:%M:%S').time()

        # تحديث البيانات في الكائن
        gift.name = title
        gift.total_gifts = number_of_gifts
        gift.name1 = title1
        gift.total_gifts1 = number_of_gifts1
        gift.name2 = title2
        gift.total_gifts2 = number_of_gifts2
        gift.name3 = title3
        gift.total_gifts3 = number_of_gifts3
        gift.name4 = title4
        gift.total_gifts4 = number_of_gifts4
        gift.description = description
        gift.start_date = start_date
        gift.end_date = end_date
        gift.start_time = start_time
        gift.end_time = end_time
        gift.is_internal = True if gift_type == 'Internal' else False
        gift.percentage = percentage
        gift.exam_count = number_of_tests
        gift.practice_count = number_of_practices if number_of_practices else 4
        gift.account_type = account_type

        # حفظ الصورة (إذا تم رفعها)
        if 'image' in request.files:
            image = request.files['image']
            if image and allowed_file(image.filename):
                filename = f"{gift.id}_{secure_filename(image.filename)}"
                image_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                if gift.image and os.path.exists(os.path.join(app.config['UPLOAD_FOLDER'], gift.image)):
                    os.remove(os.path.join(app.config['UPLOAD_FOLDER'], gift.image))
                image.save(image_path)
                gift.image = f"uploads/{filename}"

        # حفظ التغييرات في قاعدة البيانات
        db.session.commit()
        return redirect(url_for('gift_details', gift_id=gift.id))

    return render_template('gift_details.html', gift=gift,sector_code=sector_code)

@app.route('/delete_gift/<int:gift_id>', methods=['POST'])
def delete_gift(gift_id):
    
    gift = Reward2.query.get_or_404(gift_id)

    if gift.image and os.path.exists(os.path.join(app.config['UPLOAD_FOLDER'], gift.image)):
        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], gift.image))

    db.session.delete(gift)
    db.session.commit()

    return redirect(url_for('gift'))  


@app.route('/gift', methods=['GET', 'POST'])
def gift():
    sector_code = session.get('sector_code')
    gifts = Reward2.query.all()
    return render_template('gift.html', gifts=gifts,sector_code=sector_code)

@app.route('/work_shop', methods=['GET', 'POST'])
def work_shop():
    sector_code = session.get('sector_code')
    work_shops = Work_shop4.query.all()
    return render_template('work_shop.html', work_shops=work_shops, sector_code=sector_code)

def calculate_similarity(keywords_query, keywords_workshop):
 if not keywords_query or not keywords_workshop:
    return 0
 query_words = set(keywords_query.lower().split())
 workshop_words = set(keywords_workshop.lower().split())
 common_words = query_words.intersection(workshop_words)
 return len(common_words)

@app.route('/work_shop_search', methods=['GET', 'POST'])
def work_shop_search():
    sector_code = session.get('sector_code')
    search_query = request.args.get('search', '').strip()
    work_shops = Work_shop4.query.all()
   
    if search_query:
         workshops_with_similarity = []
         for workshop in work_shops:
             similarity = calculate_similarity(search_query, workshop.keywords)
             if similarity > 0:
                 workshops_with_similarity.append((workshop, similarity))
                 
         workshops_with_similarity.sort(key=lambda item: item[1], reverse=True)
         work_shops = [item[0] for item in workshops_with_similarity]
    return render_template('work_shop_search.html', work_shops=work_shops, sector_code=sector_code)

@app.route('/delete_workshop/<int:workshop_id>', methods=['POST'])
def delete_workshop(workshop_id):
    workshop = Work_shop4.query.get(workshop_id)
    
    if workshop:
        db.session.delete(workshop)
        db.session.commit()  
        
    return redirect(url_for('work_shop')) 


@app.route('/add_workshop', methods=['POST'])
def add_workshop():
    workshop_data = {
        "name": "Default Workshop",
        "image": None,
        "start_date": datetime.strptime("2024-01-01", "%Y-%m-%d"),
        "end_date": datetime.strptime("2024-12-31", "%Y-%m-%d"),
        "start_time": time(hour=9, minute=0),
        "end_time": time(hour=17, minute=0),
        "type": "Forum", 
        "days_count": 5,
        "person_count": 25,
        "location": "Cairo, Egypt"
    }
    
    workshop = Work_shop4(**workshop_data)
    db.session.add(workshop)
    db.session.commit()
    
    return redirect(url_for('work_shop'))

@app.route('/add_gift', methods=['POST'])
def add_gift():
    gift_data = {
        "name": "Default Gift",
        "image": None,
        "start_date": datetime.strptime("2024-01-01", "%Y-%m-%d"),
        "end_date": datetime.strptime("2024-12-31", "%Y-%m-%d"),
        "start_time": time(hour=9, minute=0),  
        "end_time": time(hour=17, minute=0), 
        "total_gifts": 10,
        "is_internal": True,
        "exam_count": 5,
        "percentage": 75,
        "account_type": "free",
        "practice_count": 4,
        "description": "This is a default description for the gift." 
    }
    gift = Reward2(**gift_data)
    db.session.add(gift)
    db.session.commit()
    return redirect(url_for('gift'))

@app.route('/add_Article/<int:lesson_id>', methods=['POST'])
def add_Article(lesson_id):
    question = Article_question(
        lesson_id=lesson_id,
        question_text="Default question Click to edit",
        difficulty="Easy",
        importance="Necessary",
        bloom_taxonomy="Analyze",
        grade=1,
        times=7,
        
    )
    db.session.add(question)
    db.session.commit()
    
    return redirect(url_for('article', lesson_id=lesson_id))

def calculate_expected_time_article(question):
    expected_time = 0

    # حساب الوقت بناءً على طول النص
    if question.question_text:
        expected_time += len(question.question_text) // 5  # ثانية لكل 5 حروف

    # إضافة 20 ثانية إذا كانت هناك صورة
    if question.photo_path:
        expected_time += 20

    # إضافة وقت الصوت و 5 ثواني إذا كان هناك ملف صوتي
    if question.audio_path:
        try:
            audio_path = os.path.join(app.config['UPLOAD_FOLDER'], question.audio_path)
            if audio_path.lower().endswith(('.mp3')):
                audio = MP3(audio_path)
                audio_duration = audio.info.length
                expected_time += audio_duration + 5
            elif audio_path.lower().endswith(('.wav')):
                audio = WAVE(audio_path)
                audio_duration = audio.info.length
                expected_time += audio_duration + 5
            elif audio_path.lower().endswith(('.ogg')):
                audio = OggVorbis(audio_path)
                audio_duration = audio.info.length
                expected_time += audio_duration + 5
            elif audio_path.lower().endswith(('.aac')):
                audio = AAC(audio_path)
                audio_duration = audio.info.length
                expected_time += audio_duration + 5
            else:
                expected_time += 5

        except Exception as e:
            print(f"Error getting audio duration: {e}")
            expected_time += 5
    return int(expected_time)

@app.route('/Article_question', methods=['GET', 'POST'])
def article_question():
    sector_code = session.get('sector_code')
    question_id = request.args.get('question_id', type=int)
    question = Article_question.query.get_or_404(question_id)
    
    if request.method == 'POST':
        question.question_text = request.form['question_text']
        question.difficulty = request.form['difficulty']
        question.importance = request.form['importance']
        question.bloom_taxonomy = request.form['bloom_taxonomy']
        question.grade = request.form['grade']
        question.times = request.form['times']
        if 'photo' in request.files:
            photo = request.files['photo']
            if photo and allowed_file1(photo.filename):
                if question.photo_path:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.photo_path))
                    except FileNotFoundError:
                        pass

                filename = f"T_F_{question_id}_{secure_filename(photo.filename).rsplit('.', 1)[0]}.png"
                try:
                    photo.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.photo_path = filename
                except Exception as e:
                    print(f"Error saving photo: {e}")

     
        if 'audio' in request.files:
            audio = request.files['audio']
            if audio and allowed_file1(audio.filename):
        
                if question.audio_path:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.audio_path))
                    except FileNotFoundError:
                        pass

                filename = f"T_F_{question_id}_{secure_filename(audio.filename)}"
                try:
                    audio.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.audio_path = filename
                except Exception as e:
                    print(f"Error saving audio: {e}")

        try:
            db.session.commit()
        except Exception as e:
            print(f"Error saving data: {e}")
    expected_time = calculate_expected_time_article(question)    
    return render_template('article_question.html', question=question,sector_code=sector_code,expected_time=expected_time)

@app.route('/edit_question_Multiple_single', methods=['GET', 'POST'])
def edit_question_Multiple_single():
    sector_code = session.get('sector_code')
    lesson_id = request.args.get('lesson_id')
    questions = Question_Multipleone_choice.query.filter_by(lesson_id=lesson_id).all()
    return render_template('edit_question_Multiple_single.html',lesson_id=lesson_id,questions=questions,sector_code=sector_code)   

@app.route('/edit_question_sequence', methods=['GET', 'POST'])
def edit_question_sequence():
    sector_code = session.get('sector_code')
    lesson_id = request.args.get('lesson_id')
    questions = Questionsequence1.query.filter_by(lesson_id=lesson_id).all()
    return render_template('edit_question_sequence.html',lesson_id=lesson_id,questions=questions,sector_code=sector_code)  

@app.route('/add_question_sequence/<int:lesson_id>', methods=['POST'])
def add_question_sequence(lesson_id):
    question = Questionsequence1(
        lesson_id=lesson_id,
        question_name="Default question Click to edit",
        difficulty="Easy",
        importance="Necessary",
        bloom_taxonomy="Analyze",
        grade=1,
        times=7
    )
    db.session.add(question)
    db.session.commit()
    answers = [
        Answer_sequence(
            question_id=question.question_id,
            answer_text=f"Default answer",
            index=i+1
        )
        for i in range(3)  
    ]
    db.session.add_all(answers)
    db.session.commit()

    return redirect(url_for('edit_question_sequence', lesson_id=lesson_id))


@app.route('/edit_matching', methods=['GET', 'POST'])
def edit_matching():
    lesson_id = request.args.get('lesson_id')
    questions = questionmatching1.query.filter_by(lesson_id=lesson_id).all()
    sector_code = session.get('sector_code')
    return render_template('edit_matching.html',lesson_id=lesson_id,questions=questions,sector_code=sector_code)  

@app.route('/edit_classification', methods=['GET', 'POST'])
def edit_classification():
    sector_code = session.get('sector_code')
    lesson_id = request.args.get('lesson_id')
    questions = questionclassiffication.query.filter_by(lesson_id=lesson_id).all()
    return render_template('edit_classification.html',lesson_id=lesson_id,questions=questions,sector_code=sector_code)  

@app.route('/add_classification/<int:lesson_id>', methods=['POST'])
def add_classification(lesson_id):
    question = questionclassiffication(
        lesson_id=lesson_id,
        label_text="Default question Click to edit",
        label_text1="Default answer 1",
        label_text2="Default answer 2",
        label_text3="Default answer 3",
        difficulty="Easy",
        importance="Necessary",
        bloom_taxonomy="Analyze",
        grade=1,
        times=7
    )

    db.session.add(question)
    db.session.commit()

  
    question_id = question.question_id
    for i in range(4):
        answer_text = f"Default answer {i+1}"
        answer = Answer_classiffication1(
            question_id=question_id,
            answer_text=answer_text
        )
        db.session.add(answer)

    db.session.commit()
    return redirect(url_for('edit_classification', lesson_id=lesson_id))
    # Route for classification page
def calculate_expected_time_classification(question, answers):
    expected_time = 0

     # حساب الوقت بناءً على طول النص في السؤال
    if question.label_text:
         expected_time += len(question.label_text) // 5
    if question.label_text1:
         expected_time += len(question.label_text1) // 5
    if question.label_text2:
         expected_time += len(question.label_text2) // 5
    if question.label_text3:
         expected_time += len(question.label_text3) // 5

     # حساب الوقت بناءً على النصوص والوسائط في الإجابات
    if answers:
        for answer in answers:
            if answer.answer_text:
              expected_time += len(answer.answer_text) // 5
              expected_time+=3
           # إضافة 20 ثانية إذا كانت هناك صورة في الإجابة
            #if answer.photo_path:
            #   expected_time += 20

            # إضافة وقت الصوت و 5 ثواني إذا كان هناك ملف صوتي في الإجابة
            #if answer.audio_path:
            #  try:
            #      audio_path = os.path.join(app.config['UPLOAD_FOLDER'], answer.audio_path)
            #      if audio_path.lower().endswith(('.mp3')):
            #          audio = MP3(audio_path)
            #          audio_duration = audio.info.length
            #          expected_time += audio_duration + 5
            #      elif audio_path.lower().endswith(('.wav')):
            #             audio = WAVE(audio_path)
            #             audio_duration = audio.info.length
            #             expected_time += audio_duration + 5
            #      elif audio_path.lower().endswith(('.ogg')):
            #          audio = OggVorbis(audio_path)
            #          audio_duration = audio.info.length
            #          expected_time += audio_duration + 5
            #      elif audio_path.lower().endswith(('.aac')):
            #          audio = AAC(audio_path)
            #          audio_duration = audio.info.length
            #          expected_time += audio_duration + 5
            #      else:
            #          expected_time += 5

            #  except Exception as e:
            #     print(f"Error getting audio duration: {e}")
            #     expected_time += 5
    return int(expected_time+7)
@app.route('/classification/<int:question_id>', methods=['GET', 'POST'])
def classification(question_id):
    sector_code = session.get('sector_code')
    question = questionclassiffication.query.get_or_404(question_id)
    answers = Answer_classiffication1.query.filter_by(question_id=question_id).all()

    if request.method == 'POST':
        question.label_text = request.form.get('label_text_1')
        question.label_text1 = request.form.get('label_text_2')
        question.label_text2 = request.form.get('label_text_3')
        question.label_text3 = request.form.get('label_text_4')
        question.difficulty = request.form.get('difficulty')
        question.importance = request.form.get('importance')
        question.bloom_taxonomy = request.form.get('bloom_taxonomy')
        question.grade = request.form.get('grade')
        question.times = request.form.get('times')
        print(request.form)

        # تحديث الإجابات القديمة
        for answer in answers:
            answer_text = request.form.get(f'answer_text_{answer.answer_id}')
            label = request.form.get(f'label_{answer.answer_id}')
            print(label)
            if answer_text is not None:
                answer.answer_text = answer_text
            if label is not None and label != '':  # هنا يتم التحقق
                answer.label = int(label)

        # معالجة الإجابات الجديدة
        new_answers_text = request.form.getlist('new_answer_text[]')
        new_labels = request.form.getlist('new_label[]')

        for answer_text, label in zip(new_answers_text, new_labels):
            if answer_text and label and label != '':  # هنا يتم التحقق
                new_answer = Answer_classiffication1(question_id=question_id, answer_text=answer_text, label=int(label))
                db.session.add(new_answer)

        db.session.commit()
        return redirect(request.url)

    expected_time = calculate_expected_time_classification(question, answers)
    return render_template('classification.html', question=question, answers=answers, sector_code=sector_code,
                           expected_time=expected_time)
@app.route('/add_matching/<int:lesson_id>', methods=['POST'])
def add_matching(lesson_id):
    question = questionmatching1(
        lesson_id=lesson_id,
        question_text="Default question Click to edit",
        question_text1="Default question Click to edit",
        question_text2="Default question Click to edit",
     
        answer_text="Default answer",
        answer_text1="Default answer",
        answer_text2="Default answer",
        difficulty="Easy",
        importance="Necessary",
        bloom_taxonomy="Analyze",
        grade=1,
        times=7
    )
    db.session.add(question)
    db.session.commit()


    return redirect(url_for('edit_matching', lesson_id=lesson_id))



def get_audio_duration(audio_path):
    try:
        if audio_path.lower().endswith(('.mp3')):
            audio = MP3(audio_path)
            return audio.info.length
        elif audio_path.lower().endswith(('.wav')):
            audio = WAVE(audio_path)
            return audio.info.length
        elif audio_path.lower().endswith(('.ogg')):
            audio = OggVorbis(audio_path)
            return audio.info.length
        elif audio_path.lower().endswith(('.aac')):
            audio = AAC(audio_path)
            return audio.info.length
    except Exception as e:
        print(f"Error getting audio duration: {e}")
    return 0

def calculate_expected_time_matching(question, app):
    expected_time = 0
    time_per_char = 0.2 # ثانية لكل 5 حروف (قابلة للتخصيص)
    image_time = 20      # وقت إضافي للصورة
    audio_base_time = 5    # وقت إضافي للصوت بغض النظر عن المدة

    # حساب الوقت بناءً على طول النص في السؤال الرئيسي
    if question.question_text:
        expected_time += len(question.question_text) * time_per_char

    # إضافة وقت الصورة في السؤال الرئيسي
    if question.photo_path:
        expected_time += image_time

    # إضافة وقت الصوت في السؤال الرئيسي
    if question.audio_path:
        try:
            audio_path = os.path.join(app.config['UPLOAD_FOLDER'], question.audio_path)
            audio_duration = get_audio_duration(audio_path)
            expected_time += audio_duration + audio_base_time
        except Exception as e:
            print(f"Error getting audio duration: {e}")
            expected_time += audio_base_time

    # حساب الوقت بناءً على طول النص والصوت والصورة في الأسئلة الفرعية
    for i in range(1, 4):
        question_text = getattr(question, f'question_text{i}', None)
        photo_path = getattr(question, f'photo_path{i}', None)
        audio_path = getattr(question, f'audio_path{i}', None)
        answer_text = getattr(question, f'answer_text{i}', None)
        answer_photo_path = getattr(question, f'answer_photo_path{i}', None)
        answer_audio_path = getattr(question, f'answer_audio_path{i}', None)

        if question_text:
              expected_time += len(question_text) * time_per_char
        if photo_path:
              expected_time += image_time
        if audio_path:
            try:
                audio_path = os.path.join(app.config['UPLOAD_FOLDER'], audio_path)
                audio_duration = get_audio_duration(audio_path)
                expected_time += audio_duration + audio_base_time
            except Exception as e:
                print(f"Error getting audio duration: {e}")
                expected_time += audio_base_time
        if answer_text:
             expected_time += len(answer_text) * time_per_char
        if answer_photo_path:
               expected_time += image_time
        if answer_audio_path:
            try:
                audio_path = os.path.join(app.config['UPLOAD_FOLDER'], answer_audio_path)
                audio_duration = get_audio_duration(audio_path)
                expected_time += audio_duration + audio_base_time
            except Exception as e:
                print(f"Error getting audio duration: {e}")
                expected_time += audio_base_time


    #حساب الوقت في الاجابه الرئيسيه
    if question.answer_text:
        expected_time += len(question.answer_text) * time_per_char
    if question.answer_photo_path:
        expected_time += image_time
    if question.answer_audio_path:
        try:
            audio_path = os.path.join(app.config['UPLOAD_FOLDER'], question.answer_audio_path)
            audio_duration = get_audio_duration(audio_path)
            expected_time += audio_duration + audio_base_time
        except Exception as e:
            print(f"Error getting audio duration: {e}")
            expected_time += audio_base_time
    return int(expected_time + 7)


@app.route('/matching/<int:question_id>', methods=['GET', 'POST'])
def matching(question_id):
    question = questionmatching1.query.get_or_404(question_id)
    sector_code = session.get('sector_code')
    if request.method == 'POST':
        question.question_text = request.form.get('question_text')
        question.question_text1 = request.form.get('question_text1')
        question.question_text2 = request.form.get('question_text2')
        question.question_text3 = request.form.get('question_text3')
        question.answer_text = request.form.get('answer_text')
        question.answer_text1 = request.form.get('answer_text1')
        question.answer_text2 = request.form.get('answer_text2')
        question.answer_text3 = request.form.get('answer_text3')
        question.difficulty = request.form.get('difficulty')
        question.importance = request.form.get('importance')
        question.bloom_taxonomy = request.form.get('bloom_taxonomy')
        question.grade = request.form.get('grade')
        question.times = request.form.get('times')

        # ملفات السؤال
        if 'photo' in request.files:
            photo = request.files['photo']
            if photo and allowed_file1(photo.filename):
                if question.photo_path:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.photo_path))
                    except FileNotFoundError:
                        pass
                filename = f"M_{question_id}_{secure_filename(photo.filename).rsplit('.', 1)[0]}.png"
                try:
                    photo.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.photo_path = filename
                except Exception as e:
                    print(f"Error saving photo: {e}")

        if 'audio' in request.files:
            audio = request.files['audio']
            if audio and allowed_file1(audio.filename):
                if question.audio_path:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.audio_path))
                    except FileNotFoundError:
                        pass
                filename = f"M_{question_id}_{secure_filename(audio.filename)}"
                try:
                    audio.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.audio_path = filename
                except Exception as e:
                    print(f"Error saving audio: {e}")

        if 'photo1' in request.files:
            photo1 = request.files['photo1']
            if photo1 and allowed_file1(photo1.filename):
                if question.photo_path1:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.photo_path1))
                    except FileNotFoundError:
                        pass
                filename = f"M_{question_id}_{secure_filename(photo1.filename).rsplit('.', 1)[0]}.png"
                try:
                    photo1.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.photo_path1 = filename
                except Exception as e:
                    print(f"Error saving photo1: {e}")

        if 'audio1' in request.files:
            audio1 = request.files['audio1']
            if audio1 and allowed_file1(audio1.filename):
                if question.audio_path1:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.audio_path1))
                    except FileNotFoundError:
                        pass
                filename = f"M_{question_id}_{secure_filename(audio1.filename)}"
                try:
                    audio1.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.audio_path1 = filename
                except Exception as e:
                    print(f"Error saving audio1: {e}")

        if 'photo2' in request.files:
            photo2 = request.files['photo2']
            if photo2 and allowed_file1(photo2.filename):
                if question.photo_path2:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.photo_path2))
                    except FileNotFoundError:
                        pass
                filename = f"M_{question_id}_{secure_filename(photo2.filename).rsplit('.', 1)[0]}.png"
                try:
                    photo2.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.photo_path2 = filename
                except Exception as e:
                    print(f"Error saving photo2: {e}")

        if 'audio2' in request.files:
            audio2 = request.files['audio2']
            if audio2 and allowed_file1(audio2.filename):
                if question.audio_path2:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.audio_path2))
                    except FileNotFoundError:
                        pass
                filename = f"M_{question_id}_{secure_filename(audio2.filename)}"
                try:
                    audio2.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.audio_path2 = filename
                except Exception as e:
                    print(f"Error saving audio2: {e}")

        if 'photo3' in request.files:
            photo3 = request.files['photo3']
            if photo3 and allowed_file1(photo3.filename):
                if question.photo_path3:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.photo_path3))
                    except FileNotFoundError:
                        pass
                filename = f"M_{question_id}_{secure_filename(photo3.filename).rsplit('.', 1)[0]}.png"
                try:
                    photo3.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.photo_path3 = filename
                except Exception as e:
                    print(f"Error saving photo3: {e}")

        if 'audio3' in request.files:
            audio3 = request.files['audio3']
            if audio3 and allowed_file1(audio3.filename):
                if question.audio_path3:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.audio_path3))
                    except FileNotFoundError:
                        pass
                filename = f"M_{question_id}_{secure_filename(audio3.filename)}"
                try:
                    audio3.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.audio_path3 = filename
                except Exception as e:
                    print(f"Error saving audio3: {e}")

       # ملفات الإجابات
        if 'answer_photo' in request.files:
             answer_photo = request.files['answer_photo']
             if answer_photo and allowed_file1(answer_photo.filename):
                 if question.answer_photo_path:
                     try:
                          os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.answer_photo_path))
                     except FileNotFoundError:
                          pass
                 filename = f"M_{question_id}_answer_{secure_filename(answer_photo.filename).rsplit('.', 1)[0]}.png"
                 try:
                     answer_photo.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                     question.answer_photo_path = filename
                 except Exception as e:
                        print(f"Error saving answer_photo: {e}")

        if 'answer_audio' in request.files:
            answer_audio = request.files['answer_audio']
            if answer_audio and allowed_file1(answer_audio.filename):
                if question.answer_audio_path:
                    try:
                         os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.answer_audio_path))
                    except FileNotFoundError:
                         pass
                filename = f"M_{question_id}_answer_{secure_filename(answer_audio.filename)}"
                try:
                    answer_audio.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.answer_audio_path = filename
                except Exception as e:
                    print(f"Error saving answer_audio: {e}")

        if 'answer_photo1' in request.files:
             answer_photo1 = request.files['answer_photo1']
             if answer_photo1 and allowed_file1(answer_photo1.filename):
                 if question.answer_photo_path1:
                     try:
                          os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.answer_photo_path1))
                     except FileNotFoundError:
                          pass
                 filename = f"M_{question_id}_answer1_{secure_filename(answer_photo1.filename).rsplit('.', 1)[0]}.png"
                 try:
                     answer_photo1.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                     question.answer_photo_path1 = filename
                 except Exception as e:
                        print(f"Error saving answer_photo1: {e}")

        if 'answer_audio1' in request.files:
            answer_audio1 = request.files['answer_audio1']
            if answer_audio1 and allowed_file1(answer_audio1.filename):
                if question.answer_audio_path1:
                    try:
                         os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.answer_audio_path1))
                    except FileNotFoundError:
                         pass
                filename = f"M_{question_id}_answer1_{secure_filename(answer_audio1.filename)}"
                try:
                    answer_audio1.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.answer_audio_path1 = filename
                except Exception as e:
                    print(f"Error saving answer_audio1: {e}")

        if 'answer_photo2' in request.files:
             answer_photo2 = request.files['answer_photo2']
             if answer_photo2 and allowed_file1(answer_photo2.filename):
                 if question.answer_photo_path2:
                     try:
                          os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.answer_photo_path2))
                     except FileNotFoundError:
                          pass
                 filename = f"M_{question_id}_answer2_{secure_filename(answer_photo2.filename).rsplit('.', 1)[0]}.png"
                 try:
                     answer_photo2.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                     question.answer_photo_path2 = filename
                 except Exception as e:
                        print(f"Error saving answer_photo2: {e}")

        if 'answer_audio2' in request.files:
            answer_audio2 = request.files['answer_audio2']
            if answer_audio2 and allowed_file1(answer_audio2.filename):
                if question.answer_audio_path2:
                    try:
                         os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.answer_audio_path2))
                    except FileNotFoundError:
                         pass
                filename = f"M_{question_id}_answer2_{secure_filename(answer_audio2.filename)}"
                try:
                    answer_audio2.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.answer_audio_path2 = filename
                except Exception as e:
                    print(f"Error saving answer_audio2: {e}")

        if 'answer_photo3' in request.files:
             answer_photo3 = request.files['answer_photo3']
             if answer_photo3 and allowed_file1(answer_photo3.filename):
                 if question.answer_photo_path3:
                     try:
                          os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.answer_photo_path3))
                     except FileNotFoundError:
                          pass
                 filename = f"M_{question_id}_answer3_{secure_filename(answer_photo3.filename).rsplit('.', 1)[0]}.png"
                 try:
                     answer_photo3.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                     question.answer_photo_path3 = filename
                 except Exception as e:
                        print(f"Error saving answer_photo3: {e}")

        if 'answer_audio3' in request.files:
            answer_audio3 = request.files['answer_audio3']
            if answer_audio3 and allowed_file1(answer_audio3.filename):
                if question.answer_audio_path3:
                    try:
                         os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.answer_audio_path3))
                    except FileNotFoundError:
                         pass
                filename = f"M_{question_id}_answer3_{secure_filename(answer_audio3.filename)}"
                try:
                    answer_audio3.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.answer_audio_path3 = filename
                except Exception as e:
                    print(f"Error saving answer_audio3: {e}")
        

        db.session.commit()
        return redirect(url_for('matching', question_id=question_id))
    expected_time = calculate_expected_time_matching(question, app)
    print(expected_time)
    return render_template('matching.html', question=question, sector_code=sector_code,expected_time=expected_time)

@app.route('/delete_classification/<int:question_id>', methods=['GET', 'POST'])
def delete_classification(question_id):

    question = questionclassiffication.query.get_or_404(question_id)
    answers = Answer_classiffication1.query.filter_by(question_id=question_id).all()
    lesson_id = question.lesson_id 

    if request.method == 'POST':

        for answer in answers:
            db.session.delete(answer)

        db.session.delete(question)
        db.session.commit()


    return redirect(url_for('edit_classification', lesson_id=lesson_id))

@app.route('/delete_matching/<int:question_id>', methods=['POST'])
def delete_matching(question_id):
    question = questionmatching1.query.get_or_404(question_id)
    lesson_id = question.lesson_id 
    def delete_file(file_path):
        if file_path:
            try:
                os.remove(os.path.join(app.config['UPLOAD_FOLDER'], file_path))
            except FileNotFoundError:
                pass

    delete_file(question.photo_path)
    delete_file(question.audio_path)
    delete_file(question.photo_path1)
    delete_file(question.audio_path1)
    delete_file(question.photo_path2)
    delete_file(question.audio_path2)
    delete_file(question.photo_path3)
    delete_file(question.audio_path3)
    question.photo_path = None
    question.audio_path = None
    question.photo_path1 = None
    question.audio_path1 = None
    question.photo_path2 = None
    question.audio_path2 = None
    question.photo_path3 = None
    question.audio_path3 = None
    db.session.delete(question)

    db.session.commit()

    return redirect(url_for('edit_matching', lesson_id=lesson_id))

def calculate_expected_time_sequence(question, answers):
    expected_time = 0

    # حساب الوقت بناءً على طول النص في السؤال
    if question.question_name:
        expected_time += len(question.question_name) // 5  # ثانية لكل 5 حروف

    # إضافة 20 ثانية إذا كانت هناك صورة في السؤال
    if question.photo_path:
        expected_time += 20

    # إضافة وقت الصوت و 5 ثواني إذا كان هناك ملف صوتي في السؤال
    if question.audio_path:
       try:
            audio_path = os.path.join(app.config['UPLOAD_FOLDER'], question.audio_path)
            if audio_path.lower().endswith(('.mp3')):
                audio = MP3(audio_path)
                audio_duration = audio.info.length
                expected_time += audio_duration + 5
            elif audio_path.lower().endswith(('.wav')):
                audio = WAVE(audio_path)
                audio_duration = audio.info.length
                expected_time += audio_duration + 5
            elif audio_path.lower().endswith(('.ogg')):
                audio = OggVorbis(audio_path)
                audio_duration = audio.info.length
                expected_time += audio_duration + 5
            elif audio_path.lower().endswith(('.aac')):
                audio = AAC(audio_path)
                audio_duration = audio.info.length
                expected_time += audio_duration + 5
            else:
                  expected_time += 5
       except Exception as e:
           print(f"Error getting audio duration: {e}")
           expected_time += 5

    if answers:
        expected_time += len(answers) * 5


    return int(expected_time+12)


@app.route('/sequence/<int:question_id>', methods=['GET', 'POST'])
def sequence(question_id):
    sector_code = session.get('sector_code')
    question = Questionsequence1.query.get(question_id)
    answers = Answer_sequence.query.filter_by(question_id=question_id).all()
    
    if request.method == 'GET':
        answers = Answer_sequence.query.filter_by(question_id=question_id).all()
        expected_time = calculate_expected_time_sequence(question, answers)
        return render_template('sequence.html', question=question, answers=answers,sector_code=sector_code,expected_time=expected_time)

    if request.method == 'POST':
        question.question_name = request.form.get('question_text')
        question.difficulty = request.form.get('difficulty')
        question.importance = request.form.get('importance')
        question.bloom_taxonomy = request.form.get('bloom_taxonomy')
        question.grade = request.form.get('grade')
        question.times = request.form.get('times')

        if 'photo' in request.files:
            photo = request.files['photo']
            if photo and allowed_file1(photo.filename):
                if question.photo_path:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.photo_path))
                    except FileNotFoundError:
                        pass
                filename = f"s_{question_id}_{secure_filename(photo.filename).rsplit('.', 1)[0]}.png"
                try:
                    photo.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.photo_path = filename
                except Exception as e:
                    print(f"Error saving photo: {e}")

        # معالجة الصوت
        if 'audio' in request.files:
            audio = request.files['audio']
            if audio and allowed_file1(audio.filename):
                if question.audio_path:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.audio_path))
                    except FileNotFoundError:
                        pass
                filename = f"s_{question_id}_{secure_filename(audio.filename)}"
                try:
                    audio.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.audio_path = filename
                except Exception as e:
                    print(f"Error saving audio: {e}")

        answers = Answer_sequence.query.filter_by(question_id=question_id).all()
        for i, answer in enumerate(answers):
            answer.answer_text = request.form.get(f'answer_text_{i}')
            order_value = request.form.get(f'order_{i}')
            if order_value is not None and order_value.isdigit():
                answer.index = int(order_value)
            else:
                answer.index = 1

        new_answer_index = len(answers) 
        while True:
            answer_text = request.form.get(f'answer_text_{new_answer_index}')
            order_value = request.form.get(f'order_{new_answer_index}')
            
            if answer_text is None or order_value is None:
                break 
            
          
            new_answer = Answer_sequence(question_id=question_id, answer_text=answer_text)
            if order_value.isdigit():
                 new_answer.index = int(order_value)
            else:
                 new_answer.index = 1
            db.session.add(new_answer)
            
            new_answer_index += 1
            
        db.session.commit()
        return redirect(url_for('sequence', question_id=question_id))
    expected_time = calculate_expected_time_sequence(question, answers)
    print(expected_time)

@app.route('/delete_article_question/<int:question_id>', methods=['POST'])
def delete_article_question(question_id):
    # جلب السؤال من قاعدة البيانات
    question = Article_question.query.get_or_404(question_id)
    lesson_id = question.lesson_id 

    try:
        # حذف الملفات المرتبطة بالسؤال (إذا كانت موجودة)
        if question.photo_path:
            try:
                os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.photo_path))
            except FileNotFoundError:
                pass
        
        if question.audio_path:
            try:
                os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.audio_path))
            except FileNotFoundError:
                pass

        db.session.delete(question)
        db.session.commit()

    except Exception as e:
        db.session.rollback()
       
    return redirect(url_for('article',lesson_id=lesson_id))  


@app.route('/delete_sequence/<int:question_id>', methods=['POST'])
def delete_sequence(question_id):
    question = Questionsequence1.query.get_or_404(question_id)
    lesson_id = question.lesson_id 

    try:

        if question.photo_path:
            try:
                os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.photo_path))
            except FileNotFoundError:
                pass
        if question.audio_path:
            try:
                os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.audio_path))
            except FileNotFoundError:
                pass

        Answer_sequence.query.filter_by(question_id=question_id).delete()

        db.session.delete(question)
        db.session.commit()

    except Exception as e:
        db.session.rollback()

    return redirect(url_for('edit_question_sequence',lesson_id=lesson_id))  

@app.route('/add_question_Multiple_single_choice/<int:lesson_id>', methods=['POST'])
def add_question_Multiple_single_choice(lesson_id):
    question = Question_Multipleone_choice(
        lesson_id=lesson_id,
        question_name="Default question Click to edit ",
        difficulty="Easy",
        importance="Necessary",
        bloom_taxonomy="Analyze",
        grade=1,
        times=7
    )
    db.session.add(question)
    db.session.commit()
    answers = [
        Answer_Multiple_Multiple_choice(
            question_id=question.question_id,
            is_correct=(i == 0)  
        )
        for i in range(8)  
    ]
    db.session.add_all(answers)
    db.session.commit()

    return redirect(url_for('edit_question_Multiple_single', lesson_id=lesson_id))




@app.route('/mcq_multiple_single/<int:question_id>', methods=['GET', 'POST'])
def mcq_multiple_single(question_id):
    sector_code = session.get('sector_code')
    question = Question_Multipleone_choice.query.get(question_id)
    answers = Answer_Multiple_Multiple_choice.query.filter_by(question_id=question_id).all()
    if request.method == 'GET':
        answers = Answer_Multiple_Multiple_choice.query.filter_by(question_id=question_id).all()
        expected_time = calculate_expected_time_MCQ(question, answers)
        print(expected_time)

        return render_template('mcq_multiple_single.html', question=question, answers=answers,sector_code=sector_code,expected_time=expected_time)
    
    if request.method == 'POST':
        expected_time = calculate_expected_time_MCQ(question, answers)
        print(expected_time)
        question.question_name = request.form.get('question_text')
        question.difficulty = request.form.get('difficulty')
        question.importance = request.form.get('importance')
        question.bloom_taxonomy = request.form.get('bloom_taxonomy')
        question.grade = request.form.get('grade')
        question.times = request.form.get('times')

        answers = Answer_Multiple_Multiple_choice.query.filter_by(question_id=question_id).all()
        correct_index = int(request.form.get('correct'))
        for i, answer in enumerate(answers):
            answer.answer_text = request.form.get(f'answer_text_{i}')
            answer.is_correct = (i == correct_index)  
            if 'photo_' + str(i) in request.files:
                photo = request.files['photo_' + str(i)]
                if photo and allowed_file1(photo.filename):
                   
                    if answer.photo_path:
                        try:
                            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], answer.photo_path))
                        except FileNotFoundError:
                            pass

                    filename = f"M_S_{question_id}_answer_{i}_{secure_filename(photo.filename).rsplit('.', 1)[0]}.png"
                    try:
                        photo.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                        answer.photo_path = filename
                    except Exception as e:
                        print(f"Error saving photo: {e}")

         
            if 'audio_' + str(i) in request.files:
                audio = request.files['audio_' + str(i)]
                if audio and allowed_file1(audio.filename):
                  
                    if answer.audio_path:
                        try:
                            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], answer.audio_path))
                        except FileNotFoundError:
                            pass

                   
                    filename = f"M_S_{question_id}_answer_{i}_{secure_filename(audio.filename)}"
                    try:
                        audio.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                        answer.audio_path = filename
                    except Exception as e:
                        print(f"Error saving audio: {e}")


        if 'photo' in request.files:
            photo = request.files['photo']
            if photo and allowed_file1(photo.filename):
      
                if question.photo_path:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.photo_path))
                    except FileNotFoundError:
                        pass

                filename = f"M_S_{question_id}_{secure_filename(photo.filename).rsplit('.', 1)[0]}.png"
                try:
                    photo.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.photo_path = filename
                except Exception as e:
                    print(f"Error saving photo: {e}")

        # معالجة رفع الصوت
        if 'audio' in request.files:
            audio = request.files['audio']
            if audio and allowed_file1(audio.filename):
          
                if question.audio_path:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.audio_path))
                    except FileNotFoundError:
                        pass

                
                filename = f"M_S_{question_id}_{secure_filename(audio.filename)}"
                try:
                    audio.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.audio_path = filename
                except Exception as e:
                    print(f"Error saving audio: {e}")


        db.session.commit()
        expected_time = calculate_expected_time_MCQ(question, answers)
        print(expected_time)
        return render_template('mcq_multiple_single.html', question=question, answers=answers,sector_code=sector_code,expected_time=expected_time)

def calculate_expected_time_MCQ(question, answers):
    expected_time = 0

    # حساب الوقت بناءً على طول النص في السؤال
    if question.question_name:
        expected_time += len(question.question_name) // 5  # ثانية لكل 5 حروف

    # إضافة 20 ثانية إذا كانت هناك صورة في السؤال
    if question.photo_path:
        expected_time += 20

    # إضافة وقت الصوت و 5 ثواني إذا كان هناك ملف صوتي في السؤال
    if question.audio_path:
         try:
            audio_path = os.path.join(app.config['UPLOAD_FOLDER'], question.audio_path)
            if audio_path.lower().endswith(('.mp3')):
                  audio = MP3(audio_path)
                  audio_duration = audio.info.length
                  expected_time += audio_duration + 5
            elif audio_path.lower().endswith(('.wav')):
                audio = WAVE(audio_path)
                audio_duration = audio.info.length
                expected_time += audio_duration + 5
            elif audio_path.lower().endswith(('.ogg')):
                  audio = OggVorbis(audio_path)
                  audio_duration = audio.info.length
                  expected_time += audio_duration + 5
            elif audio_path.lower().endswith(('.aac')):
                  audio = AAC(audio_path)
                  audio_duration = audio.info.length
                  expected_time += audio_duration + 5
            else:
                  expected_time += 5
         except Exception as e:
            print(f"Error getting audio duration: {e}")
            expected_time += 5


    # حساب الوقت بناءً على الصور والملفات الصوتية في الإجابات
    if answers:
       for answer in answers:
            if answer.answer_text:
               expected_time += len(answer.answer_text) // 5  # ثانية لكل 5 حروف
             # إضافة 20 ثانية إذا كانت هناك صورة في الإجابة
            if answer.photo_path:
                 expected_time += 20

            # إضافة وقت الصوت و 5 ثواني إذا كان هناك ملف صوتي في الإجابة
            if answer.audio_path:
               try:
                  audio_path = os.path.join(app.config['UPLOAD_FOLDER'], answer.audio_path)
                  if audio_path.lower().endswith(('.mp3')):
                      audio = MP3(audio_path)
                      audio_duration = audio.info.length
                      expected_time += audio_duration + 5
                  elif audio_path.lower().endswith(('.wav')):
                       audio = WAVE(audio_path)
                       audio_duration = audio.info.length
                       expected_time += audio_duration + 5
                  elif audio_path.lower().endswith(('.ogg')):
                         audio = OggVorbis(audio_path)
                         audio_duration = audio.info.length
                         expected_time += audio_duration + 5
                  elif audio_path.lower().endswith(('.aac')):
                         audio = AAC(audio_path)
                         audio_duration = audio.info.length
                         expected_time += audio_duration + 5
                  else:
                       expected_time += 5

               except Exception as e:
                   print(f"Error getting audio duration: {e}")
                   expected_time += 5
    return int(expected_time+7)

@app.route('/mcq_multiple_ans/<int:question_id>', methods=['GET', 'POST'])
def mcq_multiple_ans(question_id):
    sector_code = session.get('sector_code')
    question = QuestionMultipleMultiple_choice.query.get(question_id)
    answers = Answer_Multiple_Multiple_choice.query.filter_by(question_id=question_id).all()
    
    if request.method == 'GET':
        answers = Answer_Multiple_Multiple_choice.query.filter_by(question_id=question_id).all()
        expected_time = calculate_expected_time_MCQ(question, answers)

        return render_template('mcq_multiple_ans.html', question=question, answers=answers,sector_code=sector_code,expected_time=expected_time)
    if request.method == 'POST':
        question.question_name = request.form.get('question_text')
        question.difficulty = request.form.get('difficulty')
        question.importance = request.form.get('importance')
        question.bloom_taxonomy = request.form.get('bloom_taxonomy')
        question.grade = request.form.get('grade')
        question.times = request.form.get('times')

        answers = Answer_Multiple_Multiple_choice.query.filter_by(question_id=question_id).all()
        for i, answer in enumerate(answers):
            answer.answer_text = request.form.get(f'answer_text_{i}')
            answer.is_correct = f'correct_{i}' in request.form
            if 'photo_' + str(i) in request.files:
                photo = request.files['photo_' + str(i)]
                if photo and allowed_file1(photo.filename):
                   
                    if answer.photo_path:
                        try:
                            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], answer.photo_path))
                        except FileNotFoundError:
                            pass

                    filename = f"M_M_{question_id}_answer_{i}_{secure_filename(photo.filename).rsplit('.', 1)[0]}.png"
                    try:
                        photo.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                        answer.photo_path = filename
                    except Exception as e:
                        print(f"Error saving photo: {e}")

         
            if 'audio_' + str(i) in request.files:
                audio = request.files['audio_' + str(i)]
                if audio and allowed_file1(audio.filename):
                  
                    if answer.audio_path:
                        try:
                            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], answer.audio_path))
                        except FileNotFoundError:
                            pass

                    filename = f"M_M_{question_id}_answer_{i}_{secure_filename(audio.filename)}"
                    try:
                        audio.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                        answer.audio_path = filename
                    except Exception as e:
                        print(f"Error saving audio: {e}")


        if 'photo' in request.files:
            photo = request.files['photo']
            if photo and allowed_file1(photo.filename):
                # حذف الصورة القديمة إذا كانت موجودة
                if question.photo_path:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.photo_path))
                    except FileNotFoundError:
                        pass

                # تحديد الاسم الجديد للصورة مع تغيير الامتداد إلى .png
                filename = f"M_M_{question_id}_{secure_filename(photo.filename).rsplit('.', 1)[0]}.png"
                try:
                    photo.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.photo_path = filename
                except Exception as e:
                    print(f"Error saving photo: {e}")
        
        # معالجة رفع الصوت
        if 'audio' in request.files:
            audio = request.files['audio']
            if audio and allowed_file1(audio.filename):
                
                if question.audio_path:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.audio_path))
                    except FileNotFoundError:
                        pass

                filename = f"M_M_{question_id}_{secure_filename(audio.filename)}"
                try:
                    audio.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.audio_path = filename
                except Exception as e:
                    print(f"Error saving audio: {e}")
        db.session.commit()
        expected_time = calculate_expected_time_MCQ(question, answers)
        print(expected_time)
        return render_template('mcq_multiple_ans.html', question=question, answers=answers,sector_code=sector_code,expected_time=expected_time)
@app.route('/add_question_Multiple_Multiple_choice/<int:lesson_id>', methods=['POST'])
def add_question_Multiple_Multiple_choice(lesson_id):
    question = QuestionMultipleMultiple_choice(
        lesson_id=lesson_id,
        question_name="Default question Click to edit",
        difficulty="Easy",
        importance="Necessary",
        bloom_taxonomy="Analyze",
        grade=1,
        times=7
    )
    db.session.add(question)
    db.session.commit()
    answers = [
        Answer_Multiple_Multiple_choice(
            question_id=question.question_id,
            is_correct=(i == 0)  
        )
        for i in range(8)  
    ]
    db.session.add_all(answers)
    db.session.commit()

    return redirect(url_for('edit_question_Multiple_Multiple', lesson_id=lesson_id))


@app.route('/edit_exams', methods=['GET', 'POST'])
def edit_exams():
    exams = Exam.query.all()
    return render_template('edit_exams.html',exams=exams) 

ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
ALLOWED_AUDIO_EXTENSIONS = {'mp3', 'wav'}

def allowed_file2(filename, allowed_extensions):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions

@app.route('/Word_Puzzle', methods=['GET', 'POST'])
def Word_Puzzle():
    sector_code = session.get('sector_code')
    lesson_id = request.args.get('lesson_id', type=int)
    employee_id = session.get('employee_id')
    company_code = session.get('sector_code')

    if request.method == 'POST':
        difficulty = request.form.get('col1')
        importance = request.form.get('importance')
        bloom_taxonomy = request.form.get('bloom_taxonomy')
        grade = request.form.get('grade')
        times = int(request.form.get('times', 1))
        word_puzzle = request.form.get('Word_Puzzle')
        image_file = request.files.get('Word_Puzzle_image')
        audio_file = request.files.get('Word_Puzzle_audio')

        # معالجة رفع الصورة
        image_path = None
        if image_file and allowed_file2(image_file.filename, ALLOWED_IMAGE_EXTENSIONS):
            existing_word = WordPuzzleGame2.query.filter_by(lesson_id=lesson_id, words=word_puzzle).first()
            if existing_word and existing_word.image_path:
                try:
                    os.remove(os.path.join(app.config['UPLOAD_FOLDER'], existing_word.image_path))
                except FileNotFoundError:
                    pass
            
            filename = f"WordPuzzle_{lesson_id}_{secure_filename(image_file.filename).rsplit('.', 1)[0]}.png"
            try:
                image_file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                image_path = filename
            except Exception as e:
                print(f"Error saving image: {e}")
        audio_path = None
        if audio_file and allowed_file2(audio_file.filename, ALLOWED_AUDIO_EXTENSIONS):

            # حذف الملف الصوتي القديم إذا كان موجودًا
            if existing_word and existing_word.audio_path:
                try:
                    os.remove(os.path.join(app.config['UPLOAD_FOLDER'], existing_word.audio_path))
                except FileNotFoundError:
                    pass

            # إنشاء اسم جديد للملف الصوتي يحتوي على اسم الملف الأصلي
            original_filename = secure_filename(audio_file.filename).rsplit('.', 1)[0]  # اسم الملف بدون الامتداد
            filename = f"WordPuzzle_{lesson_id}_{original_filename}.mp3"  # إضافة اسم الملف الأصلي

            try:
                audio_file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                audio_path = filename
            except Exception as e:
                print(f"Error saving audio: {e}")

        # إدخال البيانات في قاعدة البيانات
        if word_puzzle:
            try:
                print(image_path)
                print(audio_path)
                if not existing_word:
                    new_word = WordPuzzleGame2(
                        lesson_id=lesson_id,
                        user_id=employee_id,
                        company_code=company_code,
                        difficulty=difficulty,
                        importance=importance,
                        bloom_taxonomy=bloom_taxonomy,
                        grade=grade,
                        times=times,
                        words=word_puzzle,
                        image_path=image_path,
                        audio_path=audio_path
                    )
                    db.session.add(new_word)
                else:
                    # تحديث البيانات الحالية
                    existing_word.difficulty = difficulty
                    existing_word.importance = importance
                    existing_word.bloom_taxonomy = bloom_taxonomy
                    existing_word.grade = grade
                    existing_word.times = times
                    existing_word.image_path = image_path or existing_word.image_path
                    existing_word.audio_path = audio_path or existing_word.audio_path
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")

    games = WordPuzzleGame2.query.filter_by(lesson_id=lesson_id).all()
    last_difficulty = WordPuzzleGame2.query.filter_by(lesson_id=lesson_id).order_by(WordPuzzleGame2.id.desc()).first()
    current_difficulty = last_difficulty.difficulty if last_difficulty else 'Easy' 
    current_importance = last_difficulty.importance if last_difficulty else 'Necessary'
    current_bloom_taxonomy = last_difficulty.bloom_taxonomy if last_difficulty else 'Analyze'
    current_grade = last_difficulty.grade if last_difficulty else 1
    current_times = last_difficulty.times if last_difficulty else 7

    return render_template(
        'games/Word_Puzzle.html',
        games=games,
        lesson_id=lesson_id,
        current_difficulty=current_difficulty,
        current_importance=current_importance,
        current_bloom_taxonomy=current_bloom_taxonomy,
        current_grade=current_grade,
        current_times=current_times,
        sector_code=sector_code
    )

@app.route('/match', methods=['GET', 'POST'])
def match():
    sector_code = session.get('sector_code')
    lesson_id = request.args.get('lesson_id', type=int)
    employee_id = session.get('employee_id')
    company_code = session.get('sector_code')

    if request.method == 'POST':
        # استلام القيم من النموذج
        difficulty = request.form.get('col1')
        importance = request.form.get('importance')
        bloom_taxonomy = request.form.get('bloom_taxonomy')
        grade = request.form.get('grade')
        times = int(request.form.get('times', 1))  # افتراض 1 إذا لم يتم تقديمه
        Word_Puzzle = request.form.get('Word_Puzzle')
        uploaded_file = request.files.get('Word_Puzzle_file')

        # إضافة الكلمة إلى قاعدة البيانات
        if Word_Puzzle:
            try:
                existing_word = MatchGame.query.filter_by(lesson_id=lesson_id, words=Word_Puzzle).first()
                if not existing_word:
                    new_word = MatchGame(
                        lesson_id=lesson_id,
                        user_id=employee_id,
                        company_code=company_code,
                        difficulty=difficulty,
                        importance=importance,
                        bloom_taxonomy=bloom_taxonomy,
                        grade=grade,
                        times=times,
                        words=Word_Puzzle
                    )
                    db.session.add(new_word)
                    db.session.commit()
                else:
                    print("Word already exists for this exam.")
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")

        # معالجة الملفات المرفوعة (إذا كانت موجودة)
        if uploaded_file and uploaded_file.filename != '':
            filename = secure_filename(uploaded_file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            uploaded_file.save(filepath)

            try:
                workbook = load_workbook(filepath)
                sheet = workbook.active
                for row in sheet.iter_rows(min_row=2):
                    word = row[0].value
                    if word: 
                        existing_word = MatchGame.query.filter_by(
                            lesson_id=lesson_id, words=word
                        ).first()
                        if not existing_word:
                            new_word = MatchGame(
                                lesson_id=lesson_id,
                                user_id=employee_id,
                                company_code=company_code,
                                difficulty=difficulty,
                                importance=importance,
                                bloom_taxonomy=bloom_taxonomy,
                                grade=grade,
                                times=times,
                                words=word
                            )
                            db.session.add(new_word)
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")
            finally:
                os.remove(filepath)

    games = MatchGame.query.filter_by(lesson_id=lesson_id).all()
    last_difficulty = MatchGame.query.filter_by(lesson_id=lesson_id).order_by(MatchGame.id.desc()).first()
    current_difficulty = last_difficulty.difficulty if last_difficulty else 'Easy' 
    current_importance = last_difficulty.importance if last_difficulty else 'Necessary'
    current_bloom_taxonomy = last_difficulty.bloom_taxonomy if last_difficulty else 'Analyze'
    current_grade = last_difficulty.grade if last_difficulty else 1
    current_times = last_difficulty.times if last_difficulty else 7


    return render_template('games/match.html',games=games, lesson_id=lesson_id, current_difficulty=current_difficulty, current_importance=current_importance, current_bloom_taxonomy=current_bloom_taxonomy, current_grade=current_grade, current_times=current_times,sector_code=sector_code)

@app.route('/reverse', methods=['GET', 'POST'])
def reverse():
    sector_code = session.get('sector_code')
    lesson_id = request.args.get('lesson_id', type=int)
    employee_id = session.get('employee_id')
    company_code = session.get('sector_code')

    if request.method == 'POST':
        difficulty = request.form.get('col1')
        importance = request.form.get('importance')
        bloom_taxonomy = request.form.get('bloom_taxonomy')
        grade = request.form.get('grade')
        times = int(request.form.get('times', 1))
        word_puzzle = request.form.get('Word_Puzzle')
        image_file = request.files.get('Word_Puzzle_image')
        audio_file = request.files.get('Word_Puzzle_audio')

        image_path = None
        if image_file and allowed_file2(image_file.filename, ALLOWED_IMAGE_EXTENSIONS):
            existing_word = ReverseGame1.query.filter_by(lesson_id=lesson_id, words=word_puzzle).first()
            if existing_word and existing_word.image_path:
                try:
                    os.remove(os.path.join(app.config['UPLOAD_FOLDER'], existing_word.image_path))
                except FileNotFoundError:
                    pass
            
            filename = f"WordPuzzle_{lesson_id}_{secure_filename(image_file.filename).rsplit('.', 1)[0]}.png"
            try:
                image_file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                image_path = filename
            except Exception as e:
                print(f"Error saving image: {e}")
        audio_path = None
        if audio_file and allowed_file2(audio_file.filename, ALLOWED_AUDIO_EXTENSIONS):

            if existing_word and existing_word.audio_path:
                try:
                    os.remove(os.path.join(app.config['UPLOAD_FOLDER'], existing_word.audio_path))
                except FileNotFoundError:
                    pass

            original_filename = secure_filename(audio_file.filename).rsplit('.', 1)[0]  # اسم الملف بدون الامتداد
            filename = f"WordPuzzle_{lesson_id}_{original_filename}.mp3"  # إضافة اسم الملف الأصلي

            try:
                audio_file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                audio_path = filename
            except Exception as e:
                print(f"Error saving audio: {e}")

        # إدخال البيانات في قاعدة البيانات
        if word_puzzle:
            try:
                print(image_path)
                print(audio_path)
                if not existing_word:
                    new_word = ReverseGame1(
                        lesson_id=lesson_id,
                        user_id=employee_id,
                        company_code=company_code,
                        difficulty=difficulty,
                        importance=importance,
                        bloom_taxonomy=bloom_taxonomy,
                        grade=grade,
                        times=times,
                        words=word_puzzle,
                        image_path=image_path,
                        audio_path=audio_path
                    )
                    db.session.add(new_word)
                else:
                    # تحديث البيانات الحالية
                    existing_word.difficulty = difficulty
                    existing_word.importance = importance
                    existing_word.bloom_taxonomy = bloom_taxonomy
                    existing_word.grade = grade
                    existing_word.times = times
                    existing_word.image_path = image_path or existing_word.image_path
                    existing_word.audio_path = audio_path or existing_word.audio_path
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")

    games = ReverseGame1.query.filter_by(lesson_id=lesson_id).all()
    last_difficulty = ReverseGame1.query.filter_by(lesson_id=lesson_id).order_by(ReverseGame1.id.desc()).first()
    current_difficulty = last_difficulty.difficulty if last_difficulty else 'Easy' 
    current_importance = last_difficulty.importance if last_difficulty else 'Necessary'
    current_bloom_taxonomy = last_difficulty.bloom_taxonomy if last_difficulty else 'Analyze'
    current_grade = last_difficulty.grade if last_difficulty else 1
    current_times = last_difficulty.times if last_difficulty else 7

    return render_template('games/Reverse.html',sector_code=sector_code,games=games, lesson_id=lesson_id, current_difficulty=current_difficulty, current_importance=current_importance, current_bloom_taxonomy=current_bloom_taxonomy, current_grade=current_grade, current_times=current_times)


@app.route('/quess', methods=['GET', 'POST'])
def quess():
    lesson_id = request.args.get('lesson_id', type=int)
    employee_id = session.get('employee_id')
    company_code = session.get('sector_code')
    sector_code = session.get('sector_code')

    if request.method == 'POST':
     
        difficulty = request.form.get('col1')
        importance = request.form.get('importance')
        bloom_taxonomy = request.form.get('bloom_taxonomy')
        grade = request.form.get('grade')
        times = int(request.form.get('times', 1))  
        Word_Puzzle = request.form.get('Word_Puzzle')
        translate = request.form.get('translate')
        uploaded_file = request.files.get('Word_Puzzle_file')

        if Word_Puzzle:
            try:
                existing_word = Quessgame.query.filter_by(lesson_id=lesson_id, words=Word_Puzzle).first()
                if not existing_word:
                    new_word = Quessgame(
                        lesson_id=lesson_id,
                        user_id=employee_id,
                        company_code=company_code,
                        difficulty=difficulty,
                        importance=importance,
                        bloom_taxonomy=bloom_taxonomy,
                        grade=grade,
                        times=times,
                        words=Word_Puzzle,
                        translate=translate
                    )
                    db.session.add(new_word)
                    db.session.commit()
                else:
                    print("Word already exists for this exam.")
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")

        if uploaded_file and uploaded_file.filename != '':
            filename = secure_filename(uploaded_file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            uploaded_file.save(filepath)

            try:
                workbook = load_workbook(filepath)
                sheet = workbook.active
                for row in sheet.iter_rows(min_row=2):
                    word = row[0].value
                    if word: 
                        existing_word = Quessgame.query.filter_by(
                            lesson_id=lesson_id, words=word
                        ).first()
                        if not existing_word:
                            new_word = Quessgame(
                                lesson_id=lesson_id,
                                user_id=employee_id,
                                company_code=company_code,
                                difficulty=difficulty,
                                importance=importance,
                                bloom_taxonomy=bloom_taxonomy,
                                grade=grade,
                                times=times,
                                words=word,
                                translate=translate
                            )
                            db.session.add(new_word)
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")
            finally:
                os.remove(filepath)

    games = Quessgame.query.filter_by(lesson_id=lesson_id).all()
    last_difficulty = Quessgame.query.filter_by(lesson_id=lesson_id).order_by(Quessgame.id.desc()).first()
    current_difficulty = last_difficulty.difficulty if last_difficulty else 'Easy' 
    current_importance = last_difficulty.importance if last_difficulty else 'Necessary'
    current_bloom_taxonomy = last_difficulty.bloom_taxonomy if last_difficulty else 'Analyze'
    current_grade = last_difficulty.grade if last_difficulty else 1
    current_times = last_difficulty.times if last_difficulty else 7

    return render_template('games/quess.html',games=games,sector_code=sector_code, lesson_id=lesson_id, current_difficulty=current_difficulty, current_importance=current_importance, current_bloom_taxonomy=current_bloom_taxonomy, current_grade=current_grade, current_times=current_times)

@app.route('/traslate', methods=['GET', 'POST'])
def traslate():
    lesson_id = request.args.get('lesson_id', type=int)
    employee_id = session.get('employee_id')
    company_code = session.get('sector_code')
    sector_code = session.get('sector_code')

    if request.method == 'POST':
     
        difficulty = request.form.get('col1')
        importance = request.form.get('importance')
        bloom_taxonomy = request.form.get('bloom_taxonomy')
        grade = request.form.get('grade')
        times = int(request.form.get('times', 1))  
        Word_Puzzle = request.form.get('Word_Puzzle')
        translate = request.form.get('translate')
        uploaded_file = request.files.get('Word_Puzzle_file')

        if Word_Puzzle:
            try:
                existing_word = TraslateArabic.query.filter_by(lesson_id=lesson_id, words=Word_Puzzle).first()
                if not existing_word:
                    new_word = TraslateArabic(
                        lesson_id=lesson_id,
                        user_id=employee_id,
                        company_code=company_code,
                        difficulty=difficulty,
                        importance=importance,
                        bloom_taxonomy=bloom_taxonomy,
                        grade=grade,
                        times=times,
                        words=Word_Puzzle,
                        translate=translate
                    )
                    db.session.add(new_word)
                    db.session.commit()
                else:
                    print("Word already exists for this exam.")
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")

        if uploaded_file and uploaded_file.filename != '':
            filename = secure_filename(uploaded_file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            uploaded_file.save(filepath)

            try:
                workbook = load_workbook(filepath)
                sheet = workbook.active
                for row in sheet.iter_rows(min_row=2):
                    word = row[0].value
                    if word: 
                        existing_word = TraslateArabic.query.filter_by(
                            lesson_id=lesson_id, words=word
                        ).first()
                        if not existing_word:
                            new_word = TraslateArabic(
                                lesson_id=lesson_id,
                                user_id=employee_id,
                                company_code=company_code,
                                difficulty=difficulty,
                                importance=importance,
                                bloom_taxonomy=bloom_taxonomy,
                                grade=grade,
                                times=times,
                                words=word,
                                translate=translate
                            )
                            db.session.add(new_word)
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")
            finally:
                os.remove(filepath)

    games = TraslateArabic.query.filter_by(lesson_id=lesson_id).all()
    last_difficulty = TraslateArabic.query.filter_by(lesson_id=lesson_id).order_by(TraslateArabic.id.desc()).first()
    current_difficulty = last_difficulty.difficulty if last_difficulty else 'Easy' 
    current_importance = last_difficulty.importance if last_difficulty else 'Necessary'
    current_bloom_taxonomy = last_difficulty.bloom_taxonomy if last_difficulty else 'Analyze'
    current_grade = last_difficulty.grade if last_difficulty else 1
    current_times = last_difficulty.times if last_difficulty else 7

    return render_template('games/traslate.html',games=games,sector_code=sector_code, lesson_id=lesson_id, current_difficulty=current_difficulty, current_importance=current_importance, current_bloom_taxonomy=current_bloom_taxonomy, current_grade=current_grade, current_times=current_times)

@app.route('/traslate_icon', methods=['GET', 'POST'])
def traslate_icon():
    lesson_id = request.args.get('lesson_id', type=int)
    employee_id = session.get('employee_id')
    company_code = session.get('sector_code')
    sector_code = session.get('sector_code')

    if request.method == 'POST':
     
        difficulty = request.form.get('col1')
        importance = request.form.get('importance')
        bloom_taxonomy = request.form.get('bloom_taxonomy')
        grade = request.form.get('grade')
        times = int(request.form.get('times', 1))  
        Word_Puzzle = request.form.get('Word_Puzzle')
        translate = request.form.get('translate')
        uploaded_file = request.files.get('Word_Puzzle_file')

        if Word_Puzzle:
            try:
                existing_word = Traslateicon.query.filter_by(lesson_id=lesson_id, words=Word_Puzzle).first()
                if not existing_word:
                    new_word = Traslateicon(
                        lesson_id=lesson_id,
                        user_id=employee_id,
                        company_code=company_code,
                        difficulty=difficulty,
                        importance=importance,
                        bloom_taxonomy=bloom_taxonomy,
                        grade=grade,
                        times=times,
                        words=Word_Puzzle,
                        translate=translate
                    )
                    db.session.add(new_word)
                    db.session.commit()
                else:
                    print("Word already exists for this exam.")
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")

        if uploaded_file and uploaded_file.filename != '':
            filename = secure_filename(uploaded_file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            uploaded_file.save(filepath)

            try:
                workbook = load_workbook(filepath)
                sheet = workbook.active
                for row in sheet.iter_rows(min_row=2):
                    word = row[0].value
                    if word: 
                        existing_word = Traslateicon.query.filter_by(
                            lesson_id=lesson_id, words=word
                        ).first()
                        if not existing_word:
                            new_word = Traslateicon(
                                lesson_id=lesson_id,
                                user_id=employee_id,
                                company_code=company_code,
                                difficulty=difficulty,
                                importance=importance,
                                bloom_taxonomy=bloom_taxonomy,
                                grade=grade,
                                times=times,
                                words=word,
                                translate=translate
                            )
                            db.session.add(new_word)
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")
            finally:
                os.remove(filepath)

    games = Traslateicon.query.filter_by(lesson_id=lesson_id).all()
    last_difficulty = Traslateicon.query.filter_by(lesson_id=lesson_id).order_by(Traslateicon.id.desc()).first()
    current_difficulty = last_difficulty.difficulty if last_difficulty else 'Easy' 
    current_importance = last_difficulty.importance if last_difficulty else 'Necessary'
    current_bloom_taxonomy = last_difficulty.bloom_taxonomy if last_difficulty else 'Analyze'
    current_grade = last_difficulty.grade if last_difficulty else 1
    current_times = last_difficulty.times if last_difficulty else 7

    return render_template('games/traslate_icon.html',games=games,sector_code=sector_code, lesson_id=lesson_id, current_difficulty=current_difficulty, current_importance=current_importance, current_bloom_taxonomy=current_bloom_taxonomy, current_grade=current_grade, current_times=current_times)

@app.route('/word_game', methods=['GET', 'POST'])
def word_game():
    lesson_id = request.args.get('lesson_id', type=int)
    user_id = session.get('user_id')
    company_code = session.get('sector_code')
    games = WordPuzzleGame2.query.filter_by(lesson_id=lesson_id).all()
    words = WordPuzzleGame2.query.filter_by(lesson_id=lesson_id).all()
    word_list = [word.words for word in words]

    last_difficulty = WordPuzzleGame2.query.filter_by(lesson_id=lesson_id).order_by(WordPuzzleGame2.id.desc()).first()
    current_difficulty = last_difficulty.difficulty if last_difficulty else 'Easy'
    current_importance = last_difficulty.importance if last_difficulty else 'Necessary'
    current_bloom_taxonomy = last_difficulty.bloom_taxonomy if last_difficulty else 'Analyze'
    current_grade = last_difficulty.grade if last_difficulty else 1
    current_times = last_difficulty.times if last_difficulty else 7

    return render_template('games/word_game.html', 
                           games=games, 
                           lesson_id=lesson_id, 
                           current_difficulty=current_difficulty, 
                           word_list=word_list,
                           current_times=current_times,current_grade=current_grade,current_bloom_taxonomy=current_bloom_taxonomy,
                           current_importance=current_importance)

@app.route('/reverse_game', methods=['GET', 'POST'])
def reverse_game():
    lesson_id = request.args.get('lesson_id', type=int)
    user_id = session.get('user_id')
    company_code = session.get('sector_code')
    sector_code = session.get('sector_code')
    games = ReverseGame1.query.filter_by(lesson_id=lesson_id).all()
    words = ReverseGame1.query.filter_by(lesson_id=lesson_id).all()
    word_list = [word.words for word in words]
    last_difficulty = ReverseGame1.query.filter_by(lesson_id=lesson_id).order_by(ReverseGame1.id.desc()).first()
    current_difficulty = last_difficulty.difficulty if last_difficulty else 'Easy'
    current_importance = last_difficulty.importance if last_difficulty else 'Necessary'
    current_bloom_taxonomy = last_difficulty.bloom_taxonomy if last_difficulty else 'Analyze'
    current_grade = last_difficulty.grade if last_difficulty else 1
    current_times = last_difficulty.times if last_difficulty else 7

    return render_template('games/reverse_game.html', 
                           games=games, 
                           lesson_id=lesson_id, 
                           current_difficulty=current_difficulty, 
                           word_list=word_list,
                           current_times=current_times,current_grade=current_grade,current_bloom_taxonomy=current_bloom_taxonomy,
                           current_importance=current_importance)

@app.route('/quess_game', methods=['GET', 'POST'])
def quess_game():
    lesson_id = request.args.get('lesson_id', type=int)
    user_id = session.get('user_id')
    company_code = session.get('sector_code')
    games = Quessgame.query.filter_by(lesson_id=lesson_id).all()
    words = Quessgame.query.filter_by(lesson_id=lesson_id).all()
    word_list = [word.words for word in words]
    translate_list = [word.translate for word in words]
    last_difficulty = Quessgame.query.filter_by(lesson_id=lesson_id).order_by(Quessgame.id.desc()).first()
    current_difficulty = last_difficulty.difficulty if last_difficulty else 'Easy'

    return render_template('games/quess_game.html', 
                           games=games, 
                           lesson_id=lesson_id, 
                           current_difficulty=current_difficulty, 
                           word_list=word_list,translate_list=translate_list)

@app.route('/match_game', methods=['GET', 'POST'])
def match_game():
    lesson_id = request.args.get('lesson_id', type=int)
    user_id = session.get('user_id')
    company_code = session.get('sector_code')
    games = MatchGame.query.filter_by(lesson_id=lesson_id).all()
    words = MatchGame.query.filter_by(lesson_id=lesson_id).all()
    word_list = [word.words for word in words]
    last_difficulty = MatchGame.query.filter_by(lesson_id=lesson_id).order_by(MatchGame.id.desc()).first()
    current_difficulty = last_difficulty.difficulty if last_difficulty else 'Easy'
    current_importance = last_difficulty.importance if last_difficulty else 'Necessary'
    current_bloom_taxonomy = last_difficulty.bloom_taxonomy if last_difficulty else 'Analyze'
    current_grade = last_difficulty.grade if last_difficulty else 1
    current_times = last_difficulty.times if last_difficulty else 7

    return render_template('games/match_game.html', 
                           games=games, 
                           lesson_id=lesson_id, 
                           current_difficulty=current_difficulty, 
                           word_list=word_list,
                           current_times=current_times,current_grade=current_grade,current_bloom_taxonomy=current_bloom_taxonomy,
                           current_importance=current_importance)

@app.route('/traslate_game', methods=['GET', 'POST'])
def traslate_game():
    lesson_id = request.args.get('lesson_id', type=int)
    user_id = session.get('user_id')
    company_code = session.get('sector_code')
    games = TraslateArabic.query.filter_by(lesson_id=lesson_id).all()
    words = TraslateArabic.query.filter_by(lesson_id=lesson_id).all()
    word_list = [word.words for word in words]
    translate_list = [word.translate for word in words]
    last_difficulty = TraslateArabic.query.filter_by(lesson_id=lesson_id).order_by(TraslateArabic.id.desc()).first()
    current_difficulty = last_difficulty.difficulty if last_difficulty else 'Easy'
    current_importance = last_difficulty.importance if last_difficulty else 'Necessary'
    current_bloom_taxonomy = last_difficulty.bloom_taxonomy if last_difficulty else 'Analyze'
    current_grade = last_difficulty.grade if last_difficulty else 1
    current_times = last_difficulty.times if last_difficulty else 7

    return render_template('games/traslate_game.html', 
                            games=games, 
                            lesson_id=lesson_id, 
                            translate_list=translate_list,
                            current_difficulty=current_difficulty, 
                            word_list=word_list,
                            current_times=current_times,current_grade=current_grade,
                            current_bloom_taxonomy=current_bloom_taxonomy,
                            current_importance=current_importance)

@app.route('/traslate_icon_game', methods=['GET', 'POST'])
def traslate_icon_game():
    lesson_id = request.args.get('lesson_id', type=int)
    user_id = session.get('user_id')
    company_code = session.get('sector_code')
    games = Traslateicon.query.filter_by(lesson_id=lesson_id).all()
    words = Traslateicon.query.filter_by(lesson_id=lesson_id).all()
    word_list = [word.words for word in words]
    translate_list = [word.translate for word in words]
    last_difficulty = Traslateicon.query.filter_by(lesson_id=lesson_id).order_by(Traslateicon.id.desc()).first()
    current_difficulty = last_difficulty.difficulty if last_difficulty else 'Easy'
    current_importance = last_difficulty.importance if last_difficulty else 'Necessary'
    current_bloom_taxonomy = last_difficulty.bloom_taxonomy if last_difficulty else 'Analyze'
    current_grade = last_difficulty.grade if last_difficulty else 1
    current_times = last_difficulty.times if last_difficulty else 7

    return render_template('games/traslate_icon_game.html', 
                           games=games, translate_list=translate_list,
                            lesson_id=lesson_id, 
                           current_difficulty=current_difficulty, 
                           word_list=word_list,
                           current_times=current_times,current_grade=current_grade,current_bloom_taxonomy=current_bloom_taxonomy,
                           current_importance=current_importance)

@app.route('/delete_word/<int:word_id>', methods=['GET'])
def delete_word(word_id):
    word = WordPuzzleGame2.query.get_or_404(word_id)
    try:
        db.session.delete(word)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Error: {e}")
    return redirect(url_for('Word_Puzzle', lesson_id=word.lesson_id))

@app.route('/delete_match/<int:word_id>', methods=['GET'])
def delete_match(word_id):
    word = MatchGame.query.get_or_404(word_id)
    try:
        db.session.delete(word)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Error: {e}")
    return redirect(url_for('match', lesson_id=word.lesson_id))


@app.route('/delete_reverse/<int:word_id>', methods=['GET'])
def delete_reverse(word_id):
    word = ReverseGame1.query.get_or_404(word_id)
    try:
        db.session.delete(word)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Error: {e}")
    return redirect(url_for('reverse', lesson_id=word.lesson_id))

@app.route('/delete_quess/<int:word_id>', methods=['GET'])
def delete_quess(word_id):
    word = Quessgame.query.get_or_404(word_id)
    try:
        db.session.delete(word)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Error: {e}")
    return redirect(url_for('quess', lesson_id=word.lesson_id))

@app.route('/delete_traslate/<int:word_id>', methods=['GET'])
def delete_traslate(word_id):
    word = TraslateArabic.query.get_or_404(word_id)
    try:
        db.session.delete(word)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Error: {e}")
    return redirect(url_for('traslate', lesson_id=word.lesson_id))

@app.route('/delete_traslate_icon/<int:word_id>', methods=['GET'])
def delete_traslate_icon(word_id):
    word = Traslateicon.query.get_or_404(word_id)
    try:
        db.session.delete(word)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Error: {e}")
    return redirect(url_for('traslate_icon', lesson_id=word.lesson_id))


@app.route('/redirect_exam/<int:exam_type>/<int:exam_id>', methods=['GET', 'POST'])
def redirect_exam(exam_type, exam_id):
    
    exam = Exam.query.get_or_404(exam_id)

    if exam_type == 1:  
        return redirect(url_for('edit_question_Multiple_Multiple', exam_id=exam_id))
    elif exam_type == 2:  
        return redirect(url_for('edit_question_true_or_false', exam_id=exam_id))
    elif exam_type == 3:  
        return redirect(url_for('edit_fill_in_blank', exam_id=exam_id))
    elif exam_type == 4:  
        return redirect(url_for('edit_question_sequence', exam_id=exam_id))
    elif exam_type == 6:  
        return redirect(url_for('edit_question_Multiple_single', exam_id=exam_id))
    elif exam_type == 12:  
        return redirect(url_for('Word_Puzzle', exam_id=exam_id))
    elif exam_type == 13:  
        return redirect(url_for('match', exam_id=exam_id))
    elif exam_type == 14:  
        return redirect(url_for('reverse', exam_id=exam_id))
    elif exam_type == 15:  
        return redirect(url_for('traslate', exam_id=exam_id))
    elif exam_type == 16:  
        return redirect(url_for('traslate_icon', exam_id=exam_id))
    elif exam_type == 17:
        return redirect(url_for('quess', exam_id=exam_id))
    else:
        return redirect(url_for('general_exam_page', exam_id=exam_id))
    
@app.route('/delete_exam/<int:exam_id>', methods=['POST'])
def delete_exam(exam_id):
    exam = Exam.query.get_or_404(exam_id)
    
    db.session.delete(exam)
    db.session.commit()
    
    return redirect(url_for('edit_exams'))

@app.route('/questions', methods=['GET', 'POST'])
def questions():
    lesson_id = request.args.get('id')
    sector_code = session.get('sector_code')
    return render_template('questions.html',lesson_id=lesson_id,sector_code=sector_code)  

@app.route('/submit_exam/<int:lesson_id>', methods=['POST'])
def submit_exam(lesson_id):
    
    exam_name = request.form['exam_name']
    description = request.form['description']
    exam_type = request.form['exam_type']

    new_exam = Exam(
        lesson_id=lesson_id,
        exam_name=exam_name,
        description=description,
        exam_type=exam_type
    )
    db.session.add(new_exam)
    db.session.commit()

    return redirect(url_for('edit_exams')) 

@app.route('/exam_cards', methods=['GET', 'POST'])
def exam_cards():
    employee_id = session.get('employee_id')
    sector_code = session.get('sector_code')

    if not sector_code or not employee_id:
        return redirect(url_for('access'))

    lessons = category.query.filter_by(employee_id=employee_id).all()
    lessons_with_counts = []
    for lesson in lessons:
        lesson_count = Chapter.query.filter_by(id_chapter=lesson.id).count()
        lessons_with_counts.append({
            'lesson': lesson,
            'lesson_count': lesson_count
        })

    return render_template('exam_cards.html', lessons_with_counts=lessons_with_counts,sector_code=sector_code)

@app.route('/lesson_cards', methods=['GET', 'POST'])
def lesson_cards():
    employee_id = session.get('employee_id')
    sector_code = session.get('sector_code')
    if not sector_code or not employee_id:
        return redirect(url_for('access'))
    chapter_id = request.args.get('chapter_id', type=int)
    lessons = Chapter.query.filter_by(id_chapter=chapter_id).all()
    lessons_with_counts = []
    
    for lesson in lessons:
        lesson_count = Lesson8.query.filter_by(id_chapter=lesson.id).count()
        lessons_with_counts.append({
            'lesson': lesson,
            'lesson_count': lesson_count
        })
    return render_template('lesson_cards.html', lessons=lessons,sector_code=sector_code,lessons_with_counts=lessons_with_counts)

@app.route('/lesson', methods=['GET', 'POST'])
def lesson():
    employee_id = session.get('employee_id')
    sector_code = session.get('sector_code')
    if not sector_code or not employee_id:
        return redirect(url_for('access'))
    chapter_id = request.args.get('chapter_id', type=int)
    lessons = Lesson8.query.filter_by(id_chapter=chapter_id).all()
    
    lessons_with_question_counts = []
    for lesson in lessons:
        mcq_multiple_count = QuestionMultipleMultiple_choice.query.filter_by(lesson_id=lesson.id).count()
        mcq_single_count = Question_Multipleone_choice.query.filter_by(lesson_id=lesson.id).count()
        Questionfort_or_f_count = Questionfort_or_f.query.filter_by(lesson_id=lesson.id).count()
        Question_Fillin_blank4_count = Question_Fillin_blank4.query.filter_by(lesson_id=lesson.id).count()
        Questionsequence1_count = Questionsequence1.query.filter_by(lesson_id=lesson.id).count()
        questionmatching1_count = questionmatching1.query.filter_by(lesson_id=lesson.id).count()
        questionclassiffication_count = questionclassiffication.query.filter_by(lesson_id=lesson.id).count()
        Question_Fillin_blank_choice2_count = Question_Fillin_blank_choice2.query.filter_by(lesson_id=lesson.id).count()
        Article_question_count = Article_question.query.filter_by(lesson_id=lesson.id).count()
        
        
        total_questions = mcq_multiple_count + mcq_single_count+Questionfort_or_f_count+Question_Fillin_blank4_count+Questionsequence1_count+questionmatching1_count+questionclassiffication_count+Question_Fillin_blank_choice2_count+Article_question_count
        lessons_with_question_counts.append({
        'lesson': lesson,
        'question_count': total_questions
        })
        print(f" Question Count: {total_questions}")
    return render_template('lesson.html', lessons_with_question_counts=lessons_with_question_counts,sector_code=sector_code)

@app.route('/lesson_workshop', methods=['GET', 'POST'])
def lesson_workshop():
    employee_id = session.get('employee_id')
    sector_code = session.get('sector_code')
    if not sector_code or not employee_id:
        return redirect(url_for('access'))
    workshop_id = request.args.get('workshop_id', type=int)
    lessons = lessons_workshop.query.filter_by(workshop_id=workshop_id).all()
    return render_template('lesson_workshop.html', lessons=lessons,sector_code=sector_code)


@app.route('/lesson_workshop_details', methods=['GET', 'POST'])
def lesson_workshop_details():
    employee_id = session.get('employee_id')
    sector_code = session.get('sector_code')
    if not sector_code or not employee_id:
        return redirect(url_for('access'))
    
    lesson_id = request.args.get('lesson_id', type=int)
    lesson = lessons_workshop.query.filter_by(id=lesson_id).first()  
   
    if not lesson: 
        return "Lesson not found", 404  
    
    if request.method == 'POST':
        title = request.form.get('title')
        link = request.form.get('link')
        description = request.form.get('description')
        
        file = request.files.get('file')
        image = request.files.get('image')
        video = request.files.get('video')
        audio = request.files.get('audio')
        
        lesson.title = title
        lesson.link = link
        lesson.description = description
        
        if file:
            lesson.file_path = save_file(file, 'F', lesson_id, lesson.file_path) 
        if image:
            lesson.image_path = save_file(image, 'P', lesson_id, lesson.image_path) 
        if video:
            lesson.video_path = save_file(video, 'V', lesson_id, lesson.video_path)  
        if audio:
            lesson.audio_path = save_file(audio, 'A', lesson_id, lesson.audio_path) 
        db.session.commit()  
        print(lesson.file_path)
        return redirect(url_for('lesson_workshop_details', lesson_id=lesson_id)) 
    
    return render_template('lesson_workshop_details.html', lesson=lesson, sector_code=sector_code)


@app.route('/add_lesson_workshop', methods=['GET', 'POST'])
def add_lesson_workshop():
    title = request.form.get('title') 
    image = request.files.get('image')  
    employee_id = session.get('employee_id')

    workshop_id = request.args.get('workshop_id', type=int) 
   
    if not title or not employee_id or not workshop_id:
        return "Title, Employee ID, and Chapter ID are required", 400

  
    new_lesson = lessons_workshop(
        title=title,
        employee_id=employee_id,
        workshop_id=workshop_id,
        image_path=None  
    )
    db.session.add(new_lesson)
    db.session.commit() 

    if image:
        image_path = f"static/uploads/{new_lesson.id}_L_{image.filename}"
        image.save(image_path)
        
        new_lesson.image_path = image_path
        db.session.commit()

    return redirect(url_for('lesson_workshop', workshop_id=workshop_id))    

@app.route('/add_lesson1', methods=['GET', 'POST'])
def add_lesson1():
    title = request.form.get('title') 
    image = request.files.get('image')  
    employee_id = session.get('employee_id')

    chapter_id = request.args.get('chapter_id', type=int) 
   
    if not title or not employee_id or not chapter_id:
        return "Title, Employee ID, and Chapter ID are required", 400

  
    new_lesson = Lesson8(
        title=title,
        employee_id=employee_id,
        id_chapter=chapter_id,
        image_path=None  
    )
    db.session.add(new_lesson)
    db.session.commit() 

    if image:
        image_path = f"static/uploads/{new_lesson.id}_L_{image.filename}"
        image.save(image_path)
        
        new_lesson.image_path = image_path
        db.session.commit()

    return redirect(url_for('lesson', chapter_id=chapter_id))



upload_folder = os.path.join(app.root_path, 'static', 'uploads')

def save_file(file, file_type, lesson_id, current_file_path=None):
    if file:
        file_extension = os.path.splitext(file.filename)[1]
        filename = f"{lesson_id}_{file_type}{file_extension}"
        file_path = os.path.join(upload_folder, filename)
        
        if current_file_path and os.path.exists(current_file_path):
            os.remove(current_file_path)

        file.save(file_path)

        return os.path.join('uploads', filename).replace("\\", "/")
    return None

@app.route('/lesson_details', methods=['GET', 'POST'])
def lesson_details():
    employee_id = session.get('employee_id')
    sector_code = session.get('sector_code')
    if not sector_code or not employee_id:
        return redirect(url_for('access'))
    
    lesson_id = request.args.get('lesson_id', type=int)
    lesson = Lesson8.query.filter_by(id=lesson_id).first()
    
    if not lesson:
        return "Lesson not found", 404
    
    if request.method == 'POST':
        # التحقق إذا كان طلب حذف
        delete_file_field = request.form.get('delete_file')
        if delete_file_field:
            file_path = getattr(lesson, delete_file_field, None)
            if file_path:
                # حذف الملف من النظام
                delete_file(file_path)
                # إزالة المسار من قاعدة البيانات
                setattr(lesson, delete_file_field, None)
                db.session.commit()
            return redirect(url_for('lesson_details', lesson_id=lesson_id))
        
        # تعديل البيانات
        title = request.form.get('title', lesson.title)
        link = request.form.get('link', lesson.link)
        description = request.form.get('description', lesson.description)
        
        lesson.title = title
        lesson.link = link
        lesson.description = description
        
        files = request.files.getlist('file')
        image = request.files.get('image')
        video = request.files.get('video')
        audio = request.files.get('audio')
        
        if files:
            for file in files:
                if not lesson.file_path:
                    lesson.file_path = save_file(file, 'file', lesson_id, lesson.file_path)
                elif not lesson.file_path1:
                    lesson.file_path1 = save_file(file, 'file1', lesson_id, lesson.file_path1)
                elif not lesson.file_path2:
                    lesson.file_path2 = save_file(file, 'file2', lesson_id, lesson.file_path2)
                elif not lesson.file_path3:
                    lesson.file_path3 = save_file(file, 'file3', lesson_id, lesson.file_path3)
                elif not lesson.file_path4:
                    lesson.file_path4 = save_file(file, 'file4', lesson_id, lesson.file_path4)

        if image:
            lesson.image_path = save_file(image, 'image', lesson_id, lesson.image_path)
        if video:
            lesson.video_path = save_file(video, 'video', lesson_id, lesson.video_path)
        if audio:
            lesson.audio_path = save_file(audio, 'audio', lesson_id, lesson.audio_path)
        
        db.session.commit()
        return redirect(url_for('lesson_details', lesson_id=lesson_id))
    
    return render_template('lesson_details.html', lesson=lesson, sector_code=sector_code)
@app.route('/view_lesson')
def view_lesson():
    lesson_id = request.args.get('lesson_id', type=int)
    if not lesson_id:
        return "Lesson ID not provided", 400

    lesson = Lesson8.query.filter_by(id=lesson_id).first()
    if not lesson:
        return "Lesson not found", 404

    return render_template('view_lesson.html', lesson=lesson,lesson_id=lesson_id)
def delete_file(file_path):
    full_path = os.path.join('static', file_path)
    if os.path.exists(full_path):
        os.remove(full_path)

@app.route('/edit_lessons/<int:lesson_id>', methods=['POST'])
def edit_lessons(lesson_id):
    lesson = Chapter.query.get_or_404(lesson_id)
    chapter_id = lesson.id_chapter
    title = request.form.get('title')
    sector_code = session.get('sector_code')
    if title:
        lesson.title = title
        db.session.commit()

    return redirect(url_for('lesson_cards', chapter_id=chapter_id,sector_code=sector_code))

@app.route('/delete_lessons/<int:lesson_id>', methods=['POST'])
def delete_lessons(lesson_id):
    lesson = Chapter.query.get_or_404(lesson_id)
    chapter_id = lesson.id_chapter
    if lesson.image_path:
        os.remove(lesson.image_path) 
    db.session.delete(lesson)
    db.session.commit()
    return redirect(url_for('lesson_cards', chapter_id=chapter_id))
os.makedirs('static/uploads', exist_ok=True)
@app.route('/add_lesson', methods=['GET', 'POST'])
def add_lesson():
    
    title = request.form.get('title') 
    image = request.files.get('image')  
    employee_id = session.get('employee_id')
    if not title or not employee_id:
        return "Title and Employee ID are required", 400

    new_lesson = category(
        title=title,
        employee_id=employee_id,
        image_path=None  
    )
    db.session.add(new_lesson)
    db.session.commit() 

    if image:
        image_path = f"static/uploads/{new_lesson.id}_L_{image.filename}"
        image.save(image_path)

        new_lesson.image_path = image_path
        db.session.commit()

    return redirect(url_for('exam_cards'))

@app.route('/edit_lesson/<int:lesson_id>', methods=['POST'])
def edit_lesson(lesson_id):
    lesson = category.query.get_or_404(lesson_id)
    title = request.form.get('title')

    if title:
        lesson.title = title
        db.session.commit()

    return redirect(url_for('exam_cards'))

@app.route('/add_lessons', methods=['GET', 'POST'])
def add_lessons():
    title = request.form.get('title') 
    image = request.files.get('image')  
    employee_id = session.get('employee_id')

    chapter_id = request.args.get('chapter_id', type=int) 
   
    if not title or not employee_id or not chapter_id:
        return "Title, Employee ID, and Chapter ID are required", 400

  
    new_lesson = Chapter(
        title=title,
        employee_id=employee_id,
        id_chapter=chapter_id,
        image_path=None  
    )
    db.session.add(new_lesson)
    db.session.commit() 

    if image:
        image_path = f"static/uploads/{new_lesson.id}_L_{image.filename}"
        image.save(image_path)
        
        new_lesson.image_path = image_path
        db.session.commit()


    return redirect(url_for('lesson_cards', chapter_id=chapter_id))


@app.route('/delete_lesson/<int:lesson_id>', methods=['POST'])
def delete_lesson(lesson_id):
    lesson = category.query.get_or_404(lesson_id)
    if lesson.image_path:
        os.remove(lesson.image_path) 
    db.session.delete(lesson)
    db.session.commit()
    return redirect(url_for('exam_cards'))

@app.route('/delete_lesson2/<int:lesson_id>', methods=['POST'])
def delete_lesson2(lesson_id):
    lesson = Lesson8.query.get_or_404(lesson_id)
    chapter_id = lesson.id_chapter
    if lesson.image_path:
        os.remove(lesson.image_path) 
    db.session.delete(lesson)
    db.session.commit()
    return redirect(url_for('lesson', chapter_id=chapter_id))

@app.route('/edit_lesson2/<int:lesson_id>', methods=['POST'])
def edit_lesson2(lesson_id):
    lesson = Lesson8.query.get_or_404(lesson_id)
    chapter_id = lesson.id_chapter
    title = request.form.get('title')

    if title:
        lesson.title = title
        db.session.commit()

    return redirect(url_for('lesson', chapter_id=chapter_id))

@app.route('/delete_question_T_F/<int:question_id>', methods=['POST'])
def delete_question_T_F(question_id):
    question = Questionfort_or_f.query.get_or_404(question_id)
    lesson_id = question.lesson_id 
    if question.photo_path:
        try:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.photo_path))
        except FileNotFoundError:
            pass

    if question.audio_path:
        try:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.audio_path))
        except FileNotFoundError:
            pass

    try:
        db.session.delete(question)
        db.session.commit()
      
    except Exception as e:
        print(f"Error deleting question: {e}")
       

    return redirect(url_for('edit_question_true_or_false',lesson_id=lesson_id))

@app.route('/delete_mcq_multiple_single/<int:question_id>', methods=['POST'])
def delete_mcq_multiple_single(question_id):
    # جلب السؤال من قاعدة البيانات
    question = QuestionMultipleMultiple_choice.query.get_or_404(question_id)
    lesson_id = question.lesson_id 
    # حذف الصورة الخاصة بالسؤال إذا كانت موجودة
    if question.photo_path:
        try:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.photo_path))
        except FileNotFoundError:
            pass

    # حذف الصوت الخاص بالسؤال إذا كان موجودًا
    if question.audio_path:
        try:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.audio_path))
        except FileNotFoundError:
            pass

    # جلب الإجابات المرتبطة بالسؤال وحذف الصور والصوت المرتبطة بها
    answers = Answer_Multiple_Multiple_choice.query.filter_by(question_id=question_id).all()
    for answer in answers:
        if answer.photo_path:
            try:
                os.remove(os.path.join(app.config['UPLOAD_FOLDER'], answer.photo_path))
            except FileNotFoundError:
                pass
        if answer.audio_path:
            try:
                os.remove(os.path.join(app.config['UPLOAD_FOLDER'], answer.audio_path))
            except FileNotFoundError:
                pass
        # حذف الإجابة نفسها
        db.session.delete(answer)

    # حذف السؤال نفسه من قاعدة البيانات
    db.session.delete(question)

    # حفظ التغييرات في قاعدة البيانات
    try:
        db.session.commit()
      
    except Exception as e:
        print(f"Error deleting question: {e}")
    

    return redirect(url_for('edit_question_Multiple_Multiple',lesson_id=lesson_id))



@app.route('/delete_mcq_multiple_s/<int:question_id>', methods=['POST'])
def delete_mcq_multiple_s(question_id):
    # جلب السؤال من قاعدة البيانات
    question = Question_Multipleone_choice.query.get_or_404(question_id)
    lesson_id = question.lesson_id 
    # حذف الصورة الخاصة بالسؤال إذا كانت موجودة
    if question.photo_path:
        try:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.photo_path))
        except FileNotFoundError:
            pass

    # حذف الصوت الخاص بالسؤال إذا كان موجودًا
    if question.audio_path:
        try:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.audio_path))
        except FileNotFoundError:
            pass

    # جلب الإجابات المرتبطة بالسؤال وحذف الصور والصوت المرتبطة بها
    answers = Answer_Multiple_Multiple_choice.query.filter_by(question_id=question_id).all()
    for answer in answers:
        if answer.photo_path:
            try:
                os.remove(os.path.join(app.config['UPLOAD_FOLDER'], answer.photo_path))
            except FileNotFoundError:
                pass
        if answer.audio_path:
            try:
                os.remove(os.path.join(app.config['UPLOAD_FOLDER'], answer.audio_path))
            except FileNotFoundError:
                pass
        # حذف الإجابة نفسها
        db.session.delete(answer)

    # حذف السؤال نفسه من قاعدة البيانات
    db.session.delete(question)

    # حفظ التغييرات في قاعدة البيانات
    try:
        db.session.commit()
       
    except Exception as e:
        print(f"Error deleting question: {e}")
    

    return redirect(url_for('edit_question_Multiple_single',lesson_id=lesson_id))

@app.route('/company_cards', methods=['GET', 'POST'])
def company_cards():
    organizer_id = session.get('organizer_id')
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    sector_code = session.get('sector_code')
    branches = CompanyName.query.filter_by(code=sector_code).all()
    col3 = Administrative_Title.query.filter_by(code=sector_code).all()
    col4 = JobTitle.query.filter_by(code=sector_code).all()
    col5 = specializationEmployee.query.filter_by(code=sector_code).all()
    data = []
    employees = Employee6.query.filter_by(code=sector_code).all()
    department_employees = DepartmentEmployee.query.filter_by(code=sector_code).all()
    department_employees1 = DepartmentEmployee1.query.filter_by(code=sector_code).all()
    for branch in branches:
        branch_name = branch.name
        branch_code = branch.code

        total_employees = Employee6.query.filter_by(branch=branch_name,code=sector_code).count()
        total_managers = Employee6.query.filter_by(branch=branch_name,code=sector_code,is_manager=True).count()
        total_heads = Employee6.query.filter_by(branch=branch_name,code=sector_code,is_head="ttt").count()

        head_details = head.query.filter_by(branch=branch_name,code=sector_code).all()
        manager_details = ManagerEmployee.query.filter_by(branch=branch_name,code=sector_code).all()
        employee_details = Employee6.query.filter_by(branch=branch_name,code=sector_code).all()

      
        col2_counts = {}
        department_values = [d.name for d in department_employees]
        for value in department_values:
            count = Employee6.query.filter_by(col2=value, branch=branch_name).count()
            col2_counts[value] = count
        
        col1_counts = {}
        department_values = [d.name for d in department_employees1]
        for value in department_values:
            count = Employee6.query.filter_by(col1=value, branch=branch_name).count()
            col1_counts[value] = count
        col3_counts = {}
        department_values = [d.name for d in col3]
        for value in department_values:
            count = Employee6.query.filter_by(col3=value, branch=branch_name).count()
            col3_counts[value] = count

        col4_counts = {}
        department_values = [d.name for d in col4]
        for value in department_values:
            count = Employee6.query.filter_by(col4=value, branch=branch_name).count()
            col4_counts[value] = count

        col5_counts = {}
        department_values = [d.name for d in col5]
        for value in department_values:
            count = Employee6.query.filter_by(col5=value, branch=branch_name).count()
            col5_counts[value] = count

        data.append({
            'branch_name': branch_name,
            'total_employees': total_employees,
            'total_managers': total_managers,
            'total_heads': total_heads,
            'head_details': head_details,
            'manager_details': manager_details,
            'employee_details': employee_details,
            'col2_counts': col2_counts, 
            'col1_counts': col1_counts,
            'col3_counts': col3_counts,
            'col4_counts': col4_counts,
            'col5_counts': col5_counts,
        })

    return render_template('company_cards.html',employees=employees, data=data,sector_id=sector_id,sector_code=sector_code)



@app.route('/company_cards_adm', methods=['GET', 'POST'])
def company_cards_adm():
    user_id = session.get('user_id')
    sector_code = session.get('sector_code')
    branches = CompanyName.query.filter_by(code=sector_code).all()
    col3 = Administrative_Title.query.filter_by(code=sector_code,id_admin=user_id).all()
    col4 = JobTitle.query.filter_by(code=sector_code,id_admin=user_id,).all()
    col5 = specializationEmployee.query.filter_by(code=sector_code,id_admin=user_id).all()
    data = []
    employees = Employee6.query.filter_by(code=sector_code,id_admin=user_id).all()
    department_employees = DepartmentEmployee.query.filter_by(code=sector_code,id_admin=user_id,).all()
    department_employees1 = DepartmentEmployee1.query.filter_by(code=sector_code,id_admin=user_id,).all()
    for branch in branches:
        branch_name = branch.name
        branch_code = branch.code

        total_employees = Employee6.query.filter_by(branch=branch_name,id_admin=user_id).count()
        total_managers = Employee6.query.filter_by(branch=branch_name,id_admin=user_id,is_manager=True).count()
        total_heads = Employee6.query.filter_by(branch=branch_name,id_admin=user_id,is_head="ttt").count()

        head_details = head.query.filter_by(branch=branch_name,id_admin=user_id).all()
        manager_details = ManagerEmployee.query.filter_by(branch=branch_name,id_admin=user_id,).all()
        employee_details = Employee6.query.filter_by(branch=branch_name,id_admin=user_id,).all()

      
        col2_counts = {}
        department_values = [d.name for d in department_employees]
        for value in department_values:
            count = Employee6.query.filter_by(col2=value, branch=branch_name,id_admin=user_id,).count()
            col2_counts[value] = count
        
        col1_counts = {}
        department_values = [d.name for d in department_employees1]
        for value in department_values:
            count = Employee6.query.filter_by(col1=value, branch=branch_name,id_admin=user_id,).count()
            col1_counts[value] = count
        col3_counts = {}
        department_values = [d.name for d in col3]
        for value in department_values:
            count = Employee6.query.filter_by(col3=value, branch=branch_name,id_admin=user_id,).count()
            col3_counts[value] = count

        col4_counts = {}
        department_values = [d.name for d in col4]
        for value in department_values:
            count = Employee6.query.filter_by(col4=value, branch=branch_name,id_admin=user_id,).count()
            col4_counts[value] = count

        col5_counts = {}
        department_values = [d.name for d in col5]
        for value in department_values:
            count = Employee6.query.filter_by(col5=value, branch=branch_name,id_admin=user_id,).count()
            col5_counts[value] = count

        data.append({
            'branch_name': branch_name,
            'total_employees': total_employees,
            'total_managers': total_managers,
            'total_heads': total_heads,
            'head_details': head_details,
            'manager_details': manager_details,
            'employee_details': employee_details,
            'col2_counts': col2_counts, 
            'col1_counts': col1_counts,
            'col3_counts': col3_counts,
            'col4_counts': col4_counts,
            'col5_counts': col5_counts,
        })

    return render_template('company_cards_adm.html', data=data,sector_code=sector_code,employees=employees)

from werkzeug.utils import secure_filename
from flask_sqlalchemy import SQLAlchemy
import random
from mutagen.mp3 import MP3
from mutagen.wave import WAVE
from mutagen.oggvorbis import OggVorbis
from mutagen.aac import AAC

ALLOWED_EXTENSIONS_T_F = {'png', 'jpg', 'jpeg', 'gif', 'mp3', 'wav'}

def allowed_file1(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS_T_F

def calculate_expected_time(question):
    expected_time = 0

    # حساب الوقت بناءً على طول النص
    if question.question_name:
        expected_time += len(question.question_name) // 5  # ثانية لكل 5 حروف

    # إضافة 20 ثانية إذا كانت هناك صورة
    if question.photo_path:
        expected_time += 20

    if question.audio_path:
         try:
            audio_path = os.path.join(app.config['UPLOAD_FOLDER'], question.audio_path)
            if audio_path.lower().endswith(('.mp3')):
                  audio = MP3(audio_path)
                  audio_duration = audio.info.length
                  expected_time += audio_duration + 5
            elif audio_path.lower().endswith(('.wav')):
                audio = WAVE(audio_path)
                audio_duration = audio.info.length
                expected_time += audio_duration + 5
            elif audio_path.lower().endswith(('.ogg')):
                  audio = OggVorbis(audio_path)
                  audio_duration = audio.info.length
                  expected_time += audio_duration + 5
            elif audio_path.lower().endswith(('.aac')):
                  audio = AAC(audio_path)
                  audio_duration = audio.info.length
                  expected_time += audio_duration + 5
            else:
                  expected_time += 5
         except Exception as e:
            print(f"Error getting audio duration: {e}")
            expected_time += 5
    return int(expected_time+7)



@app.route('/true_false', methods=['GET', 'POST'])
def true_false():
    sector_code = session.get('sector_code')
    question_id = request.args.get('question_id', type=int)
    question = Questionfort_or_f.query.get_or_404(question_id)
     
   
    if request.method == 'POST':
        # تحديث البيانات الأخرى
        question.question_name = request.form['question_name']
        question.difficulty = request.form['difficulty']
        question.importance = request.form['importance']
        question.bloom_taxonomy = request.form['bloom_taxonomy']
        question.grade = request.form['grade']
        question.times = request.form['times']
        correct_answer_value = request.form['correct_answer'] == 'True'
        question.correct_answer = correct_answer_value

       
        # معالجة رفع الصورة
        if 'photo' in request.files:
            photo = request.files['photo']
            if photo and allowed_file1(photo.filename):
                # حذف الصورة القديمة إذا كانت موجودة
                if question.photo_path:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.photo_path))
                    except FileNotFoundError:
                        pass

                # تحديد الاسم الجديد للصورة مع تغيير الامتداد إلى .png
                filename = f"T_F_{question_id}_{secure_filename(photo.filename).rsplit('.', 1)[0]}.png"
                try:
                    photo.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.photo_path = filename
                except Exception as e:
                    print(f"Error saving photo: {e}")

     
        if 'audio' in request.files:
            audio = request.files['audio']
            if audio and allowed_file1(audio.filename):
        
                if question.audio_path:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], question.audio_path))
                    except FileNotFoundError:
                        pass

                filename = f"T_F_{question_id}_{secure_filename(audio.filename)}"
                try:
                    audio.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                    question.audio_path = filename
                except Exception as e:
                    print(f"Error saving audio: {e}")

        try:
            db.session.commit()
        except Exception as e:
            print(f"Error saving data: {e}")
    expected_time = calculate_expected_time(question)
    print(expected_time)
    return render_template('true_false.html', question=question,sector_code=sector_code,expected_time=expected_time)

@app.route('/signup_sector', methods=['GET', 'POST'])
def signup_sector():
    if request.method == 'POST':
        foundation_name = request.form.get('foundation_name')
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')

        if not foundation_name or len(foundation_name) < 3:
            flash("اسم الشركة يجب أن يكون أطول من 3 أحرف", "error")
            return render_template('signup_sector.html')
        
        if not username or len(username) < 3:
            flash("اسم المستخدم يجب أن يكون أطول من 3 أحرف", "error")
            return render_template('signup_sector.html')

        email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
        if not email or not re.match(email_regex, email):
            flash("صيغة البريد الإلكتروني غير صحيحة", "error")
            return render_template('signup_sector.html')

        if not email or Sector.query.filter_by(email=email).first():
            flash("الإيميل مستخدم بالفعل أو غير صالح", "error")
            return render_template('signup_sector.html')
        
        if not password or len(password) < 6:
            flash("كلمة المرور يجب أن تكون أطول من 6 أحرف", "error")
            return render_template('signup_sector.html')

      
        code = None
        while not code or Sector.query.filter_by(code=code).first():
            code = str(random.randint(10**7, 10**8 - 1))
        
        
        new_sector = Sector(
            foundation_name=foundation_name,
            username=username,
            email=email,
            password=password,
            code=code
        )

        db.session.add(new_sector)
        db.session.commit()
        new_notification = Notification_owner(
        subject=f"Hello {username} 😊",
        message = "Welcome to Noor Academy for Sustainable Development, where we focus on nurturing human potential sustainably. We wish you a successful and continuous learning journey filled with growth .",
        company_code=code,
        timestamp=datetime.now()
        )
        db.session.add(new_notification)
        db.session.commit()


        session['sector_id'] = new_sector.id
        session['sector_code'] = new_sector.code

        return redirect(url_for('owner'))

    return render_template('signup_sector.html')
@app.route('/signup_personal', methods=['GET', 'POST'])
def signup_personal():
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        account_type = request.form.get('account_type')

        if username and email and password and account_type:
            new_personal = Personal(
                username=username,
                email=email,
                password=password,
                account_type=account_type
            )
            db.session.add(new_personal)
            db.session.commit()

            return redirect(url_for('success', username=username))

    return render_template('signup_personal.html')


@app.route('/', methods=['GET', 'POST'])
def home():
    return render_template('index.html') 

@app.route('/register', methods=['GET', 'POST'])
def index():
    return render_template('register.html')    

@app.route('/login', methods=['GET', 'POST'])
def login():
    return render_template('login.html')    

@app.route('/sector_login', methods=['GET', 'POST'])
def sector_login():
    return render_template('sector_login.html')

@app.route('/login_sector', methods=['GET', 'POST'])
def login_sector():
    if request.method == 'POST':
        email = request.form.get('email')
        print(email)
        password = request.form.get('password')

        sector = Sector.query.filter_by(email=email).first()
        if not sector:
            flash("الإيميل غير مسجل في النظام", "error")
            return render_template('login_sector.html')

        if sector.password != password:
            flash("كلمة المرور غير صحيحة", "error")
            return render_template('login_sector.html')

        session['sector_id'] = sector.id
        session['sector_code'] = sector.code

        return redirect(url_for('owner'))

    return render_template('login_sector.html')


@app.route('/signup_eml', methods=['GET', 'POST'])
def signup_eml():
    if request.method == 'POST':
        sector_code = request.args.get('sector_code') 
        id_admin = request.args.get('id') 
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')

        if not username or len(username) < 3:
            flash("اسم المستخدم قصير جدًا. يجب أن يكون 3 أحرف على الأقل.", "error")
        elif not email or not re.match(r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$', email):
            flash("صيغة البريد الإلكتروني غير صحيحة.", "error")
        elif Employee6.query.filter_by(email=email).first():
            flash("البريد الإلكتروني مستخدم بالفعل.", "error")
        elif not password or len(password) < 6:
            flash("كلمة المرور قصيرة جدًا. يجب أن تكون على الأقل 6 أحرف.", "error")
        else:
            try:
                new_employee = Employee6(code=sector_code, username=username, email=email,id_admin=id_admin, password=password)
                db.session.add(new_employee)
                db.session.commit()
                new_notification = Notification_admin(
                subject=f"New Employee Added",
                message=f"A new employee {username}, has been added to Noor Academy for Sustainable Development.",
                id_user=id_admin,
                company_code=sector_code,
                timestamp=datetime.now()
                )
                db.session.add(new_notification)
                db.session.commit()

                session['employee_id'] = new_employee.id
                session['employee_code'] = new_employee.code

                return redirect('/employee_profile')
            except Exception as e:
                db.session.rollback()
                flash("حدث خطأ أثناء إنشاء الحساب. يرجى المحاولة مرة أخرى.", "error")

    return render_template('signup_eml.html')
@app.route('/signup_organizer', methods=['GET', 'POST'])
def signup_organizer():
    code = request.args.get('code') 
    admin_data = ColOrganizer.query.filter_by(code=code).first()

    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        col1 = request.form.get('c1')
        col2 = request.form.get('c2')
        col3 = request.form.get('c3')
        col4 = request.form.get('c4')
        col5 = request.form.get('c5')
        password = request.form['password']
        sector_code = request.args.get('code') 
        
        email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
        if not email or not re.match(email_regex, email):
            flash("صيغة البريد الإلكتروني غير صحيحة", "error")
            return render_template('signup_organizer.html',
                                   c1=admin_data.c1,
                                   c2=admin_data.c2,
                                   c3=admin_data.c3,
                                   c4=admin_data.c4,
                                   c5=admin_data.c5)

        # تحقق من إذا كان البريد الإلكتروني مستخدمًا بالفعل
        existing_organizer = organizer2.query.filter_by(email=email).first()
        if existing_organizer:
            flash("البريد الإلكتروني مستخدم بالفعل", "error")
            return render_template('signup_organizer.html',
                                   c1=admin_data.c1,
                                   c2=admin_data.c2,
                                   c3=admin_data.c3,
                                   c4=admin_data.c4,
                                   c5=admin_data.c5)

        # تحقق من اسم المستخدم
        if len(username) < 3:
            flash("اسم المستخدم قصير جدًا. يجب أن يكون 3 أحرف على الأقل.", "error")
            return render_template('signup_organizer.html',
                                   c1=admin_data.c1,
                                   c2=admin_data.c2,
                                   c3=admin_data.c3,
                                   c4=admin_data.c4,
                                   c5=admin_data.c5)

        # تحقق من كلمة المرور
        if len(password) < 6:
            flash("كلمة المرور قصيرة جدًا. يجب أن تكون على الأقل 6 أحرف.", "error")
            return render_template('signup_organizer.html',
                                   c1=admin_data.c1,
                                   c2=admin_data.c2,
                                   c3=admin_data.c3,
                                   c4=admin_data.c4,
                                   c5=admin_data.c5)

        # إنشاء منظم جديد
        new_organizer = organizer2(
            name=username,
            email=email,
            col1=col1,
            col2=col2,
            col3=col3,
            col4=col4,
            col5=col5,
            password=password,
            code=sector_code
        )
        
        db.session.add(new_organizer)
        db.session.commit()


        new_notification = Notification_owner(
            subject="New Organizer Added",
            message=f"Organizer {username} has been added successfully.",
            company_code=sector_code,
            timestamp=datetime.now()
        )
        db.session.add(new_notification)
        db.session.commit()
        

        session['organizer_id'] = new_organizer.id
        session['sector_code'] = sector_code

        return redirect(url_for('organizer'))


    return render_template('signup_organizer.html',
                           c1=admin_data.c1,
                           c2=admin_data.c2,
                           c3=admin_data.c3,
                           c4=admin_data.c4,
                           c5=admin_data.c5)

@app.route('/notification', methods=['GET'])
def notification():
    
    company_code = session.get('sector_code')
    sector_code = session.get('sector_code')
    notifications = Notification_owner.query.filter_by(company_code=company_code).order_by(desc(Notification_owner.id)).all()


    for notification in notifications:
        if not notification.viewed: 
            notification.viewed = True

    db.session.commit()

    return render_template('notification.html', notifications=notifications,sector_code=sector_code)

@app.route('/notification_organizer', methods=['GET'])
def notification_organizer():
    
    company_code = session.get('sector_code')
    sector_code = session.get('sector_code')
    notifications = Notification_organizer.query.filter_by(company_code=company_code).order_by(desc(Notification_organizer.id)).all()

    for notification in notifications:
        if not notification.viewed: 
            notification.viewed = True

    db.session.commit()

    return render_template('notification_organizer.html', notifications=notifications,sector_code=sector_code)

@app.route('/notification_admin', methods=['GET'])
def notification_admin():
    user_id = session.get('user_id')
    company_code = session.get('sector_code')
    sector_code = session.get('sector_code')
    notifications = Notification_admin.query.filter_by(company_code=company_code, id_user=user_id).order_by(desc(Notification_admin.id)).all()

    for notification in notifications:
        if not notification.viewed: 
            notification.viewed = True

    db.session.commit()

    return render_template('notification_admin.html', notifications=notifications,sector_code=sector_code)

@app.route('/sector_register', methods=['GET', 'POST'])
def sector_register():

    return render_template('sector_register.html')



@app.route('/signup_administrator', methods=['GET', 'POST'])
def signup_administrator():
    sector_code = request.args.get('code')

    if request.method == 'POST':
        name = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        col1 = request.form.get('c1')
        col2 = request.form.get('c2')
        col3 = request.form.get('c3')
        col4 = request.form.get('c4')
        col5 = request.form.get('c5')
        code = request.args.get('code')

        if not name or len(name) < 3:
            flash("اسم المستخدم قصير جدًا. يجب أن يكون 3 أحرف على الأقل.", "error")
        elif not email or not re.match(r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$', email):
            flash("صيغة البريد الإلكتروني غير صحيحة.", "error")
        elif  administrator2.query.filter_by(email=email).first():
            flash("البريد الإلكتروني مستخدم بالفعل.", "error")
        elif not password or len(password) < 6:
            flash("كلمة المرور قصيرة جدًا. يجب أن تكون على الأقل 6 أحرف.", "error")
        else:
         
            new_admin =  administrator2(
                name=name,
                email=email,
                col1=col1,
                col2=col2,
                col3=col3,
                col4=col4,
                col5=col5,
                code=code,
                password=password
            )
            db.session.add(new_admin)
            db.session.commit()
            new_notification = Notification_owner(
            subject="New Administrator Added",
            message=f"Administrator {name} has been added successfully.",
            company_code=sector_code,
            timestamp=datetime.now()
            
            )
            db.session.add(new_notification)
            db.session.commit()

            new_notification = Notification_organizer(
            subject="New Administrator Added",
            message=f"Administrator {name} has been added successfully.",
            company_code=sector_code
            ,
            timestamp=datetime.now()
            )
            db.session.add(new_notification)
            db.session.commit()

            new_notification = Notification_admin(
                subject=f"Hello {name}",
                message = f"Welcome to Noor Academy for Sustainable Development, where we focus on nurturing human potential sustainably. We wish you a successful and continuous learning journey filled with growth.",
                id_user=new_admin.id,
                company_code=sector_code,
                timestamp=datetime.now()
            )
            db.session.add(new_notification)
            db.session.commit()


            session['user_id'] = new_admin.id
            session['sector_code'] = code

             
            return redirect('/administrator')

    admin_data = ColAdministrator.query.filter_by(code=sector_code).first()
    if not admin_data:
        return "Invalid code", 404

    return render_template(
        'signup_administrator.html',
        c1=admin_data.c1,
        c2=admin_data.c2,
        c3=admin_data.c3,
        c4=admin_data.c4,
        c5=admin_data.c5
    )

@app.route('/employee_profile', methods=['GET', 'POST'])
def employee_profile():
    employee_id = session.get('employee_id')
    sector_code = session.get('sector_code')
    if not sector_code or not employee_id:
        return redirect(url_for('access'))
    employee_data = Employee6.query.filter_by(id=employee_id).first()
    col1 = DepartmentEmployee1.query.filter_by(code=sector_code,id_admin=employee_data.id_admin).all()
    col2 = DepartmentEmployee.query.filter_by(code=sector_code,id_admin=employee_data.id_admin).all()
    col3 = Administrative_Title.query.filter_by(code=sector_code,id_admin=employee_data.id_admin).all()
    col4 = JobTitle.query.filter_by(code=sector_code,id_admin=employee_data.id_admin).all()
    col5 = specializationEmployee.query.filter_by(code=sector_code,id_admin=employee_data.id_admin).all()
    employee_name = ColEmployee.query.filter_by(code=sector_code,id_admin=employee_data.id_admin).first()
    
    print(employee_data.id_admin)
    
    CompanyNames = CompanyName.query.filter_by(code=sector_code).all()
    if not employee_id or not sector_code:
        return redirect(url_for('access')) 
    if not employee_data:
        return redirect('/')

    if request.method == 'POST':
        email = request.form.get('email')
        phone = request.form.get('phone')
        email_pattern = r"(^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$)"
        if not re.match(email_pattern, email):
            flash("البريد الإلكتروني غير صالح")
            return redirect(request.url) 
        phone_pattern = r"^\d+$"
        if not re.match(phone_pattern, phone):
            flash("رقم الهاتف غير صالح. يجب أن يحتوي على أرقام فقط")
            return redirect(request.url)
        employee_data.username = request.form.get('username')
        employee_data.email = request.form.get('email')
        birthdate_str = request.form.get('birthdate')
        employee_data.birthdate = datetime.strptime(birthdate_str, '%Y-%m-%d').date()
        employee_data.major = request.form.get('major')
        employee_data.phone = request.form.get('phone')
        employee_data.country = request.form.get('country')
        employee_data.province = request.form.get('province')

        employee_data.col1 = request.form.get('col1')
        employee_data.col2 = request.form.get('col2')
        employee_data.col3 = request.form.get('col3')
        employee_data.col4 = request.form.get('col4')
        employee_data.col5 = request.form.get('col5')
        employee_data.branch = request.form.get('branch')
       
        db.session.commit()

    return render_template('employee_profile.html', sector_code=sector_code,
                           employee_data=employee_data,
                           col1=col1, col2=col2, 
                           col3=col3, col4=col4, 
                           col5=col5, employee_name=employee_name, 
                           CompanyNames=CompanyNames)



@app.route('/Studentprofile', methods=['GET', 'POST'])
def Studentprofile():
    student_id=session.get('student_id')
    sector_code = session.get('sector_code')
    if not sector_code or not student_id:
        return redirect(url_for('access'))
    student_data = SectorStudent8.query.filter_by(id=student_id).first() 
    student_name = ColStudent.query.filter_by(code=sector_code,id_admin=student_data.id_admin).first()
    print(student_data.id_admin)
    if not student_id or not sector_code:
        return redirect(url_for('access')) 
    col1 = yearStudent.query.filter_by(code=sector_code,id_admin=student_data.id_admin).all()
    col2 = DepartmentStudent.query.filter_by(code=sector_code,id_admin=student_data.id_admin).all()
    col3 = ClassStudent.query.filter_by(code=sector_code,id_admin=student_data.id_admin).all()
    col4 = GradeStudent.query.filter_by(code=sector_code,id_admin=student_data.id_admin).all()
    CompanyNames = CompanyName.query.filter_by(code=sector_code).all()
    col5 = specializationStudent.query.filter_by(code=sector_code,id_admin=student_data.id_admin).all()

    if request.method == 'POST':
        student_data.username = request.form.get('username')
        student_data.email = request.form.get('email')
        email = request.form.get('email')
        phone = request.form.get('phone')
        email_pattern = r"(^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$)"
        if not re.match(email_pattern, email):
            flash("البريد الإلكتروني غير صالح")
            return redirect(request.url) 
        phone_pattern = r"^\d+$"
        if not re.match(phone_pattern, phone):
            flash("رقم الهاتف غير صالح. يجب أن يحتوي على أرقام فقط")
            return redirect(request.url)
        birthdate_str = request.form.get('birthdate')
        student_data.birthdate = datetime.strptime(birthdate_str, '%Y-%m-%d').date()
        student_data.stage = request.form.get('stage')
        student_data.university = request.form.get('university')
        student_data.major = request.form.get('major')
        student_data.phone = request.form.get('phone')
        student_data.country = request.form.get('country')
        student_data.province = request.form.get('province')
        student_data.col1 = request.form.get('col1')
        student_data.col2 = request.form.get('col2')
        student_data.col3 = request.form.get('col3')
        student_data.col4 = request.form.get('col4')
        student_data.col5 = request.form.get('col5')
        db.session.commit()
    return render_template('Studentprofile.html',sector_code=sector_code,student_data=student_data,student_name=student_name,
    col1=col1,col2=col2,col3=col3,col4=col4,col5=col5,CompanyNames=CompanyNames)

@app.route('/delete_job_title/<int:job_id>', methods=['GET'])
def delete_job_title(job_id):
    job = JobTitle.query.get(job_id)
    if job:
        db.session.delete(job)
        db.session.commit()
    return redirect(url_for('employee'))

@app.route('/delete_specialization/<int:specialization_id>', methods=['GET'])
def delete_specialization(specialization_id):
    specialization = specializationEmployee.query.get(specialization_id)
    if specialization:
        db.session.delete(specialization)
        db.session.commit()
    return redirect(url_for('employee'))

@app.route('/signup_student', methods=['GET', 'POST'])
def signup_student():
    
    sector_code = request.args.get('sector_code') 
    id_admin = request.args.get('id')  
    print(sector_code)
    print(id_admin)
  
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')

        # التحقق من المدخلات
        if not username or len(username) < 3:
            flash("اسم المستخدم قصير جدًا. يجب أن يكون 3 أحرف على الأقل.", "error")
        elif not email or not re.match(r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$', email):
            flash("صيغة البريد الإلكتروني غير صحيحة.", "error")
        elif SectorStudent8.query.filter_by(email=email).first():
            flash("البريد الإلكتروني مستخدم بالفعل.", "error")
        elif not password or len(password) < 6:
            flash("كلمة المرور قصيرة جدًا. يجب أن تكون على الأقل 6 أحرف.", "error")
        else:
           
            new_sector = SectorStudent8(
                username=username,
                email=email,
                password=password,
                code=sector_code,
                id_admin=id_admin,
                accepted=False
            )
            db.session.add(new_sector)
            db.session.commit()
            new_notification = Notification_admin(
                subject=f"New Student Added",
                message=f"A new student {username}, has been added to Noor Academy for Sustainable Development. The student is currently awaiting your approval.",
                id_user=id_admin,
                company_code=sector_code,
                timestamp=datetime.now()
            )
            db.session.add(new_notification)
            db.session.commit()

            session['student_id'] = new_sector.id
            session['sector_code'] = new_sector.code

            return redirect(url_for('Studentprofile'))

    departments = DepartmentStudent.query.filter_by(code=sector_code).all()
    ClassStudents = ClassStudent.query.filter_by(code=sector_code).all()

    return render_template('signup_student.html', sector_code=sector_code,id=id_admin, departments=departments, ClassStudents=ClassStudents)

@app.route('/validate_code', methods=['POST'])
def validate_code():
    user_code = request.form.get('codeInput')
    for i, char in enumerate(user_code):
            if char.isalpha():
                sector_code = user_code[:i]  
                person_code = user_code[i + 1:]  
                user_type = char.upper()  
                break

    sector = Sector.query.filter_by(code=sector_code).first()  
    if user_code and user_code.lower().startswith('pa'):  # تحقق من البادئة "pa" (بغض النظر عن حالة الأحرف)
        try:
            person_code = int(user_code[2:])  # استخرج الرقم بعد "pa" وحاول تحويله إلى عدد صحيح
            return redirect(url_for('parent_grades', user_id=person_code))  # توجيه إلى parent_grades مع تمرير user_id
        except ValueError:
            print("Invalid code format. Please enter 'pa' followed by a number.", "error")
            return redirect(url_for('sector_register'))  # أو صفحة مناسبة في حالة وجود تنسيق غير صحيح
    if sector:
        if user_type == 'S': 
            return redirect(url_for('signup_student', sector_code=sector_code, id=person_code))
        elif user_type == 'E':  
            return redirect(url_for('signup_eml', sector_code=sector_code, id=person_code))
        else:
            if user_type == 'A':
                return redirect(url_for('signup_administrator', code=sector_code))
            elif user_type == 'M':  
                return redirect(url_for('signup_organizer', code=sector_code))
            else:
                print("Invalid person code.")
                return redirect(url_for('sector_register'))
    else:
        print("Sector code is not valid. Please try again.", "error")
        return redirect(url_for('sector_register'))


@app.route('/validate_code1', methods=['POST'])
def validate_code1():
    
    user_code = request.form.get('codeInput')

    if user_code == 'O':
        return redirect(url_for('login_organizer')) 
    elif user_code == 'A':
        return redirect(url_for('login_administrator')) 
    elif user_code == 'S':
        return redirect(url_for('login_Student'))  
    elif user_code == 'E':
        return redirect(url_for('login_employee'))  
    else:
        flash("الرمز غير صالح", "error")



@app.route('/login_organizer', methods=['GET', 'POST'])
def login_organizer():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']

        organizer = organizer2.query.filter_by(email=email).first()
        if not organizer:
            flash("البريد الإلكتروني غير مسجل في النظام", "error")
            return render_template('login_organizer.html')

        if organizer.password != password:
            flash("كلمة المرور غير صحيحة", "error")
            return render_template('login_organizer.html')

        session['organizer_id'] = organizer.id
        session['sector_code'] = organizer.code

        return redirect(url_for('organizer'))

    return render_template('login_organizer.html')
    
    
@app.route('/login_administrator', methods=['GET', 'POST'])
def login_administrator():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']

        admin =  administrator2.query.filter_by(email=email).first()
        if not admin:
            flash("البريد الإلكتروني غير مسجل في النظام", "error")
            return render_template('login_administrator.html')

        if admin.password != password:
            flash("كلمة المرور غير صحيحة", "error")
            return render_template('login_administrator.html')

        session['user_id'] = admin.id
        session['sector_code'] = admin.code

        return redirect('/administrator')

    return render_template('login_administrator.html')


@app.route('/login_Student', methods=['GET', 'POST'])
def login_Student():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']

        student = SectorStudent8.query.filter_by(email=email).first()
        if not student:
            flash("البريد الإلكتروني غير مسجل في النظام.", "error")
            return render_template('login_Student.html')

        if student.password != password:
            flash("كلمة المرور غير صحيحة.", "error")
            return render_template('login_Student.html')


        session['student_id'] = student.id
        session['sector_code'] = student.code

        return redirect(url_for('Studentprofile'))

    return render_template('login_Student.html')

@app.route('/login_employee', methods=['GET', 'POST'])
def login_employee():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']

        employee = Employee6.query.filter_by(email=email).first()

        if not employee:
            flash("البريد الإلكتروني غير مسجل في النظام.", "error")
            return render_template('login_employee.html')

        if employee.password != password:
            flash("كلمة المرور غير صحيحة.", "error")
            return render_template('login_employee.html')

        session['employee_id'] = employee.id
        session['employee_code'] = employee.code

        
        return redirect('/employee_profile')

    return render_template('login_employee.html')


@app.route('/organizer', methods=['GET', 'POST'])
def organizer():
    organizer_id = session.get('organizer_id')
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    unread_notifications = Notification_organizer.query.filter_by(company_code=sector_code, viewed=False).all()
    if unread_notifications:
        notification_dot_display = True
    else:
        notification_dot_display = False
    if not sector_code or not (organizer_id):
        return redirect(url_for('access'))
    admin_data = ColOrganizer.query.filter_by(code=sector_code).first()
    organizer_data = organizer2.query.filter_by(id=organizer_id).first()
    administrators =  administrator2.query.filter_by(code=sector_code).all()
    sector_id = session.get('sector_id')
    sector_code = session.get('sector_code')
    column_names = ColAdministrator.query.filter_by(code=sector_code).first()
    company_names = CompanyName.query.filter_by(code=sector_code).all()
    print(company_names)
    col_data = ColAdministrator.query.filter_by(code=sector_code).first()
    if request.method == 'POST' and request.form.get('company_name')!=None:
        company_name = request.form.get('company_name')

        if company_name and sector_code:
            existing_company = CompanyName.query.filter_by(name=company_name, code=sector_code).first()
            if existing_company:
                return render_template('owner.html',notification_dot_display=notification_dot_display,sector_code=sector_code, company_names=company_names, error="اسم الشركة وكودها موجودان بالفعل")
            else:
                try:
                    new_company = CompanyName(name=company_name, code=sector_code)
                    db.session.add(new_company)
                    db.session.commit()
                    return redirect(url_for('organizer')) 
                except exc.IntegrityError as e:
                    db.session.rollback()
                    return render_template('organizer.html',notification_dot_display=notification_dot_display,sector_code=sector_code, company_names=company_names, error=f"خطأ في قاعدة البيانات: {e}")
                except Exception as e:
                    db.session.rollback()
                    return render_template('organizer.html',notification_dot_display=notification_dot_display,sector_code=sector_code, company_names=company_names, error=f"حدث خطأ: {e}")
    if 'excel_file' in request.files:
        excel_file = request.files['excel_file']
        if excel_file.filename != '':
            filename = secure_filename(excel_file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            excel_file.save(filepath)

            try:
                workbook = load_workbook(filepath)
                sheet = workbook.active  

                for row in sheet.iter_rows(min_row=2):
                    name = row[0].value
                    print(name) 
                    print(sector_code)

                    existing_company = CompanyName.query.filter_by(name=name, code=sector_code).first()

                    if existing_company:
                        print(f"Duplicate entry detected: {name} with code {sector_code}")
                        continue 

                    new_company = CompanyName(name=name, code=sector_code)
                    db.session.add(new_company)

                db.session.commit()
            except exc.IntegrityError as e:
                db.session.rollback()
                print(f"Database integrity error: {e}")
            except Exception as e:
                db.session.rollback()
                print(f"An error occurred: {e}")
                print(f"Error processing Excel file: {e}")
                return render_template('organizer.html',unread_notifications=unread_notifications,organizer=organizer_data,administrators=administrators, columns=column_names,sector_code=sector_code,col_data=col_data,  error=f"Error processing Excel file: {e}")
            finally:
                os.remove(filepath)
    if request.method == 'POST' and  request.form.get('name'):


            new_name = request.form['name']
            new_email = request.form['email']
            organizer_data.col1 = request.form.get('c1', organizer_data.col1)
            organizer_data.col2 = request.form.get('c2', organizer_data.col2)
            organizer_data.col3 = request.form.get('c3', organizer_data.col3)
            organizer_data.col4 = request.form.get('c4', organizer_data.col4)
            organizer_data.col5 = request.form.get('c5', organizer_data.col5)

            organizer_data.name = new_name
            organizer_data.email = new_email
            
            db.session.commit() 
    if not col_data:
        col_data = ColAdministrator(code=sector_code, c1='', c2='', c3='', c4='', c5='')
        db.session.add(col_data)
        db.session.commit()
    if request.method == 'POST':
        col_data.c1 = request.form.get('column1', col_data.c1)
        col_data.c2 = request.form.get('column2', col_data.c2)
        col_data.c3 = request.form.get('column3', col_data.c3)
        col_data.c4 = request.form.get('column4', col_data.c4)
        col_data.c5 = request.form.get('column5', col_data.c5)
        db.session.commit()
        return redirect(url_for('organizer'))

     
    
    return render_template('organizer.html',notification_dot_display=notification_dot_display,organizer_id=organizer_id,organizer=organizer_data, columns=column_names,sector_code=sector_code,col_data=col_data,administrators=administrators,  company_names=company_names
    ,
    c1=admin_data.c1,
    c2=admin_data.c2,
    c3=admin_data.c3,
    c4=admin_data.c4,
    c5=admin_data.c5)    
    

@app.route('/Education_entity', methods=['GET', 'POST'])
def Education_entityr():
    return render_template('Education_entity.html')
@app.route('/delete_student/<int:student_id>', methods=['GET','POST'])
def delete_student(student_id):
    student = SectorStudent8.query.get_or_404(student_id)
    db.session.delete(student)
    db.session.commit()
    return redirect(url_for('administrator')) 
@app.route('/administrator', methods=['GET', 'POST'])
def administrator():

    user_id = session.get('user_id')
    organizer_id = session.get('organizer_id')
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    unaccepted_students = SectorStudent8.query.filter_by(accepted=False, code=sector_code,id_admin=user_id).all()
    accepted_students = SectorStudent8.query.filter_by(accepted=True, code=sector_code,id_admin=user_id).all()
    if not sector_code or not (user_id):
        return redirect(url_for('access'))
    col_name = ColAdministrator.query.filter_by(code=sector_code).first()
   
    col_data =  administrator2.query.filter_by(id=user_id).first()
    unread_notifications = Notification_admin.query.filter_by(company_code=sector_code, viewed=False,id_user=user_id).all()
    print(unread_notifications)
    if unread_notifications:
        notification_dot_display = True
    else:
        notification_dot_display = False
 
    if request.method == 'POST':
     
        col_data.name = request.form.get('name', col_data.name)
        col_data.email = request.form.get('email', col_data.email)
        col_data.col1 = request.form.get('c1', col_data.col1)
        col_data.col2 = request.form.get('c2', col_data.col2)
        col_data.col3 = request.form.get('c3', col_data.col3)
        col_data.col4 = request.form.get('c4', col_data.col4)
        col_data.col5 = request.form.get('c5', col_data.col5)

        db.session.commit()


        return redirect(url_for('administrator',notification_dot_display=notification_dot_display))

    return render_template('administrator.html',notification_dot_display=notification_dot_display, sector_code=sector_code, col_data=col_data,col_name=col_name,unaccepted_students=unaccepted_students,
        accepted_students=accepted_students,user_id=user_id
    )

@app.route('/accept_student/<int:student_id>', methods=['GET','POST'])
def accept_student(student_id):
    student = SectorStudent8.query.get(student_id)
    if student:
        student.accepted = True
        db.session.commit()
    return redirect(url_for('administrator'))

@app.route('/delete_sector/<int:sector_id>', methods=['GET'])
def delete_sector(sector_id):
    sector = organizer2.query.get_or_404(sector_id)
    db.session.delete(sector)
    db.session.commit()
    return redirect(url_for('show_org'))

@app.route('/delete_company/<int:company_id>')
def delete_company(company_id):

    company = CompanyName.query.get_or_404(company_id)

    db.session.delete(company)
    db.session.commit()

    return redirect(url_for('organizer'))

@app.route('/delete_year/<int:company_id>')
def delete_year(company_id):

    year_Student = yearStudent.query.get_or_404(company_id)
    
    db.session.delete(year_Student)
    db.session.commit()

    return redirect(url_for('student'))

@app.route('/delete_grades/<int:grade_id>')
def delete_grades(grade_id):

    year_Student = GradeStudent.query.get_or_404(grade_id)
    
    db.session.delete(year_Student)
    db.session.commit()

    return redirect(url_for('student'))

@app.route('/delete_specializations/<int:company_id>')
def delete_specializations(company_id):
   
    year_Student = specializationStudent.query.get_or_404(company_id)
    
    db.session.delete(year_Student)
    db.session.commit()

    return redirect(url_for('student'))

@app.route('/delete_class/<int:class_id>')
def delete_class(class_id):
    year_Student = ClassStudent.query.get_or_404(class_id)
    
    db.session.delete(year_Student)
    db.session.commit()

    return redirect(url_for('student'))


@app.route('/delete-administrative-title/<int:title_id>', methods=['GET', 'POST'])
def delete_administrative_title(title_id):
    title_to_delete = Administrative_Title.query.get_or_404(title_id)

    try:
        db.session.delete(title_to_delete)
        db.session.commit()
    except Exception as e:
        db.session.rollback()

    return redirect(url_for('employee'))

@app.route('/employee', methods=['GET', 'POST'])
def employee():
    user_id = session.get('user_id')
    organizer_id = session.get('organizer_id')
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    sector_code_display = f"{sector_code}E{user_id}"  
    if not sector_code or not (user_id):
        return redirect(url_for('access'))
    
    student_data = ColEmployee.query.filter_by(code=sector_code,id_admin=user_id).first()
    if not student_data:
        student_data = ColEmployee(code=sector_code,id_admin=user_id, c1="", c2="", c3="", c4="", c5="")
        db.session.add(student_data)
        db.session.commit()
    
    if request.method == 'POST':
        new_c1_value = request.form.get('col1')
        new_c2_value = request.form.get('col2')
        new_c3_value = request.form.get('col3')
        new_c4_value = request.form.get('col4')
        new_c5_value = request.form.get('col5')

        if new_c1_value is not None:
            student_data.c1 = new_c1_value
            db.session.commit()
        
        
        if new_c2_value is not None:
            student_data.c2 = new_c2_value
            db.session.commit()

        if new_c3_value is not None:
            student_data.c3 = new_c3_value
            db.session.commit()

        if new_c4_value is not None:
            student_data.c4 = new_c4_value
            db.session.commit()
        if new_c5_value is not None:
            student_data.c5 = new_c5_value
            db.session.commit()

    if request.method == 'POST':
        job_title = request.form.get('job_title')
        if job_title:
            try:
                existing_job = JobTitle.query.filter_by(name=job_title, code=sector_code , id_admin=user_id).first()
                if not existing_job:
                    new_job = JobTitle(name=job_title, code=sector_code, id_admin=user_id)
                    db.session.add(new_job)
                    db.session.commit()
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")

        if 'job_excel_file' in request.files:
            job_excel_file = request.files['job_excel_file']
            if job_excel_file.filename != '':
                filename = secure_filename(job_excel_file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                job_excel_file.save(filepath)

                try:
                    workbook = load_workbook(filepath)
                    sheet = workbook.active
                    for row in sheet.iter_rows(min_row=2):
                        name = row[0].value
                        if not JobTitle.query.filter_by(name=name, code=sector_code, id_admin=user_id).first():
                            new_job = JobTitle(name=name, code=sector_code, id_admin=user_id)
                            db.session.add(new_job)
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    print(f"Error: {e}")
                finally:
                    os.remove(filepath)

        specialization_name = request.form.get('specialization_name')
        if specialization_name:
            try:
                existing_specialization = specializationEmployee.query.filter_by(name=specialization_name, code=sector_code, id_admin=user_id).first()
                if not existing_specialization:
                    new_specialization = specializationEmployee(name=specialization_name, code=sector_code, id_admin=user_id)
                    db.session.add(new_specialization)
                    db.session.commit()
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")

        if 'specialization_excel_file' in request.files:
            specialization_excel_file = request.files['specialization_excel_file']
            if specialization_excel_file.filename != '':
                filename = secure_filename(specialization_excel_file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                specialization_excel_file.save(filepath)

                try:
                    workbook = load_workbook(filepath)
                    sheet = workbook.active
                    for row in sheet.iter_rows(min_row=2):
                        name = row[0].value
                        if not specializationEmployee.query.filter_by(name=name, code=sector_code, id_admin=user_id).first():
                            new_specialization = specializationEmployee(name=name, code=sector_code, id_admin=user_id)
                            db.session.add(new_specialization)
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    print(f"Error: {e}")
                finally:
                    os.remove(filepath)

        administrative_title_name = request.form.get('administrative_title_name')
        if administrative_title_name:
            try:
                existing_specialization = Administrative_Title.query.filter_by(name=administrative_title_name, code=sector_code, id_admin=user_id).first()
                if not existing_specialization:
                    new_specialization = Administrative_Title(name=administrative_title_name, code=sector_code, id_admin=user_id)
                    db.session.add(new_specialization)
                    db.session.commit()
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")

        if 'administrative_titles_file' in request.files:
            administrative_titles_file = request.files['administrative_titles_file']
            if administrative_titles_file.filename != '':
                filename = secure_filename(administrative_titles_file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                administrative_titles_file.save(filepath)

                try:
                    workbook = load_workbook(filepath)
                    sheet = workbook.active
                    for row in sheet.iter_rows(min_row=2):
                        name = row[0].value
                        if not Administrative_Title.query.filter_by(name=name, code=sector_code, id_admin=user_id).first():
                            new_specialization = Administrative_Title(name=name, code=sector_code, id_admin=user_id)
                            db.session.add(new_specialization)
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    print(f"Error: {e}")
                finally:
                    os.remove(filepath)
            return redirect(url_for('employee'))

        department_name = request.form.get('department_name')
        if department_name:
            try:
                existing_specialization = DepartmentEmployee.query.filter_by(name=department_name, code=sector_code, id_admin=user_id).first()
                if not existing_specialization:
                    new_specialization = DepartmentEmployee(name=department_name, code=sector_code, id_admin=user_id)
                    db.session.add(new_specialization)
                    db.session.commit()
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")

        if 'departments_file' in request.files:
            departments_file = request.files['departments_file']
            if departments_file.filename != '':
                filename = secure_filename(departments_file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                departments_file.save(filepath)

                try:
                    workbook = load_workbook(filepath)
                    sheet = workbook.active
                    for row in sheet.iter_rows(min_row=2):
                        name = row[0].value
                        if not DepartmentEmployee.query.filter_by(name=name, code=sector_code, id_admin=user_id).first():
                            new_specialization = DepartmentEmployee(name=name, code=sector_code, id_admin=user_id)
                            db.session.add(new_specialization)
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    print(f"Error: {e}")
                finally:
                    os.remove(filepath)
            return redirect(url_for('employee'))

        department_name = request.form.get('department_name1')
        if department_name:
            try:
                existing_specialization = DepartmentEmployee1.query.filter_by(name=department_name, code=sector_code, id_admin=user_id).first()
                if not existing_specialization:
                    new_specialization = DepartmentEmployee1(name=department_name, code=sector_code, id_admin=user_id)
                    db.session.add(new_specialization)
                    db.session.commit()
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")

        if 'departments_file1' in request.files:
            departments_file = request.files['departments_file1']
            if departments_file.filename != '':
                filename = secure_filename(departments_file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                departments_file.save(filepath)

                try:
                    workbook = load_workbook(filepath)
                    sheet = workbook.active
                    for row in sheet.iter_rows(min_row=2):
                        name = row[0].value
                        if not DepartmentEmployee1.query.filter_by(name=name, code=sector_code, id_admin=user_id).first():
                            new_specialization = DepartmentEmployee1(name=name, code=sector_code, id_admin=user_id)
                            db.session.add(new_specialization)
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    print(f"Error: {e}")
                finally:
                    os.remove(filepath)
            return redirect(url_for('employee'))

    DepartmentEmployees1 = DepartmentEmployee1.query.filter_by(code=sector_code,id_admin=user_id).all()
    DepartmentEmployees = DepartmentEmployee.query.filter_by(code=sector_code,id_admin=user_id).all()
    Administrative_Titles = Administrative_Title.query.filter_by(code=sector_code,id_admin=user_id).all()
    job_titles = JobTitle.query.filter_by(code=sector_code,id_admin=user_id).all()
    specializations_list = specializationEmployee.query.filter_by(code=sector_code,id_admin=user_id).all()

    return render_template('employee.html',sector_code_display=sector_code_display,student_data=student_data, job_titles=job_titles,sector_code=sector_code, specializations_list=specializations_list,Administrative_Titles=Administrative_Titles,DepartmentEmployees=DepartmentEmployees
    ,DepartmentEmployees1=DepartmentEmployees1)

@app.route('/student', methods=['GET', 'POST'])
def student():
    user_id = session.get('user_id')
    organizer_id = session.get('organizer_id')
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    if not sector_code or not (user_id):
        return redirect(url_for('access'))
    
    student_data = ColStudent.query.filter_by(code=sector_code,id_admin=user_id).first()

    if not student_data:
        student_data = ColStudent(code=sector_code,id_admin=user_id, c1="", c2="", c3="", c4="", c5="")
        db.session.add(student_data)
        db.session.commit()
    sector_code_display = f"{sector_code}S{user_id}"  
    if request.method == 'POST':
        # تحديث القيم
        updated = False
        new_values = {
            'c1': request.form.get('col1'),
            'c2': request.form.get('col2'),
            'c3': request.form.get('col3'),
            'c4': request.form.get('col4'),
            'c5': request.form.get('col5')
        }
        for col, value in new_values.items():
            if value is not None:
                setattr(student_data, col, value)
                updated = True
        
        if updated:
            db.session.commit()

    classes = ClassStudent.query.filter_by(code=sector_code,id_admin=user_id).all()
    year_Student = yearStudent.query.filter_by(code=sector_code,id_admin=user_id).all()
    specializations = specializationStudent.query.filter_by(code=sector_code,id_admin=user_id).all()
    grades = GradeStudent.query.filter_by(code=sector_code,id_admin=user_id).all()
    departments = DepartmentStudent.query.filter_by(code=sector_code,id_admin=user_id).all()

    if request.method == 'POST':
        company_name = request.form.get('company_name')
        if company_name:
            if sector_code:
                existing_company = yearStudent.query.filter_by(name=company_name, code=sector_code,id_admin=user_id).first()
                if existing_company:
                    return render_template('student.html',departments=departments,sector_code=sector_code,classes=classes,year_Student=year_Student,grades=grades, specializations=specializations, error="اسم الشركة وكودها موجودان بالفعل")
                else:
                    try:
                        new_company = yearStudent(name=company_name, code=sector_code,id_admin=user_id)
                        db.session.add(new_company)
                        db.session.commit()
                    except exc.IntegrityError as e:
                        db.session.rollback()
                       
                    except Exception as e:
                        db.session.rollback()
                       

       
        specialization_name = request.form.get('specialization_name')
        if specialization_name:
            try:
                existing_job = specializationStudent.query.filter_by(name=specialization_name, code=sector_code,id_admin=user_id).first()
                if not existing_job:
                    new_job = specializationStudent(name=specialization_name, code=sector_code,id_admin=user_id)
                    db.session.add(new_job)
                    db.session.commit()
            except Exception as e:
                db.session.rollback()
                print(f"Error: {e}")
            return redirect(url_for('student'))

        grade_name = request.form.get('grade_name')
        if grade_name:
            if sector_code:
                existing_specialization = GradeStudent.query.filter_by(name=grade_name, code=sector_code,id_admin=user_id).first()

                if existing_specialization:
                    return render_template('student.html',departments=departments,sector_code=sector_code,classes=classes,year_Student=year_Student,grades=grades, specializations=specializations, error="التخصص موجود بالفعل")
                else:
                    try:
                        new_specialization = GradeStudent(name=grade_name, code=sector_code,id_admin=user_id)
                        db.session.add(new_specialization)
                        db.session.commit()
                        return redirect(url_for('student'))
                    except exc.IntegrityError as e:
                        db.session.rollback()
                        return render_template('student.html',departments=departments,sector_code=sector_code,classes=classes, year_Student=year_Student, grades=grades,specializations=specializations, error=f"خطأ في قاعدة البيانات: {e}")
                    except Exception as e:
                        db.session.rollback()
                        return render_template('student.html',departments=departments,sector_code=sector_code,classes=classes, year_Student=year_Student, grades=grades,specializations=specializations, error=f"حدث خطأ: {e}")
        
        class_name = request.form.get('class_name')
        if class_name:
            if sector_code:
                existing_specialization = ClassStudent.query.filter_by(name=class_name, code=sector_code,id_admin=user_id).first()

                if existing_specialization:
                    return render_template('student.html', departments=departments,sector_code=sector_code,year_Student=year_Student,grades=grades, specializations=specializations,classes=classes,error="التخصص موجود بالفعل")
                else:
                    try:
                        new_specialization = ClassStudent(name=class_name, code=sector_code,id_admin=user_id)
                        db.session.add(new_specialization)
                        db.session.commit()
                        return redirect(url_for('student'))
                    except exc.IntegrityError as e:
                        db.session.rollback()
                        return render_template('student.html',departments=departments,sector_code=sector_code,classes=classes, year_Student=year_Student, grades=grades,specializations=specializations, error=f"خطأ في قاعدة البيانات: {e}")
                    except Exception as e:
                        db.session.rollback()
                        return render_template('student.html',departments=departments,sector_code=sector_code,year_Student=year_Student, grades=grades,classes=classes,specializations=specializations, error=f"حدث خطأ: {e}")

        department_name = request.form.get('department_name')
        if department_name:
            if sector_code:
                existing_specialization = DepartmentStudent.query.filter_by(name=department_name, code=sector_code,id_admin=user_id).first()

                if existing_specialization:
                    return render_template('student.html', departments=departments,sector_code=sector_code,year_Student=year_Student,grades=grades, specializations=specializations,classes=classes,error="التخصص موجود بالفعل")
                else:
                    try:
                        new_specialization =DepartmentStudent(name=department_name, code=sector_code,id_admin=user_id)
                        db.session.add(new_specialization)
                        db.session.commit()
                        return redirect(url_for('student'))
                    except exc.IntegrityError as e:
                        db.session.rollback()
                        return render_template('student.html',departments=departments,sector_code=sector_code,classes=classes, year_Student=year_Student, grades=grades,specializations=specializations, error=f"خطأ في قاعدة البيانات: {e}")
                    except Exception as e:
                        db.session.rollback()
                        return render_template('student.html',sector_code=sector_code,departments=departments,year_Student=year_Student, grades=grades,classes=classes,specializations=specializations, error=f"حدث خطأ: {e}")
        
        if 'excel_file' in request.files:
            excel_file = request.files['excel_file']
            if excel_file.filename != '':
                filename = secure_filename(excel_file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                excel_file.save(filepath)

                try:
                    workbook = load_workbook(filepath)
                    sheet = workbook.active

                    for row in sheet.iter_rows(min_row=2):
                        name = row[0].value
                        existing_company = yearStudent.query.filter_by(name=name, code=sector_code,id_admin=user_id).first()

                        if not existing_company:
                            new_company = yearStudent(name=name, code=sector_code,id_admin=user_id)
                            db.session.add(new_company)

                    db.session.commit()

                except IntegrityError as e:
                    db.session.rollback()
                    print(f"Database integrity error: {e}")
                except Exception as e:
                    db.session.rollback()
                    print(f"An error occurred: {e}")
                finally:
                    os.remove(filepath)
                return redirect(url_for('student'))

        if 'excel_file_specialization' in request.files:
            specialization_excel_file = request.files['excel_file_specialization']
            if specialization_excel_file.filename != '':
                filename = secure_filename(specialization_excel_file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                specialization_excel_file.save(filepath)

                try:
                    workbook = load_workbook(filepath)
                    sheet = workbook.active
                    for row in sheet.iter_rows(min_row=2):
                        name = row[0].value
                        if not specializationStudent.query.filter_by(name=name, code=sector_code,id_admin=user_id).first():
                            new_specialization = specializationStudent(name=name, code=sector_code,id_admin=user_id)
                            db.session.add(new_specialization)
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    print(f"Error: {e}")
                finally:
                    os.remove(filepath)
                return redirect(url_for('student'))


        if 'excel_file_grade' in request.files:
            excel_file_grade = request.files['excel_file_grade']
            if excel_file_grade.filename != '':
                filename = secure_filename(excel_file_grade.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                excel_file_grade.save(filepath)

                try:
                    workbook = load_workbook(filepath)
                    sheet = workbook.active

                    for row in sheet.iter_rows(min_row=2):
                        grade_name = row[0].value  

                        existing_grade = GradeStudent.query.filter_by(name=grade_name, code=sector_code,id_admin=user_id).first()

                        if not existing_grade:
                            new_grade = GradeStudent(name=grade_name, code=sector_code,id_admin=user_id)
                            db.session.add(new_grade)

                    db.session.commit()

                except IntegrityError as e:
                    db.session.rollback()
                    print(f"Database integrity error: {e}")
                except Exception as e:
                    db.session.rollback()
                    print(f"An error occurred: {e}")
                finally:
                    os.remove(filepath)
                return redirect(url_for('student'))

        if 'excel_file_class' in request.files:
            excel_file_class = request.files['excel_file_class']
            if excel_file_class.filename != '':
                filename = secure_filename(excel_file_class.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                excel_file_class.save(filepath)

                try:
                    workbook = load_workbook(filepath)
                    sheet = workbook.active

                    for row in sheet.iter_rows(min_row=2):
                        class_name = row[0].value  

                        existing_class = ClassStudent.query.filter_by(name=class_name, code=sector_code,id_admin=user_id).first()

                        if not existing_class:
                            new_class = ClassStudent(name=class_name, code=sector_code,id_admin=user_id)
                            db.session.add(new_class)

                    db.session.commit()

                except IntegrityError as e:
                    db.session.rollback()
                    print(f"Database integrity error: {e}")
                except Exception as e:
                    db.session.rollback()
                    print(f"An error occurred: {e}")
                finally:
                    os.remove(filepath)
                return redirect(url_for('student'))
                
        if 'excel_file_department' in request.files:
            excel_file_department = request.files['excel_file_department']
            if excel_file_department.filename != '':
                filename = secure_filename(excel_file_department.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                excel_file_department.save(filepath)

                try:
                    workbook = load_workbook(filepath)
                    sheet = workbook.active

                    for row in sheet.iter_rows(min_row=2):
                        department_name = row[0].value  

                        existing_department = DepartmentStudent.query.filter_by(name=department_name, code=sector_code,id_admin=user_id).first()

                        if not existing_department:
                            new_department = DepartmentStudent(name=department_name, code=sector_code,id_admin=user_id)
                            db.session.add(new_department)

                    db.session.commit()

                except InterruptedError as e:
                    db.session.rollback()
                    print(f"Database integrity error: {e}")
                except Exception as e:
                    db.session.rollback()
                    print(f"An error occurred: {e}")
                finally:
                    os.remove(filepath)
                return redirect(url_for('student'))

    return render_template('student.html',sector_code_display=sector_code_display,student_data=student_data,sector_code=sector_code,departments=departments,year_Student=year_Student, specializations=specializations,grades=grades,classes=classes)
@app.route('/logout')
def logout():
    
    session.clear()
    return redirect(url_for('index'))
@app.route('/delete_department/<int:department_id>', methods=['GET'])
def delete_department(department_id):
    department = DepartmentStudent.query.get(department_id)
    if department:
        db.session.delete(department)
        db.session.commit()
    return redirect(url_for('student'))

@app.route('/delete_department1/<int:department_id>', methods=['GET'])
def delete_department1(department_id):
    department = DepartmentEmployee.query.get(department_id)
    if department:
        db.session.delete(department)
        db.session.commit()
    return redirect(url_for('employee'))

@app.route('/delete_department2/<int:department_id>', methods=['GET'])
def delete_department2(department_id):
    department = DepartmentEmployee1.query.get(department_id)
    if department:
        db.session.delete(department)
        db.session.commit()
    return redirect(url_for('employee'))

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/owner', methods=['GET', 'POST'])
def owner():
    
    sector_id = session.get('sector_id')
    sector_code = session.get('sector_code')
    if not sector_id or not sector_code:
        return redirect(url_for('access')) 
    sector = Sector.query.filter_by(id=sector_id).first()
    unread_notifications = Notification_owner.query.filter_by(company_code=sector_code, viewed=False).all()
    if unread_notifications:
        notification_dot_display = True
    else:
        notification_dot_display = False
    company_names = CompanyName.query.filter_by(code=sector_code).all()
    col_data = ColOrganizer.query.filter_by(code=sector_code).first()
    if request.method == 'POST' and request.form.get('foundation_name') is not None:
        foundation_name = request.form.get('foundation_name')
        address = request.form.get('address')
        phone_number = request.form.get('phone_number')
        state = request.form.get('state')
        country = request.form.get('country')
        description = request.form.get('description')
        email = request.form.get('email')

       
        if len(foundation_name) < 3:
            return "Error: Company name must be at least 3 characters.", 400
        if len(address) < 3:
            return "Error: address must be at least 3 characters.", 400
        if not phone_number.isdigit():
            return "Error: Phone number must contain only digits.", 400
       

        # تحديث البيانات في قاعدة البيانات
        sector.foundation_name = foundation_name
        sector.address = address
        sector.phone_number = phone_number
        sector.state = state
        sector.country = country
        sector.description = description or None
        sector.email=email
        db.session.commit()
        return redirect(url_for('owner'))

    if request.method == 'POST' and 'profile_image' in request.files:
        profile_image = request.files['profile_image']
        if profile_image and allowed_file(profile_image.filename):
            filename = f"{sector.code}.png" 
            
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            
            if os.path.exists(file_path):
                os.remove(file_path)
           
            profile_image.save(file_path)
            
          
            sector.image_path = f'{filename}'
            db.session.commit()
            
            return redirect(url_for('owner'))

    if not col_data:
        col_data = ColOrganizer(code=sector_code, c1='', c2='', c3='', c4='', c5='')
        db.session.add(col_data)
        db.session.commit()
    if request.method == 'POST':
        col_data.c1 = request.form.get('column1', col_data.c1)
        col_data.c2 = request.form.get('column2', col_data.c2)
        col_data.c3 = request.form.get('column3', col_data.c3)
        col_data.c4 = request.form.get('column4', col_data.c4)
        col_data.c5 = request.form.get('column5', col_data.c5)
        db.session.commit()
        return redirect(url_for('owner'))
    if not sector_id:
        return redirect(url_for('signup_sector')) 
    
    
    all_sectors = organizer2.query.filter_by(code=sector_code).all()
    visible_columns = ColumnPreference.query.filter_by(code=sector_code, visible=True).all()
    visible_column_names = [col.column_name for col in visible_columns]
    if not sector:
        return redirect(url_for('signup_sector'))  
    sector_id = session.get('sector_id')
    sector_code = session.get('sector_code')
    if request.method == 'POST' and request.form.get('company_name')!=None:
        company_name = request.form.get('company_name')

        if company_name and sector_code:
            existing_company = CompanyName.query.filter_by(name=company_name, code=sector_code).first()
            if existing_company:
                return render_template('owner.html',sector_code=sector_code, sector=sector, all_sectors=all_sectors, company_names=company_names, error="اسم الشركة وكودها موجودان بالفعل")
            else:
                try:
                    new_company = CompanyName(name=company_name, code=sector_code)
                    db.session.add(new_company)
                    db.session.commit()
                    return redirect(url_for('owner')) 
                except exc.IntegrityError as e:
                    db.session.rollback()
                    return render_template('owner.html',sector_code=sector_code, sector=sector, all_sectors=all_sectors, company_names=company_names, error=f"خطأ في قاعدة البيانات: {e}")
                except Exception as e:
                    db.session.rollback()
                    return render_template('owner.html',sector_code=sector_code, sector=sector, all_sectors=all_sectors, company_names=company_names, error=f"حدث خطأ: {e}")

    

    if 'excel_file' in request.files:
        excel_file = request.files['excel_file']
        if excel_file.filename != '':
            filename = secure_filename(excel_file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            excel_file.save(filepath)

            try:
                workbook = load_workbook(filepath)
                sheet = workbook.active  

                for row in sheet.iter_rows(min_row=2):
                    name = row[0].value
                    print(name) 
                    print(sector_code)

                    existing_company = CompanyName.query.filter_by(name=name, code=sector_code).first()

                    if existing_company:
                        print(f"Duplicate entry detected: {name} with code {sector_code}")
                        continue 

                    new_company = CompanyName(name=name, code=sector_code)
                    db.session.add(new_company)

                db.session.commit()
            except exc.IntegrityError as e:
                db.session.rollback()
                print(f"Database integrity error: {e}")
            except Exception as e:
                db.session.rollback()
                print(f"An error occurred: {e}")
                print(f"Error processing Excel file: {e}")
                return render_template('owner.html',col_data=col_data,sector_code=sector_code, sector=sector, all_sectors=all_sectors, error=f"Error processing Excel file: {e}")
            finally:
                os.remove(filepath) 

    return render_template('owner.html',notification_dot_display=notification_dot_display,col_data=col_data,sector_code=sector_code, sector=sector, all_sectors=all_sectors, company_names=company_names)



@app.route('/show_org', methods=['GET', 'POST'])
def show_org():
    organizer_id = session.get('organizer_id')
    sector_id = session.get('sector_id')
    sector_code = session.get('sector_code')
    if not sector_id or not sector_code:
        return redirect(url_for('access')) 
    sector = Sector.query.filter_by(id=sector_id).first()

    company_names = CompanyName.query.filter_by(code=sector_code).all()
    col_data = ColOrganizer.query.filter_by(code=sector_code).first()
    all_sectors = organizer2.query.filter_by(code=sector_code).all()
    visible_columns = ColumnPreference.query.filter_by(code=sector_code, visible=True).all()
    visible_column_names = [col.column_name for col in visible_columns]
 
    return render_template('show_org.html',col_data=col_data,sector_code=sector_code, sector=sector, all_sectors=all_sectors, company_names=company_names,organizer_id=organizer_id,sector_id=sector_id)

@app.route('/success')
def success():
    username = request.args.get('username')
    email = "user@example.com"  

    preferences = ColumnPreference.query.filter_by(email=email).all()
    visible_columns = {pref.column_name: pref.visible for pref in preferences}

    personal_data = Personal.query.all()
    return render_template('success.html', 
                           username=username, 
                           personal_data=personal_data, 
                           visible_columns=visible_columns)


@app.route('/manager', methods=['GET', 'POST'])
def manager():
    user_id = session.get('user_id')
    organizer_id = session.get('organizer_id')
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    if not sector_code or not (organizer_id or sector_id ):
        return redirect(url_for('access'))
    employee_name = ColEmployee.query.filter_by(code=sector_code).first()
    employees = Employee6.query.filter_by(code=sector_code).all()
    return render_template('manager.html', employees=employees,employee_name=employee_name,sector_code=sector_code,organizer_id=organizer_id,sector_id=sector_id)


@app.route('/manager_adm', methods=['GET', 'POST'])
def manager_adm():
    user_id = session.get('user_id')
    organizer_id = session.get('organizer_id')
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    if not sector_code or not ( user_id):
        return redirect(url_for('access'))
    employee_name = ColEmployee.query.filter_by(code=sector_code,id_admin=user_id,).first()
    employees = Employee6.query.filter_by(code=sector_code,id_admin=user_id).all()
    return render_template('manager_adm.html', employees=employees,employee_name=employee_name,sector_code=sector_code)

@app.route('/managertomanger_adm', methods=['GET', 'POST'])
def managertomanger_adm():
    all_data = head.query.all()  
    for data in all_data:
        print(f"ID: {data.id}, Username: {data.username}, Email: {data.email}, Phone: {data.phone}")
    user_id = session.get('user_id')
    organizer_id = session.get('organizer_id')
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    if not sector_code or not (user_id):
        return redirect(url_for('access'))
    employee_name = ColEmployee.query.filter_by(code=sector_code,id_admin=user_id).first()
    managers = ManagerEmployee.query.filter_by(code=sector_code,id_admin=user_id).all() 
    return render_template('managertomanger_adm.html', managers=managers,sector_code=sector_code,employee_name=employee_name)


@app.route('/assign_manager1_adm', methods=['POST'])
def assign_manager1_adm():
    if request.method == 'POST':

        manager_id = request.form.get('manager')  
    
        employee_ids = request.form.getlist('employees')  

        if manager_id and employee_ids:
            manager = ManagerEmployee.query.get(manager_id)  
            if manager:
                manager_to_update = Employee6.query.filter_by(email=manager.email).first()
                if manager_to_update:
                    manager_to_update.is_head = "ttt"
                    db.session.commit()
                
                for emp_id in employee_ids:
                    employee = ManagerEmployee.query.get(emp_id)
                    if employee:

    
                        employee.manager_id = manager.id
                        employee.manager_name = manager.username
                        db.session.commit()

                existing_manager = head.query.filter_by(email=manager.email).first()
                if not existing_manager:
                    manager_employee = head(
                        id_admin=manager.id_admin,
                        code=manager.code,
                        username=manager.username,
                        email=manager.email,
                        password=manager.password,
                        birthdate=manager.birthdate,
                        stage=manager.stage,
                        university=manager.university,
                        major=manager.major,
                        phone=manager.phone,
                        country=manager.country,
                        province=manager.province,
                        col1=manager.col1,
                        col2=manager.col2,
                        col3=manager.col3,
                        col4=manager.col4,
                        col5=manager.col5,
                        branch=manager.branch
                    )
                    db.session.add(manager_employee)  
                    db.session.commit() 

                return redirect('/managertomanger_adm') 
    return redirect(url_for('managertomanger_adm'))

@app.route('/assign_manager_adm', methods=['GET', 'POST'])
def assign_manager_adm():
    if request.method == 'POST':
        manager_id = request.form.get('manager')  
        employee_ids = request.form.getlist('employees')  
        
        if manager_id and employee_ids:
            manager = Employee6.query.get(manager_id) 
            if manager:
                if not manager.is_manager:  
                    manager.is_manager = True
                    db.session.commit()
                for emp_id in employee_ids:
                    employee = Employee6.query.get(emp_id)
                    if employee:
                  
                        employee.manager_id = manager.id
                        employee.manager_name = manager.username
                    
                        db.session.commit()

                existing_manager = ManagerEmployee.query.filter_by(email=manager.email).first()
                print(manager.id_admin)
                print(manager.id_admin)
                print(manager.id_admin)
                if not existing_manager:
                    manager_employee = ManagerEmployee(
                        id_admin=manager.id_admin,
                        code=manager.code,
                        username=manager.username,
                        email=manager.email,
                        password=manager.password,
                        birthdate=manager.birthdate,
                        stage=manager.stage,
                        university=manager.university,
                        major=manager.major,
                        phone=manager.phone,
                        country=manager.country,
                        province=manager.province,
                        col1=manager.col1,
                        col2=manager.col2,
                        col3=manager.col3,
                        col4=manager.col4,
                        col5=manager.col5,
                        branch=manager.branch
                    )
                    db.session.add(manager_employee)  
                    db.session.commit() 

                return redirect('/manager_adm') 
            else:
                return "Manager not found", 404 
        else:
            return "Error: Manager or Employees not selected", 400  
    
    return render_template('manager_adm.html')

@app.route('/show_stu', methods=['GET', 'POST'])
def show_stu():
    user_id = session.get('user_id')
    organizer_id = session.get('organizer_id')
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    if not sector_code or not (organizer_id or sector_id):
        return redirect(url_for('access'))
    employees = SectorStudent8.query.filter_by(code=sector_code).all()
    return render_template('show_stu.html', employees=employees,sector_code=sector_code,organizer_id=organizer_id,sector_id=sector_id)

@app.route('/private_bank')
def private_bank():
    lesson_id = request.args.get('id')
    student_id = session.get('student_id')

    if student_id and lesson_id:
        # التحقق مما إذا كان السجل موجودًا بالفعل
        existing_record = prvatebank.query.filter_by(lesson_id=lesson_id, student_id=student_id).first()

        if existing_record:
            return f"Lesson ID {lesson_id} already saved for student ID {student_id}"
        else:
            # إنشاء السجل الجديد وإضافته
            student_lesson = prvatebank(lesson_id=lesson_id, student_id=student_id)
            db.session.add(student_lesson)
            db.session.commit()

            return f"Lesson ID {lesson_id} saved for student ID {student_id}"
    else:
        return "Error: Student ID not found in session or Lesson ID is missing."
@app.route('/privatebank')
def privatebank():
    student_id = session.get('student_id')
    if not student_id:
        return redirect(url_for('access'))  # Redirect if student_id is missing

    # استعلام عن جميع الدروس التي قام الطالب بتسجيلها في جدول prvatebank
    student_lessons_records = prvatebank.query.filter_by(student_id=student_id).all()

    # استخراج معرفات الدروس من سجلات prvatebank
    lesson_ids = [record.lesson_id for record in student_lessons_records]

    # استعلام عن تفاصيل الدروس من جدول Lesson8 بناءً على معرفات الدروس
    lessons = Lesson8.query.filter(Lesson8.id.in_(lesson_ids)).all()

    return render_template('privatebank.html', lessons=lessons)

@app.route('/permission', methods=['GET', 'POST'])
def permission():
    user_id = session.get('user_id')
    organizer_id = session.get('organizer_id')
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    if not sector_code or not (organizer_id or sector_id or user_id):
        return redirect(url_for('access'))
    employees = Employee6.query.filter_by(code=sector_code,id_admin=user_id).all()
    return render_template('permission.html', employees=employees,sector_code=sector_code,organizer_id=organizer_id,sector_id=sector_id)

@app.route('/permission_employee', methods=['GET', 'POST'])
def permission_employee():
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    employee_id = session.get('employee_id')
    employee_data = Employee6.query.filter_by(id=employee_id).first()
    if not sector_code or not (employee_id):
        return redirect(url_for('access'))
    employees = Employee6.query.filter_by(code=sector_code,id_admin=employee_data.id_admin).all()
    
    return render_template('permission_employee.html', employees=employees,sector_code=sector_code,sector_id=sector_id,employee_data=employee_data)

@app.route('/update_permission', methods=['POST'])
def update_permission():
    employee_id = request.form.get('id')
    field = request.form.get('field')
    new_value = request.form.get('value') == 'True'

    # البحث عن الموظف في قاعدة البيانات
    employee = Employee6.query.get(employee_id)
    if not employee:
        return "Employee not found", 404

    # تحديث الصلاحية بناءً على الحقل
    if field == "permission_add_exam":
        employee.permission_add_exam = new_value
    elif field == "permission_add_permission":
        employee.permission_add_permission = new_value
    else:
        return "Invalid field", 400

    # حفظ التغييرات
    db.session.commit()

    # إعادة التوجيه للصفحة السابقة
    return redirect(url_for('permission'))

@app.route('/delete_stu/<int:sector_id>', methods=['GET'])
def delete_stu(sector_id):
    sector = SectorStudent8.query.get_or_404(sector_id)
    db.session.delete(sector)
    db.session.commit()
    return redirect(url_for('show_stu'))

@app.route('/show_adm', methods=['GET', 'POST'])
def show_adm():
    user_id = session.get('user_id')
    organizer_id = session.get('organizer_id')
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    if not sector_code or not (organizer_id or sector_id):
        return redirect(url_for('access'))
    employees = administrator2.query.filter_by(code=sector_code).all()
    admin_data = ColOrganizer.query.filter_by(code=sector_code).first()
    column_names = ColAdministrator.query.filter_by(code=sector_code).first()
    return render_template('show_adm.html', employees=employees,sector_code=sector_code,columns=column_names,sector_id=sector_id,organizer_id=organizer_id,
    c1=admin_data.c1,
    c2=admin_data.c2,
    c3=admin_data.c3,
    c4=admin_data.c4,
    c5=admin_data.c5)

@app.route('/delete_adm/<int:sector_id>', methods=['GET'])
def delete_adm(sector_id):
    sector = administrator2.query.get_or_404(sector_id)
    db.session.delete(sector)
    db.session.commit()
    return redirect(url_for('show_adm'))

@app.route('/show_emp', methods=['GET', 'POST'])
def show_emp():
    user_id = session.get('user_id')
    organizer_id = session.get('organizer_id')
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    if not sector_code or not (organizer_id or sector_id):
        return redirect(url_for('access'))
    employee_name = ColEmployee.query.filter_by(code=sector_code).first()
    employees = Employee6.query.filter_by(code=sector_code).all()
    return render_template('show_emp.html', employees=employees,employee_name=employee_name,sector_code=sector_code,organizer_id=organizer_id,sector_id=sector_id)

@app.route('/delete_emp/<int:sector_id>', methods=['GET'])
def delete_emp(sector_id):
    sector = Employee6.query.get_or_404(sector_id)
    db.session.delete(sector)
    db.session.commit()
    return redirect(url_for('show_emp'))

@app.route('/managertomanger', methods=['GET', 'POST'])
def managertomanger():
    all_data = head.query.all()  
    for data in all_data:
        print(f"ID: {data.id}, Username: {data.username}, Email: {data.email}, Phone: {data.phone}")
    user_id = session.get('user_id')
    organizer_id = session.get('organizer_id')
    sector_code = session.get('sector_code')
    sector_id = session.get('sector_id')
    if not sector_code or not (organizer_id or sector_id or user_id):
        return redirect(url_for('access'))
    employee_name = ColEmployee.query.filter_by(code=sector_code).first()
    managers = ManagerEmployee.query.filter_by(code=sector_code).all() 
    return render_template('managertomanger.html', managers=managers,sector_code=sector_code,employee_name=employee_name,organizer_id=organizer_id,sector_id=sector_id)

@app.route('/assign_manager1', methods=['POST'])
def assign_manager1():
    if request.method == 'POST':
        manager_id = request.form.get('manager')  
        employee_ids = request.form.getlist('employees')  

        if manager_id and employee_ids:
            manager = ManagerEmployee.query.get(manager_id) 
            if manager:
                manager_to_update = Employee6.query.filter_by(email=manager.email).first()
                if manager_to_update:
                    manager_to_update.is_head = "ttt"
                    db.session.commit()

                for emp_id in employee_ids:
                    employee = ManagerEmployee.query.get(emp_id)
                    if employee:
                        employee.manager_id = manager.id
                        employee.manager_name = manager.username
                        db.session.commit()

                # التحقق من وجود المدير في جدول head
                existing_manager = head.query.filter_by(email=manager.email).first()
                if not existing_manager:
                    manager_employee = head(
                        id_admin=manager.id_admin,
                        code=manager.code,
                        username=manager.username,
                        email=manager.email,
                        password=manager.password,
                        birthdate=manager.birthdate,
                        stage=manager.stage,
                        university=manager.university,
                        major=manager.major,
                        phone=manager.phone,
                        country=manager.country,
                        province=manager.province,
                        col1=manager.col1,
                        col2=manager.col2,
                        col3=manager.col3,
                        col4=manager.col4,
                        col5=manager.col5,
                        branch=manager.branch
                    )
                    db.session.add(manager_employee)  
                    db.session.commit() 

                return redirect('/managertomanger') 
    
    return redirect(url_for('managertomanger'))

@app.route('/assign_manager', methods=['GET', 'POST'])
def assign_manager():
    if request.method == 'POST':
        manager_id = request.form.get('manager')  
        employee_ids = request.form.getlist('employees')  
        
        if manager_id and employee_ids:
            manager = Employee6.query.get(manager_id) 
            if manager:
                # تعيين الموظف كمدير (is_manager = True)
                if not manager.is_manager:  # إذا لم يكن المدير الحالي بالفعل مديرًا
                    manager.is_manager = True
                    db.session.commit()

                for emp_id in employee_ids:
                    employee = Employee6.query.get(emp_id)
                    if employee:
                        # تعيين المدير للموظف
                        employee.manager_id = manager.id
                        employee.manager_name = manager.username
                        db.session.commit()

                # نقل المدير إلى جدول ManagerEmployee إذا لم يكن موجودًا بالفعل
                existing_manager = ManagerEmployee.query.filter_by(email=manager.email).first()
                if not existing_manager:
                    manager_employee = ManagerEmployee(
                        id_admin=manager.id_admin,
                        code=manager.code,
                        username=manager.username,
                        email=manager.email,
                        password=manager.password,
                        birthdate=manager.birthdate,
                        stage=manager.stage,
                        university=manager.university,
                        major=manager.major,
                        phone=manager.phone,
                        country=manager.country,
                        province=manager.province,
                        col1=manager.col1,
                        col2=manager.col2,
                        col3=manager.col3,
                        col4=manager.col4,
                        col5=manager.col5,
                        branch=manager.branch
                    )
                    db.session.add(manager_employee)  
                    db.session.commit() 

                return redirect('/manager') 
            else:
                return "Manager not found", 404 
        else:
            return "Error: Manager or Employees not selected", 400  
    
    return render_template('manager.html')

load_dotenv()

GOOGLE_API_KEY = os.environ.get('GOOGLE_API_KEY')
if not GOOGLE_API_KEY:
    raise ValueError("No API key found, please set it in the .env file")

genai.configure(api_key=GOOGLE_API_KEY) # Changed here
model = genai.GenerativeModel('gemini-2.0-flash-thinking-exp-01-21')

# Test the connection to see if it works
try:
    response = model.generate_content("What is the capital of France?")
    print("API Connection Successful")
    print(response.text)
except Exception as e:
    print(f"Error during API Connection test: {e}")


def generate_true_false_questions(text,num):
    try:
        prompt = f"""
        بناءً على النص التالي، قم بإنشاء {num} أسئلة صح أو خطأ (True/False) مباشرة وواضحة.
        يجب أن تكون الأسئلة متنوعة ومختلفة وتستخلص من المعلومات الموجودة في النص تحديدًا.
        يجب أن تكون الأسئلة واقعية وتعكس محتوى النص بشكل دقيق.
        يجب أن يكون كل سؤال وإجابته في سطر منفصل، بحيث يبدأ السطر بالسؤال، يتبعه فاصلة (,) ثم الإجابة الصحيحة (True أو False). مترقمش الاسئله اللى طالعه او تحط اى علامه
        ونوع اعمل اسئله true على false مش كله false
        مثال:
        هل الشمس نجم؟,True
        هل الأرض مسطحة؟,False
        ...إلخ
        النص: {text}
        """
        response = model.generate_content([prompt])
        return response.text
    except Exception as e:
        print(f"Error generating true/false questions: {e}")
        return None

def generate_matching_questions(text,num):
    try:
        prompt = f"""
        بناءً على النص التالي، قم بإنشاء 4 أسئلة مطابقة (Matching Questions) مباشرة وواضحة.
        يجب أن تكون الأسئلة متنوعة ومختلفة وتستخلص من المعلومات الموجودة في النص تحديدًا.
        يجب أن تكون الأسئلة واقعية وتعكس محتوى النص بشكل دقيق.
        لكل سؤال، قم بإنشاء 4 أزواج من النصوص (عمودين)، حيث العمود الأول يمثل الجزء الأول من السؤال، والعمود الثاني يمثل الإجابة الصحيحة.
        يجب أن يكون تنسيق الإجابة كالتالي JSON:
        [
          {{
              "question_text": "[نص الجزء الأول من السؤال 1]",
              "answer_text": "[نص الإجابة الصحيحة 1]",
              "question_text1": "[نص الجزء الأول من السؤال 2]",
              "answer_text1": "[نص الإجابة الصحيحة 2]",
              "question_text2": "[نص الجزء الأول من السؤال 3]",
               "answer_text2": "[نص الإجابة الصحيحة 3]",
               "question_text3": "[نص الجزء الأول من السؤال 4]",
               "answer_text3": "[نص الإجابة الصحيحة 4]"
           }}
        ]
        مثال:
        [
            {{
            "question_text": "مصر",
            "answer_text": "القاهرة",
            "question_text1": "سوريا",
            "answer_text1": "دمشق",
             "question_text2": "السعودية",
             "answer_text2":"الرياض",
              "question_text3": "فلسطين",
             "answer_text3": "القدس"
        }}
        ]
        النص: {text}
        """
        response = model.generate_content([prompt])
        return response.text
    except Exception as e:
        print(f"Error generating matching questions: {e}")
        return None
def generate_mcq_questions(text,num):
    try:
        prompt = f"""
        بناءً على النص التالي، قم بإنشاء {num} أسئلة اختيار من متعدد (Multiple Choice Questions) مباشرة وواضحة.
        يجب أن تكون الأسئلة متنوعة ومختلفة وتستخلص من المعلومات الموجودة في النص تحديدًا.
        يجب أن تكون الأسئلة واقعية وتعكس محتوى النص بشكل دقيق.
        لكل سؤال، قم بتضمين 4 خيارات إجابة، واحدة منها صحيحة على الأقل.
        يجب أن يكون تنسيق الإجابة كالتالي JSON:
        [
        {{
        "question": "[نص السؤال]",
        "options": ["الخيار1", "الخيار2", "الخيار3", "الخيار4"],
        "correct_answers": [رقم الخيار الصحيح (أو أرقام الخيارات الصحيحة إذا كانت أكثر من واحدة)]
        }},
        {{
        "question": "[نص السؤال]",
        "options": ["الخيار1", "الخيار2", "الخيار3", "الخيار4"],
        "correct_answers": [رقم الخيار الصحيح (أو أرقام الخيارات الصحيحة إذا كانت أكثر من واحدة)]
        }}
        ]
        مثال:
        [
        {{
        "question": "ما هي عاصمة مصر؟",
        "options": ["القاهرة", "الإسكندرية", "الأقصر", "أسوان"],
        "correct_answers": [1]
        }},
        {{
        "question": "من هو مكتشف الجاذبية الأرضية؟",
        "options": ["نيوتن", "أينشتاين", "غاليليو", "أديسون"],
        "correct_answers": [1]
        }}
        ]
        النص: {text}
        """
        response = model.generate_content([prompt])
        return response.text
    except Exception as e:
        print(f"Error generating MCQ questions: {e}")
        return None


def generate_multiple_multiple_choice_questions(text,num):
    try:
        prompt = f"""
        بناءً على النص التالي، قم بإنشاء {num} أسئلة اختيار من متعدد (Multiple Choice Questions) مباشرة وواضحة.
        يجب أن تكون الأسئلة متنوعة ومختلفة وتستخلص من المعلومات الموجودة في النص تحديدًا.
        يجب أن تكون الأسئلة واقعية وتعكس محتوى النص بشكل دقيق.
        لكل سؤال، قم بتضمين 4 خيارات إجابة، ويجب أن يكون هناك أكثر من خيار صحيح واحد على الأقل.
        يجب أن يكون تنسيق الإجابة كالتالي JSON:
        [
        {{
        "question": "[نص السؤال]",
        "options": ["الخيار1", "الخيار2", "الخيار3", "الخيار4"],
        "correct_answers": [أرقام الخيارات الصحيحة]
        }},
        {{
        "question": "[نص السؤال]",
        "options": ["الخيار1", "الخيار2", "الخيار3", "الخيار4"],
        "correct_answers": [أرقام الخيارات الصحيحة]
        }}
        ]
        مثال:
        [
        {{
        "question": "ما هي بعض المدن المصرية الهامة؟",
         "options": ["القاهرة", "الإسكندرية", "الأقصر", "أسوان"],
         "correct_answers": [1,2]
        }},
         {{
        "question": "ما هي بعض أنواع الفاكهة الحمضية؟",
        "options": ["التفاح", "البرتقال", "الموز", "الليمون"],
        "correct_answers": [2,4]
        }}
        ]
        النص: {text}
        """
        response = model.generate_content([prompt])
        return response.text
    except Exception as e:
        print(f"Error generating Multiple multiple choice questions: {e}")
        return None
def generate_fill_in_the_blank_questions(text,num):
    try:
        prompt = f"""
        بناءً على النص التالي، قم بإنشاء  4 أسئلة "أكمل الفراغ" (Fill in the blanks) مباشرة وواضحة.
        يجب أن تكون الأسئلة متنوعة ومختلفة وتستخلص من المعلومات الموجودة في النص تحديدًا.
        يجب أن تكون الأسئلة واقعية وتعكس محتوى النص بشكل دقيق.
        يجب أن يكون كل سؤال عبارة عن فراغ واحد فقط مع إجابة واحدة فقط.
        يجب أن يكون تنسيق الإجابة كالتالي JSON:
        [
          {{
            "question": "كلمة {{blank}}",
            "answers": ["إجابة"]
          }},
           {{
            "question": "كلمة {{blank}}",
            "answers": ["إجابة"]
           }},
            {{
            "question":"كلمة {{blank}}",
            "answers": ["إجابة"]
          }},
           {{
            "question": "كلمة {{blank}}",
            "answers": ["إجابة"]
           }}
         ]
        مثال:
        [
           {{
            "question": "عاصمة {{blank}}",
            "answers": ["مصر"]
           }},
            {{
            "question": "الجاذبية {{blank}}",
            "answers": ["نيوتن"]
             }},
           {{
            "question":"عدد {{blank}}",
            "answers": ["ايام"]
            }},
           {{
            "question":"مياه {{blank}}",
             "answers": ["مالحة"]
           }}
        ]
        النص: {text}
        """
        response = model.generate_content([prompt])
        return response.text
    except Exception as e:
        print(f"Error generating fill in the blank questions: {e}")
        return None
def generate_essay_questions(text,num):
    try:
        prompt = f"""
        بناءً على النص التالي، قم بإنشاء {num} أسئلة مقالية (Essay Questions) مباشرة وواضحة.
        يجب أن تكون الأسئلة متنوعة ومختلفة وتستخلص من المعلومات الموجودة في النص تحديدًا.
        يجب أن تكون الأسئلة واقعية وتعكس محتوى النص بشكل دقيق وتتطلب إجابة تفصيلية.
        يجب أن يكون تنسيق الإجابة كالتالي  قائمة من النصوص:
        [
          "[نص السؤال الأول]",
           "[نص السؤال الثاني]",
           "[نص السؤال الثالث]"
         ]
        مثال:
        [
            "اشرح أهمية موقع مصر الجغرافي.",
            "وضح كيف أثرت الحضارة المصرية القديمة على العالم.",
            "ما هي التحديات التي تواجه مصر في العصر الحديث؟"
          ]
        النص: {text}
        """
        response = model.generate_content([prompt])
        return response.text
    except Exception as e:
        print(f"Error generating essay questions: {e}")
        return None

def generate_sequence_questions(text,num):
    try:
      prompt = f"""
        بناءً على النص التالي، قم بإنشاء سؤال يتطلب ترتيب إجابات مع 4 خيارات إجابة مباشرة وواضحة.
        يجب أن يكون السؤال والإجابات متعلقة بالنص.
        يجب أن تكون الإجابات واقعية وتعكس محتوى النص بشكل دقيق.
        يجب أن يكون تنسيق الإجابة كالتالي JSON:
        {{
          "question": "[نص السؤال]",
          "answers": [
              {{ "text": "[نص الإجابة 1]", "index": 1 }},
              {{ "text": "[نص الإجابة 2]", "index": 2 }},
              {{ "text": "[نص الإجابة 3]", "index": 3 }},
              {{ "text": "[نص الإجابة 4]", "index": 4 }}
          ]
        }}
        مثال:
        {{
          "question": "رتب مراحل نمو النبات التالية:",
           "answers": [
             {{ "text": "البذرة", "index": 1 }},
             {{ "text": "الإنبات", "index": 2 }},
             {{ "text": "النمو", "index": 3 }},
             {{ "text": "الإزهار", "index": 4 }}
            ]
        }}
        النص: {text}
        """
      response = model.generate_content([prompt])
      return response.text
    except Exception as e:
        print(f"Error generating sequence questions: {e}")
        return None

# Functions to parse questions
def parse_true_false_questions(response_text):
    if not response_text:
        return []

    try:
        questions = []
        lines = response_text.strip().split('\n')
        seen_questions = set()  # Track unique questions
        for line in lines:
            parts = line.strip().split(',')
            if len(parts) == 2:
                question_text = parts[0].strip()
                answer_str = parts[1].strip()
                if answer_str.lower() == "true":
                    correct_answer = True
                elif answer_str.lower() == "false":
                    correct_answer = False
                else:
                    continue  # Skip if answer isn't valid

                # Check if the question has been seen
                if question_text not in seen_questions:
                    questions.append({"question": question_text, "answer": correct_answer})
                    seen_questions.add(question_text)
        return questions
    except Exception as e:
        print(f"Error parsing true/false questions: {e}")
        return []
def parse_matching_questions(response_text):
    if not response_text:
        return []
    try:
         cleaned_json = clean_json_response(response_text)
         questions = json.loads(cleaned_json)
         formatted_questions = []
         for q in questions:
              formatted_questions.append({
                 "question_text": q.get("question_text", ""),
                 "answer_text": q.get("answer_text", ""),
                 "question_text1": q.get("question_text1", ""),
                 "answer_text1": q.get("answer_text1", ""),
                  "question_text2": q.get("question_text2", ""),
                  "answer_text2": q.get("answer_text2", ""),
                  "question_text3": q.get("question_text3", ""),
                   "answer_text3": q.get("answer_text3", "")
              })

         return formatted_questions
    except Exception as e:
        print(f"Error parsing matching questions: {e}")
        return []
def clean_json_response(text):
    """
    حاول تنظيف النص الذي من المفترض أن يكون JSON.
    إزالة المقدمات والخواتم النصية المحتملة.
    """
    try:
        # ابحث عن بداية أول قوس مربع وبداية آخر قوس مربع
        start_index = text.find('[')
        end_index = text.rfind(']')

        if start_index != -1 and end_index != -1 and start_index < end_index:
            # استخراج جزء JSON المحتمل
            json_string = text[start_index:end_index + 1]

            # محاولة تحميل JSON للتأكد من أنه صالح.  إذا لم يكن كذلك، سيتم التقاط الخطأ.
            json.loads(json_string)
            return json_string
        else:
            return None  # لا يوجد JSON صالح
    except json.JSONDecodeError:
        return None  # JSON غير صالح

def parse_mcq_questions(response_text):
    if not response_text:
        return []
    try:
        cleaned_json = clean_json_response(response_text)
        if not cleaned_json:
            print("Error: No valid JSON found in response.")
            return []

        questions = json.loads(cleaned_json)
        formatted_questions = []
        for q in questions:
            options = q.get("options", [])  # احصل على قائمة الخيارات
            correct_indices = q.get("correct_answers", [])  # احصل على قائمة الأرقام الصحيحة

            # تحويل الأرقام الصحيحة إلى نصوص الخيارات الصحيحة
            correct_answers = []
            for index in correct_indices:
                if 0 < index <= len(options):  # تحقق من أن الرقم ضمن النطاق
                    correct_answers.append(options[index - 1])  # أضف النص المقابل للرقم
                else:
                    print(f"Warning: Invalid correct_answers index: {index}")

            formatted_questions.append({
                "question": q.get("question", ""),
                "options": options,
                "correct_answers": correct_answers  # استخدم النصوص الصحيحة هنا
            })
        return formatted_questions
    except json.JSONDecodeError as e:
        print(f"Error parsing MCQ questions: Invalid JSON - {e}")
        return []
    except Exception as e:
        print(f"Error parsing MCQ questions: {e}")
        return []

def parse_multiple_multiple_choice_questions(response_text):
    if not response_text:
        return []
    try:
        cleaned_json = clean_json_response(response_text)
        questions = json.loads(cleaned_json)
        formatted_questions = []
        for q in questions:
            formatted_questions.append({
                 "question": q.get("question",""),
                 "options": q.get("options",[]),
                 "correct_answers": [q.get("options", [])[index-1]  for index in q.get("correct_answers",[]) if 0 < index <= len(q.get("options", []))] if q.get("options", []) else []
            })
        return formatted_questions
    except Exception as e:
        print(f"Error parsing Multiple multiple choice questions: {e}")
        return []

def parse_fill_in_the_blank_questions(response_text):
    if not response_text:
        return []
    try:
        cleaned_json = clean_json_response(response_text)
        questions = json.loads(cleaned_json)
        formatted_questions = []
        for q in questions:
            formatted_questions.append({
                "question": q.get("question", ""),
                "answers": q.get("answers", [])
            })
        return formatted_questions
    except Exception as e:
        print(f"Error parsing fill in the blank questions: {e}")
        return []

def parse_essay_questions(response_text):
    if not response_text:
        return []
    try:
        cleaned_json = clean_json_response(response_text)
        questions = json.loads(cleaned_json)
        return [{"question": q} for q in questions]
    except Exception as e:
        print(f"Error parsing essay questions: {e}")
        return []
def parse_sequence_questions(response_text):
    if not response_text:
      return {}  # Return an empty dictionary

    try:
        cleaned_json = clean_json_response_s(response_text)
        if not cleaned_json:
            return {}  # Return empty dictionary if cleaning fails

        data = json.loads(cleaned_json)

        # Check if 'question' and 'answers' keys exist
        if not isinstance(data, dict) or 'question' not in data or 'answers' not in data:
            print("JSON data is missing 'question' or 'answers' keys, or is not a dictionary")
            return {}

        question_text = data.get("question", "")
        answers = data.get("answers", [])

        if not isinstance(answers, list):
            print("The 'answers' value is not a list")
            return {} # Return empty dictionary if answers is not a list


        formatted_answers = []
        for answer in answers:
          if not isinstance(answer, dict):
            print("An item in answers is not a dictionary")
            return {}
          formatted_answers.append({
              "text": answer.get("text",""),
               "index":answer.get("index","")
          })
        return {"question":question_text, "answers":formatted_answers}
    except json.JSONDecodeError as e:
        print(f"JSONDecodeError: {e}")
        print(f"Failed to parse: {response_text}")  # Print the problematic text
        return {}
    except Exception as e:
        print(f"Error parsing sequence questions: {e}")
        return {}

def clean_json_response_s(text):
    """
    Clean a text string to extract a valid JSON object.
    Remove leading/trailing text.
    """
    try:
        # Find the first and last curly braces
        start_index = text.find('{')
        end_index = text.rfind('}')

        if start_index != -1 and end_index != -1 and start_index < end_index:
            # Extract the potential JSON string
            json_string = text[start_index:end_index + 1]

            # Attempt to parse the JSON string to validate it
            json.loads(json_string)
            return json_string
        else:
            print("No valid JSON object found")
            return None
    except json.JSONDecodeError as e:
        print(f"JSON Decode Error in clean_json_response: {e}")  # Print the error
        print(f"Attempted to clean: {text}")  # Print the text being cleaned
        return None
    except Exception as e:
        print(f"Error cleaning JSON: {e}")
        return None
import random

def generate_difficulty():
    rand = random.randint(1, 100)
    
    if rand <= 70:
        return 'Medium' 
    elif rand <= 85:
        return 'Easy'   
    else:
        return 'Hard'   
@app.route('/classify', methods=['POST'])
def classify():
    try:
        lesson_id = request.args.get('lesson_id')  
        number_of_questions = request.args.get('number_of_questions')  
        print(lesson_id)
        data = request.get_json()
        if 'text' not in data or 'questionType' not in data:
            return jsonify({'error': 'Text and question type are required'}), 400

        text = data['text']
        
        question_type = data['questionType']

        if question_type == "true_false":
            response_text = generate_true_false_questions(text,number_of_questions)
            if not response_text:
                return jsonify({'error': 'Failed to generate true/false questions'}), 500
            print(response_text)
            questions_data = parse_true_false_questions(response_text)
            if not questions_data:
                return jsonify({'error': 'Failed to parse true/false questions'}), 500
            # Store Questions in the Database
            for q_data in questions_data:
                question_obj = Questionfort_or_f(
                    lesson_id=lesson_id,
                    difficulty=generate_difficulty(),
                    question_name=q_data['question'],
                    correct_answer=q_data['answer'],
        
                )
                question_obj.times = calculate_expected_time(question_obj)
                db.session.add(question_obj)
            db.session.commit()

            # Print all questions from the database
            all_questions = Questionfort_or_f.query.all()
            print("True/False Questions in the database:")
            for question in all_questions:
                if question.type == "true_false":
                    print(
                        f"ID: {question.question_id}, Text: {question.question_name}, Answer: {question.correct_answer}")

            redirect_url = url_for('edit_question_true_or_false', lesson_id=lesson_id)

            return jsonify(
                {
                    'message': 'True/False questions generated and saved to database successfully',
                    'questions': questions_data,
                    'redirect_url': redirect_url
                })

        elif question_type == "mcq":
            response_text = generate_mcq_questions(text,number_of_questions)
            print(response_text)
            if not response_text:
                return jsonify({'error': 'Failed to generate MCQ questions'}), 500

            questions_data = parse_mcq_questions(response_text)
            if not questions_data:
                return jsonify({'error': 'Failed to parse MCQ questions'}), 500

            # Store MCQ Questions and Answers in the Database
            for q_data in questions_data:
                question_obj = Question_Multipleone_choice(
                    lesson_id=lesson_id,
                    question_name=q_data['question']
                )
              
                db.session.add(question_obj)
                db.session.flush()
                answers = []  # تهيئة القائمة قبل الحلقة
                for option in q_data["options"]:
                    is_correct = option in q_data["correct_answers"]
                    answer_obj = Answer_Multiple_Multiple_choice(
                        question_id=question_obj.question_id,
                        answer_text=option,
                        is_correct=is_correct
                    )
                    db.session.add(answer_obj)
                    answers.append(answer_obj)  

                expected_time = calculate_expected_time_MCQ(question_obj, answers)
                question_obj.times = expected_time  
            db.session.commit()

            all_questions = Question_Multipleone_choice.query.all()
            print("MCQ Questions in the database:")
            for question in all_questions:
                print(f"ID: {question.question_id}, Text: {question.question_name}")
                answers = Answer_Multiple_Multiple_choice.query.filter_by(question_id=question.question_id).all()
                for answer in answers:
                    print(f"    Answer ID: {answer.answer_id}, Text: {answer.answer_text}, Is Correct: {answer.is_correct}")
            redirect_url = url_for('edit_question_Multiple_single', lesson_id=lesson_id)
            return jsonify(
                {'message': 'MCQ questions generated and saved to database successfully', 'questions': questions_data,'redirect_url': redirect_url})
        elif question_type == "multiple_mcq":
             response_text = generate_multiple_multiple_choice_questions(text,number_of_questions)
             if not response_text:
                return jsonify({'error': 'Failed to generate multiple MCQ questions'}), 500

             questions_data = parse_multiple_multiple_choice_questions(response_text)
             if not questions_data:
                return jsonify({'error': 'Failed to parse multiple MCQ questions'}), 500
            # Store MCQ Questions and Answers in the Database
             for q_data in questions_data:
                question_obj = QuestionMultipleMultiple_choice(
                    lesson_id=lesson_id,
                    question_name=q_data['question']
                )
                answers = []  
                db.session.add(question_obj)
                db.session.flush()
                for option in q_data["options"]:
                     is_correct = option in q_data["correct_answers"]
                     answer_obj = Answer_Multiple_Multiple_choice(
                        question_id=question_obj.question_id,
                        answer_text=option,
                        is_correct=is_correct
                      )
                     db.session.add(answer_obj)
                     answers.append(answer_obj)  

                expected_time = calculate_expected_time_MCQ(question_obj, answers)
                question_obj.times = expected_time  
             db.session.commit()

             all_questions = QuestionMultipleMultiple_choice.query.all()
             print("Multiple MCQ Questions in the database:")
             for question in all_questions:
                  print(f"ID: {question.question_id}, Text: {question.question_name}")
                  answers = Answer_Multiple_Multiple_choice.query.filter_by(question_id=question.question_id).all()
                  for answer in answers:
                      print(f"    Answer ID: {answer.answer_id}, Text: {answer.answer_text}, Is Correct: {answer.is_correct}")

             redirect_url = url_for('edit_question_Multiple_Multiple', lesson_id=lesson_id)
             return jsonify(
                {'message': 'Multiple MCQ questions generated and saved to database successfully', 'questions': questions_data,'redirect_url': redirect_url})
        elif question_type == "fill_blank":
             response_text = generate_fill_in_the_blank_questions(text,number_of_questions)
             if not response_text:
                return jsonify({'error': 'Failed to generate fill in the blank questions'}), 500

             questions_data = parse_fill_in_the_blank_questions(response_text)
             if not questions_data:
                  return jsonify({'error': 'Failed to parse fill in the blank questions'}), 500

             question_obj = Question_Fillin_blank4(lesson_id=lesson_id)


             for index, q_data in enumerate(questions_data):
                 parts = q_data['question'].split(' {{blank}} ')
                 answers = q_data['answers']
                 
                 if index == 0:
                    question_obj.question_text = parts[0] if len(parts) > 0 else None
                    question_obj.answer_text = answers[0] if len(answers) > 0 else None
                 elif index == 1:
                     question_obj.question_text1 = parts[0] if len(parts) > 0 else None
                     question_obj.answer_text1 = answers[0] if len(answers) > 0 else None

                 elif index == 2:
                      question_obj.question_text2 = parts[0] if len(parts) > 0 else None
                      question_obj.answer_text2 = answers[0] if len(answers) > 0 else None
                 elif index == 3:
                      question_obj.question_text3 = parts[0] if len(parts) > 0 else None
                      question_obj.answer_text3 = answers[0] if len(answers) > 0 else None

             db.session.add(question_obj)
             question_obj.times = calculate_expected_time_fill(question_obj)
             db.session.commit()

             all_questions = Question_Fillin_blank4.query.all()
             print("Fill in the blank Questions in the database:")
             for question in all_questions:
                    print(
                        f"ID: {question.question_id}, Text: {question.question_text} ,{question.question_text1}, {question.question_text2}, {question.question_text3} Answers: {question.answer_text}, {question.answer_text1},{question.answer_text2},{question.answer_text3}")

             redirect_url = url_for('edit_fill_in_blank', lesson_id=lesson_id)
             return jsonify(
                {'message': 'fill in the blank questions generated and saved to database successfully',
                 'questions': questions_data,'redirect_url': redirect_url})

        elif question_type == "fill_blank_choice":
             response_text = generate_fill_in_the_blank_questions(text,number_of_questions)
             if not response_text:
                return jsonify({'error': 'Failed to generate fill in the blank questions'}), 500

             questions_data = parse_fill_in_the_blank_questions(response_text)
             if not questions_data:
                  return jsonify({'error': 'Failed to parse fill in the blank questions'}), 500

             question_obj = Question_Fillin_blank_choice2(lesson_id=lesson_id)  

             for index, q_data in enumerate(questions_data):
                 parts = q_data['question'].split(' {{blank}} ')
                 answers = q_data['answers']
                 
                 if index == 0:
                    question_obj.question_text = parts[0] if len(parts) > 0 else None
                    question_obj.answer_text = answers[0] if len(answers) > 0 else None
                 elif index == 1:
                     question_obj.question_text1 = parts[0] if len(parts) > 0 else None
                     question_obj.answer_text1 = answers[0] if len(answers) > 0 else None

                 elif index == 2:
                      question_obj.question_text2 = parts[0] if len(parts) > 0 else None
                      question_obj.answer_text2 = answers[0] if len(answers) > 0 else None
                 elif index == 3:
                      question_obj.question_text3 = parts[0] if len(parts) > 0 else None
                      question_obj.answer_text3 = answers[0] if len(answers) > 0 else None

             db.session.add(question_obj)
             question_obj.times = calculate_expected_time_fill(question_obj)
             db.session.commit()
             redirect_url = url_for('edit_fill_in_blank_choice', lesson_id=lesson_id)
             all_questions = Question_Fillin_blank_choice2.query.all()
             print("Fill in the blank Questions in the database:")
             for question in all_questions:
                    print(
                        f"ID: {question.question_id}, Text: {question.question_text} ,{question.question_text1}, {question.question_text2}, {question.question_text3} Answers: {question.answer_text}, {question.answer_text1},{question.answer_text2},{question.answer_text3}")


             return jsonify(
                {'message': 'fill in the blank questions generated and saved to database successfully',
                 'questions': questions_data,'redirect_url': redirect_url})
        elif question_type == "essay":
            response_text = generate_essay_questions(text,number_of_questions)
            if not response_text:
                 return jsonify({'error': 'Failed to generate essay questions'}), 500
            
            questions_data = parse_essay_questions(response_text)
            if not questions_data:
                return jsonify({'error': 'Failed to parse essay questions'}), 500
            
            # Store Essay Questions in the Database
            for q_data in questions_data:
                  question_obj = Article_question(
                      lesson_id=lesson_id,
                      question_text=q_data['question']
                  )
                  db.session.add(question_obj)
                  
            db.session.commit()
            
            # Print all essay questions from the database
            all_questions = Article_question.query.all()
            redirect_url = url_for('article', lesson_id=lesson_id)
            print("Essay Questions in the database:")
            for question in all_questions:
                print(f"ID: {question.question_id}, Text: {question.question_text}")
            
            return jsonify({
                 'message': 'Essay questions generated and saved to database successfully',
                 'questions': questions_data
                 ,'redirect_url': redirect_url
            })
        elif question_type == "sequence":
            response_text = generate_sequence_questions(text,number_of_questions)
            if not response_text:
              return jsonify({'error': 'Failed to generate sequence questions'}), 500
            
            question_data = parse_sequence_questions(response_text)
            if not question_data:
              return jsonify({'error': 'Failed to parse sequence questions'}), 500
            
            # Store sequence Questions and Answers in the Database
            question_obj = Questionsequence1(
                    lesson_id=lesson_id,
                    question_name=question_data['question']
                )
            db.session.add(question_obj)
            
            db.session.flush()
            answers = []  
            for answer in question_data['answers']:
              answer_obj = Answer_sequence(
                  question_id = question_obj.question_id,
                  answer_text = answer["text"],
                  index = answer["index"]
              )
              db.session.add(answer_obj)
              answers.append(answer_obj)  

            expected_time = calculate_expected_time_sequence(question_obj, answers)
            question_obj.times = expected_time  
            db.session.commit()

            all_questions = Questionsequence1.query.all()
            print("Sequence Questions in the database:")
            for question in all_questions:
              print(f"ID: {question.question_id}, Text: {question.question_name}")
              answers = Answer_sequence.query.filter_by(question_id=question.question_id).all()
              for answer in answers:
                print(f"    Answer ID: {answer.answer_id}, Text: {answer.answer_text}, Index: {answer.index}")

            redirect_url = url_for('edit_question_sequence', lesson_id=lesson_id)
            return jsonify({
                 'message': 'Sequence questions generated and saved to database successfully',
                 'questions': question_data,'redirect_url': redirect_url
            })
        elif question_type == "matching":
            response_text = generate_matching_questions(text,number_of_questions)
            if not response_text:
                return jsonify({'error': 'Failed to generate matching questions'}), 500

            questions_data = parse_matching_questions(response_text)
            if not questions_data:
                return jsonify({'error': 'Failed to parse matching questions'}), 500
            
            # Store Matching Questions in the Database
            for q_data in questions_data:
                question_obj = questionmatching1(
                    lesson_id=lesson_id,
                    question_text=q_data['question_text'],
                    answer_text=q_data['answer_text'],
                     question_text1=q_data['question_text1'],
                    answer_text1=q_data['answer_text1'],
                     question_text2=q_data['question_text2'],
                    answer_text2=q_data['answer_text2'],
                     question_text3=q_data['question_text3'],
                    answer_text3=q_data['answer_text3'],
                )
                db.session.add(question_obj)
                question_obj.times = calculate_expected_time_matching(question_obj, app)
            db.session.commit()

            # Print all matching questions from the database
            all_questions = questionmatching1.query.all()
            print("Matching Questions in the database:")
            for question in all_questions:
                print(f"ID: {question.question_id}, Question 1: {question.question_text} Answer 1 : {question.answer_text}, Question 2: {question.question_text1} Answer 2 : {question.answer_text1}, Question 3: {question.question_text2} Answer 3 : {question.answer_text2}, Question 4: {question.question_text3} Answer 4 : {question.answer_text3}")
            redirect_url = url_for('edit_matching', lesson_id=lesson_id)
            return jsonify({
                'message': 'Matching questions generated and saved to database successfully',
                'questions': questions_data,'redirect_url': redirect_url
            })
        else:
             return jsonify({'error': 'Invalid question type'}), 400


    except Exception as e:
        print(f"Error: {e}")
        db.session.rollback()
        return jsonify({'error': 'Internal Server Error'}), 500
@app.route('/get_questions', methods=['GET'])
def get_questions():
    all_questions = Questionfort_or_f.query.all()
    questions_list = []
    for question in all_questions:
        questions_list.append({
            'id': question.question_id,
            'lesson_id': question.lesson_id,
            'question_text': question.question_text,
            'correct_answer': question.correct_answer,
            'type': question.type
        })
    return jsonify({'questions': questions_list})

@app.route('/get_mcq_questions', methods=['GET'])
def get_mcq_questions():
    all_questions = Question_Multipleone_choice.query.all()
    questions_list = []
    for question in all_questions:
        answers = Answer_Multiple_Multiple_choice.query.filter_by(question_id=question.question_id).all()
        answers_list = [{'id': answer.answer_id, 'text': answer.answer_text, 'is_correct': answer.is_correct} for
                        answer in answers]
        questions_list.append({
            'id': question.question_id,
            'lesson_id': question.lesson_id,
            'question_text': question.question_name,
            'answers': answers_list
        })
    return jsonify({'questions': questions_list})

@app.route('/get_multiple_mcq_questions', methods=['GET'])
def get_multiple_mcq_questions():
    all_questions = QuestionMultipleMultiple_choice.query.all()
    questions_list = []
    for question in all_questions:
        answers = Answer_Multiple_Multiple_choice.query.filter_by(question_id=question.question_id).all()
        answers_list = [{'id': answer.answer_id, 'text': answer.answer_text, 'is_correct': answer.is_correct} for
                        answer in answers]
        questions_list.append({
            'id': question.question_id,
            'lesson_id': question.lesson_id,
            'question_text': question.question_name,
            'answers': answers_list
        })
    return jsonify({'questions': questions_list})

@app.route('/get_fill_blank_questions', methods=['GET'])
def get_fill_blank_questions():
    all_questions = Question_Fillin_blank4.query.all()
    questions_list = []
    for question in all_questions:
        questions_list.append({
            'id': question.question_id,
            'lesson_id': question.lesson_id,
            'question_text': question.question_text,
            'question_text1': question.question_text1,
            'question_text2': question.question_text2,
             'question_text3': question.question_text3,
             'answer_text': question.answer_text,
            'answer_text1': question.answer_text1,
            'answer_text2': question.answer_text2,
            'answer_text3': question.answer_text3,
        })
    return jsonify({'questions': questions_list})


@app.route('/get_essay_questions', methods=['GET'])
def get_essay_questions():
    all_questions = Article_question.query.all()
    questions_list = []
    for question in all_questions:
        questions_list.append({
            'id': question.question_id,
            'lesson_id': question.lesson_id,
            'question_text': question.question_text,
        })
    return jsonify({'questions': questions_list})
@app.route('/bank')
def bank():
    student_id = session.get('student_id')
    if not student_id:
        return redirect(url_for('access'))  # Redirect if student_id is missing

    # Fetch all the lessons IDs for the provided student_id
    lesson_employees = SubmittionLesson.query.filter_by().all()
    
    lessons_data = []
    for lesson_employee in lesson_employees:
         lesson = Lesson8.query.filter_by(id=lesson_employee.lesson_id).first()
         if lesson:
            lessons_data.append(lesson)

    return render_template('bank.html', lessons=lessons_data)
@app.route('/ai')
def ai():
    for m in genai.list_models():
        if 'generateContent' in m.supported_generation_methods:
            print(m.name)
    return render_template('ai.html')

@app.route('/employee_grades')
def employee_grades():
    employee_id = session.get('employee_id')  

    if not employee_id:
        return "Employee ID is required in the session. Please log in."

    # استعلام لاسترجاع جميع الامتحانات التي أنشأها هذا الموظف
    lessons = Lesson8.query.filter_by(employee_id=employee_id).all()

    exam_grades = {}  # قاموس لتجميع درجات الطلاب لكل امتحان

    for lesson in lessons:
        lesson_id = lesson.id

        # استعلام لاسترجاع جميع إجابات الطلاب في هذا الامتحان
        user_answers = UserAnswer4.query.filter_by(lesson_id=lesson_id).all()

        # تجميع درجات الطلاب في هذا الامتحان
        student_grades = {}
        for answer in user_answers:
            student_id = answer.user_id
            if student_id not in student_grades:
                student = SectorStudent8.query.get(student_id)
                username = student.username if student else "Unknown User"  # Get username of the student
                student_grades[student_id] = {'username': username, 'total_grade': 0, 'max_total_grade': 0}
            student_grades[student_id]['total_grade'] += answer.grade
            student_grades[student_id]['max_total_grade'] += answer.grade_real
        
        user_exam = UserExam1.query.filter_by(lesson_id=lesson_id).first()
        exams_taken = user_exam.exams_taken if user_exam else 0

        exam_grades[lesson_id] = {
            'exam_name': lesson.title,  # اسم الامتحان
            'student_grades': student_grades,  # درجات الطلاب في هذا الامتحان
            'exams_taken': exams_taken #عدد مرات الحل
        }

    return render_template('employee_grades.html', exam_grades=exam_grades, employee_id=employee_id)



if __name__ == '__main__':
    app.run(debug=True)
